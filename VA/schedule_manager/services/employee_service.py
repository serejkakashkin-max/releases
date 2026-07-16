from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook

from VA.schedule_manager.models.employee import Employee
from VA.schedule_manager.models.schedule_grid import ScheduleGrid
from VA.schedule_manager.parsers.excel_parser import parse_employees_from_excel
from VA.schedule_manager.parsers.monthly_workbook_parser import workbook_usage_index
from VA.schedule_manager.repositories.competency_repository import CompetencyRepository
from VA.schedule_manager.repositories.employee_repository import EmployeeRepository
from VA.schedule_manager.services.competency_service import COMPETENCY_SUPPORT, CompetencyService
from VA.schedule_manager.services.schedule_service import ScheduleService
from VA.schedule_manager.config import DATA_DIR, EMPLOYEE_SHEET_NAME


EMPLOYEE_STATUSES = {
    "active": "Активен",
    "long_leave": "Длительный отпуск",
    "dismissed": "Уволен",
}

EMPLOYEE_LOCATIONS = {
    "moscow": "Москва",
    "khabarovsk": "Хабаровск",
}


class EmployeeValidationError(Exception):
    pass


class EmployeeInUseError(Exception):
    pass


@dataclass(frozen=True)
class ScheduleUsage:
    sheet_name: str
    title: str
    assigned_days: int


class EmployeeService:
    def __init__(
        self,
        repository: EmployeeRepository,
        workbook_path: Optional[Path] = None,
        schedule_service: Optional[ScheduleService] = None,
        competency_service: Optional[CompetencyService] = None,
    ) -> None:
        self.repository = repository
        self.workbook_path = workbook_path
        self.schedule_service = schedule_service
        self.competency_service = competency_service or CompetencyService(CompetencyRepository(), repository)

    def list_employees(self) -> List[Employee]:
        employees = self.repository.load_all()
        if employees:
            return employees

        seeded = self._seed_from_workbook()
        self.repository.save_all(seeded)
        return seeded

    def active_count(self) -> int:
        return sum(1 for employee in self.list_employees() if employee.status == "active")

    def add_employee(
        self,
        name: str,
        email: str,
        phone: str,
        status: str,
        location: str = "moscow",
        competencies: Optional[List[str]] = None,
        overtime_ready: bool = True,
    ) -> None:
        employees = self.list_employees()
        employee = self._build_employee(name, email, phone, status, location, competencies, overtime_ready)
        if self._find(employees, employee.name) is not None:
            raise EmployeeValidationError("Сотрудник с таким ФИО уже есть.")
        employees.append(employee)
        self.repository.save_all(self._sorted(employees))

    def update_employee(
        self,
        original_name: str,
        name: str,
        email: str,
        phone: str,
        status: str,
        location: str = "moscow",
        competencies: Optional[List[str]] = None,
        overtime_ready: bool = True,
    ) -> None:
        employees = self.list_employees()
        employee = self._build_employee(name, email, phone, status, location, competencies, overtime_ready)
        found = False
        updated: List[Employee] = []
        for current in employees:
            if current.name == original_name:
                updated.append(employee)
                found = True
            else:
                if current.name == employee.name:
                    raise EmployeeValidationError("Сотрудник с таким ФИО уже есть.")
                updated.append(current)

        if not found:
            raise EmployeeValidationError("Сотрудник не найден.")
        self.repository.save_all(self._sorted(updated))

    def update_employee_fields(
        self,
        original_name: str,
        name: str,
        email: str,
        phone: str,
        status: str,
        location: str,
        competencies: Optional[List[str]] = None,
        overtime_ready: bool = True,
    ) -> None:
        self.update_employee(original_name, name, email, phone, status, location, competencies, overtime_ready)

    def delete_employee(self, name: str) -> None:
        usage = self.find_schedule_usage(name)
        if usage:
            raise EmployeeInUseError("Сотрудник найден в графиках. Выберите, что сделать дальше.")

        employees = [employee for employee in self.list_employees() if employee.name != name]
        self.repository.save_all(employees)

    def delete_employee_with_schedule_cleanup(self, name: str) -> Path:
        usage = self.find_schedule_usage(name)
        if not usage:
            self.delete_employee(name)
            if self.schedule_service is not None:
                return Path("schedule_data.json")
            if self.workbook_path is None:
                raise EmployeeValidationError("Файл графика не найден.")
            return self.workbook_path

        if self.schedule_service is not None:
            self._remove_employee_from_saved_schedules(name)
            employees = [employee for employee in self.list_employees() if employee.name != name]
            self.repository.save_all(employees)
            return Path("schedule_data.json")

        output_path = self._save_workbook_without_employee(name)
        employees = [employee for employee in self.list_employees() if employee.name != name]
        self.repository.save_all(employees)
        return output_path

    def change_status(self, name: str, status: str) -> None:
        employees = self.list_employees()
        updated = []
        found = False
        for employee in employees:
            if employee.name == name:
                updated.append(
                    Employee(
                        name=employee.name,
                        email=employee.email,
                        phone=employee.phone,
                        status=self._validate_status(status),
                        personnel_number=employee.personnel_number,
                        role=employee.role,
                        location=employee.location,
                        competencies=employee.competencies,
                        overtime_ready=employee.overtime_ready,
                    )
                )
                found = True
            else:
                updated.append(employee)
        if not found:
            raise EmployeeValidationError("Сотрудник не найден.")
        self.repository.save_all(updated)

    def is_used_in_schedules(self, name: str) -> bool:
        return bool(self.find_schedule_usage(name))

    def find_schedule_usage(self, name: str) -> List[ScheduleUsage]:
        if self.schedule_service is not None:
            return self._find_saved_schedule_usage(name)

        if self.workbook_path is None or not self.workbook_path.exists():
            return []

        normalized_name = self._normalize_name(name)
        usage: List[ScheduleUsage] = []
        for sheet_usage in workbook_usage_index(self.workbook_path):
            for employee_name, assigned_days in sheet_usage.employee_assignments:
                if self._normalize_name(employee_name) != normalized_name:
                    continue

                usage.append(
                    ScheduleUsage(
                        sheet_name=sheet_usage.sheet_name,
                        title=sheet_usage.label,
                        assigned_days=assigned_days,
                    )
                )
                break

        return usage

    def _find_saved_schedule_usage(self, name: str) -> List[ScheduleUsage]:
        snapshot = self.schedule_service.get_current()
        if snapshot is None:
            return []

        normalized_name = self._normalize_name(name)
        usage = []
        for option in snapshot.month_options():
            grid = snapshot.get_month_grid(option["sheet_name"])
            for row in grid.employees:
                if self._normalize_name(row.employee_name) != normalized_name:
                    continue
                assigned_days = sum(1 for code in row.assignments.values() if self._normalize_name(code))
                usage.append(ScheduleUsage(option["sheet_name"], option["label"], assigned_days))
                break
        return usage

    def _remove_employee_from_saved_schedules(self, name: str) -> None:
        if self.schedule_service is None:
            return
        snapshot = self.schedule_service.get_current()
        if snapshot is None:
            return

        normalized_name = self._normalize_name(name)
        for option in snapshot.month_options():
            grid = snapshot.get_month_grid(option["sheet_name"])
            updated_grid = ScheduleGrid(
                title=grid.title,
                year=grid.year,
                month=grid.month,
                days=grid.days,
                employees=[
                    row
                    for row in grid.employees
                    if self._normalize_name(row.employee_name) != normalized_name
                ],
            )
            self.schedule_service.save_month_grid(option["sheet_name"], updated_grid)

    def _save_workbook_without_employee(self, name: str) -> Path:
        if self.workbook_path is None or not self.workbook_path.exists():
            raise EmployeeValidationError("Файл графика не найден.")

        workbook = load_workbook(self.workbook_path)
        normalized_name = self._normalize_name(name)
        for worksheet in workbook.worksheets:
            rows_to_delete = []
            for row_index in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row_index, column=1).value
                if self._normalize_name(cell_value) == normalized_name:
                    rows_to_delete.append(row_index)

            for row_index in reversed(rows_to_delete):
                worksheet.delete_rows(row_index, 1)

        if EMPLOYEE_SHEET_NAME in workbook.sheetnames:
            worksheet = workbook[EMPLOYEE_SHEET_NAME]
            rows_to_delete = []
            for row_index in range(1, worksheet.max_row + 1):
                for column_index in range(1, worksheet.max_column + 1):
                    if self._normalize_name(worksheet.cell(row=row_index, column=column_index).value) == normalized_name:
                        rows_to_delete.append(row_index)
                        break

            for row_index in reversed(sorted(set(rows_to_delete))):
                worksheet.delete_rows(row_index, 1)

        output_dir = DATA_DIR / "edited_workbooks"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{self.workbook_path.stem}_без_{self._safe_slug(name)}.xlsx"
        workbook.save(output_path)
        return output_path

    def _seed_from_workbook(self) -> List[Employee]:
        if self.workbook_path is not None and self.workbook_path.exists():
            employees = parse_employees_from_excel(self.workbook_path)
            return self._sorted(
                [
                    Employee(
                        name=employee.name,
                        status="active",
                        role="employee",
                        location="moscow",
                        competencies=(COMPETENCY_SUPPORT,),
                        overtime_ready=True,
                    )
                    for employee in employees
                    if employee.name != "Список сотрудников"
                ]
            )
        return []

    def _build_employee(
        self,
        name: str,
        email: str,
        phone: str,
        status: str,
        location: str,
        competencies: Optional[List[str]],
        overtime_ready: bool = True,
    ) -> Employee:
        name = " ".join(name.strip().split())
        if not name:
            raise EmployeeValidationError("ФИО обязательно.")
        validated_competencies = self._validate_competencies(competencies or [COMPETENCY_SUPPORT])
        return Employee(
            name=name,
            email=email.strip(),
            phone=phone.strip(),
            status=self._validate_status(status),
            role="manager" if "manager" in validated_competencies else "employee",
            location=self._validate_location(location),
            competencies=validated_competencies,
            overtime_ready=bool(overtime_ready),
        )

    def _validate_status(self, status: str) -> str:
        if status not in EMPLOYEE_STATUSES:
            raise EmployeeValidationError("Неизвестный статус сотрудника.")
        return status

    def _validate_location(self, location: str) -> str:
        if location not in EMPLOYEE_LOCATIONS:
            raise EmployeeValidationError("Неизвестная локация сотрудника.")
        return location

    def _validate_competencies(self, competencies: List[str]) -> tuple:
        valid_codes = self.competency_service.valid_codes()
        result = tuple(dict.fromkeys(code for code in competencies if code))
        unknown = [code for code in result if code not in valid_codes]
        if unknown:
            raise EmployeeValidationError("Неизвестная компетенция сотрудника.")
        return result

    def _find(self, employees: List[Employee], name: str) -> Optional[Employee]:
        return next((employee for employee in employees if employee.name == name), None)

    def _sorted(self, employees: List[Employee]) -> List[Employee]:
        return sorted(employees, key=lambda employee: employee.name)

    def _normalize_name(self, value: object) -> str:
        if value is None:
            return ""
        return " ".join(str(value).strip().split())

    def _safe_slug(self, value: str) -> str:
        slug = "".join(char if char.isalnum() else "_" for char in value).strip("_")
        return slug or "employee"
