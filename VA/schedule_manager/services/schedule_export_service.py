from io import BytesIO
from re import sub
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from VA.schedule_manager.models.schedule_grid import ScheduleGrid
from VA.schedule_manager.models.shift import ShiftDefinition
from VA.schedule_manager.services.schedule_service import ScheduleService
from VA.schedule_manager.services.shift_service import ShiftService


class ScheduleExportError(Exception):
    pass


class ScheduleExportService:
    def __init__(self, schedule_service: ScheduleService, shift_service: ShiftService) -> None:
        self.schedule_service = schedule_service
        self.shift_service = shift_service

    def export_month(self, sheet_name: str) -> tuple[str, BytesIO]:
        snapshot = self.schedule_service.get_current()
        if snapshot is None:
            raise ScheduleExportError("Сначала загрузите или создайте график.")

        try:
            grid = snapshot.get_month_grid(sheet_name)
        except KeyError as exc:
            raise ScheduleExportError("График для экспорта не найден.") from exc

        workbook = Workbook()
        schedule_sheet = workbook.active
        schedule_sheet.title = self._safe_sheet_title(grid.title)

        self._build_schedule_sheet(schedule_sheet, grid)
        self._build_publication_sheet(workbook.create_sheet("К публикации"), grid)
        self._build_timesheet_sheet(workbook.create_sheet("Для внесения в табель"), grid)

        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        stream.seek(0)
        return f"График_{self._safe_filename(grid.title)}.xlsx", stream

    def _build_schedule_sheet(self, sheet, grid: ScheduleGrid) -> None:
        self._setup_grid_sheet(sheet, grid, include_hours=True, transform="raw")
        self._add_full_legend(sheet, grid)

    def _build_publication_sheet(self, sheet, grid: ScheduleGrid) -> None:
        self._setup_grid_sheet(sheet, grid, include_hours=False, transform="publication")
        self._add_full_legend(sheet, grid)

    def _build_timesheet_sheet(self, sheet, grid: ScheduleGrid) -> None:
        self._setup_grid_sheet(sheet, grid, include_hours=False, transform="timesheet")
        self._add_timesheet_legend(sheet, grid)

    def _setup_grid_sheet(self, sheet, grid: ScheduleGrid, include_hours: bool, transform: str) -> None:
        day_start_col = 4 if include_hours else 3
        sheet.freeze_panes = f"{get_column_letter(day_start_col)}3"
        sheet.sheet_view.showGridLines = False

        if grid.days:
            sheet["A1"] = grid.days[0].date
            sheet["A1"].number_format = "dd.mm.yyyy"
        sheet["A2"] = "Фамилия И. О."
        if include_hours:
            sheet["C1"] = "кол-во"
            sheet["C2"] = "часов"

        for offset, day in enumerate(grid.days):
            column = day_start_col + offset
            sheet.cell(row=1, column=column, value=day.weekday)
            sheet.cell(row=2, column=column, value=day.day)

        lookup = self.shift_service.lookup()
        for row_index, row in enumerate(grid.employees, start=3):
            sheet.cell(row=row_index, column=1, value=row.employee_name)
            if include_hours:
                sheet.cell(row=row_index, column=3, value=int(row.hours or 0))
            for offset, day in enumerate(grid.days):
                raw_code = row.assignments.get(day.day, "")
                sheet.cell(
                    row=row_index,
                    column=day_start_col + offset,
                    value=self._display_value(raw_code, transform, lookup),
                )

        self._style_grid(sheet, grid, day_start_col, include_hours)
        for row_index, row in enumerate(grid.employees, start=3):
            for offset, day in enumerate(grid.days):
                self._apply_shift_fill(
                    sheet.cell(row=row_index, column=day_start_col + offset),
                    row.assignments.get(day.day, ""),
                    lookup,
                )

    def _style_grid(self, sheet, grid: ScheduleGrid, day_start_col: int, include_hours: bool) -> None:
        thin = Side(style="thin", color="D0D5DD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        weekend_fill = PatternFill("solid", fgColor="E2F0D9")
        name_fill = PatternFill("solid", fgColor="F8FAFC")

        max_row = len(grid.employees) + 2
        max_col = day_start_col + len(grid.days) - 1
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 0.5
        if include_hours:
            sheet.column_dimensions["C"].width = 8
        for column in range(day_start_col, max_col + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 4.2

        for row in range(1, max_row + 1):
            sheet.row_dimensions[row].height = 18
            for column in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=column)
                cell.border = border
                cell.alignment = self._alignment(horizontal="center")
                cell.font = Font(name="Calibri", size=10)

        for column in range(1, max_col + 1):
            for row in (1, 2):
                sheet.cell(row=row, column=column).fill = header_fill
                sheet.cell(row=row, column=column).font = Font(name="Calibri", size=10, bold=True)

        for offset, day in enumerate(grid.days):
            column = day_start_col + offset
            if day.weekday in {"сб", "вс"}:
                for row in range(1, max_row + 1):
                    if row <= 2 or not sheet.cell(row=row, column=column).value:
                        sheet.cell(row=row, column=column).fill = weekend_fill

        for row in range(3, max_row + 1):
            sheet.cell(row=row, column=1).fill = name_fill
            sheet.cell(row=row, column=1).alignment = self._alignment(horizontal="left")
            sheet.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)
            if include_hours:
                sheet.cell(row=row, column=3).font = Font(name="Calibri", size=10, bold=True)

    def _add_full_legend(self, sheet, grid: ScheduleGrid) -> None:
        start_row = len(grid.employees) + 5
        sheet.cell(row=start_row, column=1, value="Легенда:")
        sheet.cell(row=start_row, column=1).font = Font(name="Calibri", size=10, bold=True)
        for index, shift in enumerate(self._legend_shifts(), start=start_row):
            sheet.merge_cells(start_row=index, start_column=3, end_row=index, end_column=13)
            sheet.merge_cells(start_row=index, start_column=14, end_row=index, end_column=32)
            sheet.cell(row=index, column=3, value=self._legend_time_label(shift))
            sheet.cell(row=index, column=14, value=self._legend_description(shift))
            self._style_legend_row(sheet, index, shift)

    def _add_timesheet_legend(self, sheet, grid: ScheduleGrid) -> None:
        start_row = len(grid.employees) + 5
        sheet.cell(row=start_row, column=1, value="Легенда:")
        sheet.cell(row=start_row, column=1).font = Font(name="Calibri", size=10, bold=True)
        for index, shift in enumerate(self._legend_shifts(), start=start_row):
            sheet.merge_cells(start_row=index, start_column=3, end_row=index, end_column=13)
            sheet.merge_cells(start_row=index, start_column=14, end_row=index, end_column=32)
            sheet.cell(row=index, column=3, value=self._timesheet_value(shift.code, self.shift_service.lookup()))
            sheet.cell(row=index, column=14, value=self._legend_description(shift))
            self._style_legend_row(sheet, index, shift)

    def _style_legend_row(self, sheet, row: int, shift: ShiftDefinition) -> None:
        thin = Side(style="thin", color="D0D5DD")
        for column in range(3, 33):
            cell = sheet.cell(row=row, column=column)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = self._alignment(horizontal="left")
            cell.font = Font(name="Calibri", size=10)
        self._apply_shift_fill(sheet.cell(row=row, column=3), shift.code, self.shift_service.lookup())
        self._apply_shift_fill(sheet.cell(row=row, column=14), shift.code, self.shift_service.lookup())

    def _display_value(self, code: object, transform: str, lookup: dict) -> object:
        normalized = self._normalize_code(code)
        if not normalized:
            return None
        if transform == "timesheet":
            return self._timesheet_value(normalized, lookup)
        shift = lookup.get(normalized) or lookup.get(normalized.lower())
        if shift is None:
            return normalized
        return self._numeric_if_possible(shift.display_code)

    def _timesheet_value(self, code: object, lookup: dict) -> Optional[str]:
        normalized = self._normalize_code(code)
        if not normalized:
            return None
        shift = lookup.get(normalized) or lookup.get(normalized.lower())
        canonical = self._normalize_code(shift.code if shift else normalized).lower()
        display = self._normalize_code(shift.display_code if shift else normalized).lower()
        if canonical in {"праздник"} or display == "п":
            return "П"
        if canonical in {"отпуск"} or display == "о":
            return "О"
        if display == "хд":
            return "07-16"
        if display == "хр":
            return "09-18"
        if shift and shift.start_time and shift.end_time:
            return f"{self._hour_label(shift.start_time)}-{self._hour_label(shift.end_time)}"
        return self._normalize_code(shift.display_code if shift else normalized)

    def _legend_shifts(self) -> list[ShiftDefinition]:
        shifts = self.shift_service.list_shifts()
        preferred = ["ХД", "ХР", "8", "ДД", "ДР", "ВД", "ВР", "отпуск", "ВХ", "Праздник"]
        by_code = {self._normalize_code(shift.code).lower(): shift for shift in shifts}
        by_display = {self._normalize_code(shift.display_code).lower(): shift for shift in shifts}
        result = []
        for code in preferred:
            shift = by_code.get(code.lower()) or by_display.get(code.lower())
            if shift and shift not in result:
                result.append(shift)
        for shift in shifts:
            if shift not in result:
                result.append(shift)
        return result

    def _legend_time_label(self, shift: ShiftDefinition) -> str:
        display = self._normalize_code(shift.display_code)
        if display == "ВХ":
            return "По факту выхода"
        if display == "О":
            return "отпуск"
        if display == "П":
            return "Праздник"
        if shift.start_time and shift.end_time:
            return f"{self._hour_label(shift.start_time)}-{self._hour_label(shift.end_time)} ({shift.timezone})"
        return shift.name

    def _legend_description(self, shift: ShiftDefinition) -> str:
        return f"{shift.display_code} - {shift.name}"

    def _apply_shift_fill(self, cell, code: object, lookup: dict) -> None:
        normalized = self._normalize_code(code)
        shift = lookup.get(normalized) or lookup.get(normalized.lower())
        if shift is None or not shift.color:
            return
        color = self._excel_color(shift.color)
        if not color:
            return
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(name="Calibri", size=10, bold=True, color=self._font_color_for_fill(color))

    def _alignment(self, horizontal: str) -> Alignment:
        return Alignment(
            horizontal=horizontal,
            vertical="center",
            wrap_text=False,
            shrink_to_fit=True,
        )

    def _safe_sheet_title(self, value: str) -> str:
        title = sub(r"[:\\/?*\\[\\]]", " ", value).strip() or "График"
        return title[:31]

    def _safe_filename(self, value: str) -> str:
        return sub(r"[^0-9A-Za-zА-Яа-яЁё_-]+", "_", value).strip("_") or "schedule"

    def _normalize_code(self, value: object) -> str:
        return " ".join(str(value or "").strip().split())

    def _numeric_if_possible(self, value: object) -> object:
        text = self._normalize_code(value)
        if text.isdigit():
            return int(text)
        return text

    def _hour_label(self, value: str) -> str:
        return self._normalize_code(value).replace(":00", "")

    def _excel_color(self, value: str) -> str:
        color = self._normalize_code(value).lstrip("#").upper()
        if len(color) == 6 and all(char in "0123456789ABCDEF" for char in color):
            return color
        return ""

    def _font_color_for_fill(self, color: str) -> str:
        red = int(color[0:2], 16)
        green = int(color[2:4], 16)
        blue = int(color[4:6], 16)
        brightness = (red * 299 + green * 587 + blue * 114) / 1000
        return "1F2933" if brightness > 150 else "FFFFFF"
