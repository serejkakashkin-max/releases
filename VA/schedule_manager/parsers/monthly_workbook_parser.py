from dataclasses import dataclass
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import load_workbook

from VA.schedule_manager.models.schedule_grid import ScheduleDay, ScheduleGrid, ScheduleRow
from VA.schedule_manager.parsers.schedule_csv_parser import MONTHS_RU


TIME_ROW_CODES = {"8", "ДД", "ДР", "ВД", "ВР", "ХД", "ХР"}


@dataclass(frozen=True)
class MonthSheetOption:
    year: int
    month: int
    month_name: str
    sheet_name: str
    label: str


@dataclass(frozen=True)
class MonthSheetUsageIndex:
    sheet_name: str
    label: str
    employee_assignments: Tuple[Tuple[str, int], ...]
    shift_counts: Tuple[Tuple[str, int], ...]


class MonthlyWorkbookParseError(Exception):
    pass


def list_month_sheets(path: Path) -> List[MonthSheetOption]:
    if not path.exists():
        return []
    return list(_list_month_sheets(str(path), _file_signature(path)))


def parse_month_sheet(path: Path, sheet_name: str) -> ScheduleGrid:
    if not path.exists():
        raise MonthlyWorkbookParseError(f"Файл не найден: {path}")
    return _parse_month_sheet(str(path), _file_signature(path), sheet_name)


def parse_all_month_sheets(path: Path) -> List[Tuple[MonthSheetOption, ScheduleGrid]]:
    if not path.exists():
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        result: List[Tuple[MonthSheetOption, ScheduleGrid]] = []
        for worksheet in workbook.worksheets:
            option = _option_from_sheet(worksheet.title, worksheet["A1"].value)
            if option is None:
                continue
            try:
                result.append((option, _parse_month_worksheet(worksheet, option)))
            except MonthlyWorkbookParseError:
                continue
        return sorted(result, key=lambda item: (item[0].year, item[0].month), reverse=True)
    finally:
        workbook.close()


def workbook_usage_index(path: Path) -> List[MonthSheetUsageIndex]:
    if not path.exists():
        return []
    return list(_workbook_usage_index(str(path), _file_signature(path)))


@lru_cache(maxsize=8)
def _list_month_sheets(path_text: str, signature: Tuple[int, int]) -> Tuple[MonthSheetOption, ...]:
    path = Path(path_text)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        options: List[MonthSheetOption] = []
        for worksheet in workbook.worksheets:
            option = _option_from_sheet(worksheet.title, worksheet["A1"].value)
            if option is not None:
                options.append(option)
    finally:
        workbook.close()

    return tuple(sorted(options, key=lambda item: (item.year, item.month), reverse=True))


@lru_cache(maxsize=8)
def _workbook_usage_index(path_text: str, signature: Tuple[int, int]) -> Tuple[MonthSheetUsageIndex, ...]:
    path = Path(path_text)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        result: List[MonthSheetUsageIndex] = []
        for worksheet in workbook.worksheets:
            option = _option_from_sheet(worksheet.title, worksheet["A1"].value)
            if option is None:
                continue

            employee_assignments = {}
            shift_counts = {}
            for row in worksheet.iter_rows(min_row=3, values_only=True):
                name = _clean_cell(row[0] if row else "")
                if not name or name in TIME_ROW_CODES:
                    break

                assigned_days = 0
                for value in row[3:]:
                    code = _clean_cell(value)
                    if not code:
                        continue
                    assigned_days += 1
                    shift_counts[code] = shift_counts.get(code, 0) + 1

                employee_assignments[_normalize_name(name)] = assigned_days

            result.append(
                MonthSheetUsageIndex(
                    sheet_name=option.sheet_name,
                    label=option.label,
                    employee_assignments=tuple(employee_assignments.items()),
                    shift_counts=tuple(shift_counts.items()),
                )
            )
    finally:
        workbook.close()

    return tuple(result)


@lru_cache(maxsize=24)
def _parse_month_sheet(path_text: str, signature: Tuple[int, int], sheet_name: str) -> ScheduleGrid:
    path = Path(path_text)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise MonthlyWorkbookParseError(f"Лист не найден: {sheet_name}")

        worksheet = workbook[sheet_name]
        title_value = worksheet["A1"].value
        option = _option_from_sheet(sheet_name, title_value)
        if option is None:
            raise MonthlyWorkbookParseError(f"Лист '{sheet_name}' не похож на месячный график.")
        grid = _parse_month_worksheet(worksheet, option)
    finally:
        workbook.close()

    return grid


def _parse_month_worksheet(worksheet, option: MonthSheetOption) -> ScheduleGrid:
    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) < 3:
        raise MonthlyWorkbookParseError(f"На листе '{option.sheet_name}' слишком мало строк.")

    weekdays = [_clean_cell(value) for value in rows[0][3:]]
    day_numbers = [_parse_day(value) for value in rows[1][3:]]
    days = [
        ScheduleDay(day=day, weekday=weekday, date=date(option.year, option.month, day))
        for day, weekday in zip(day_numbers, weekdays)
        if day is not None
    ]

    employees: List[ScheduleRow] = []
    for row in rows[2:]:
        name = _clean_cell(row[0] if len(row) > 0 else "")
        if not name:
            break
        if name in TIME_ROW_CODES:
            break

        hours = _parse_int(row[2] if len(row) > 2 else None)
        values = row[3 : 3 + len(days)]
        assignments = {
            schedule_day.day: _clean_cell(values[index] if index < len(values) else "")
            for index, schedule_day in enumerate(days)
        }
        employees.append(ScheduleRow(employee_name=_normalize_name(name), hours=hours, assignments=assignments))

    if not employees:
        raise MonthlyWorkbookParseError(f"На листе '{option.sheet_name}' не найдены сотрудники.")

    return ScheduleGrid(
        title=option.label,
        year=option.year,
        month=option.month,
        days=days,
        employees=employees,
    )


def clear_monthly_workbook_cache() -> None:
    _list_month_sheets.cache_clear()
    _parse_month_sheet.cache_clear()
    _workbook_usage_index.cache_clear()


def _file_signature(path: Path) -> Tuple[int, int]:
    stat = path.stat()
    return stat.st_mtime_ns, stat.st_size


def _option_from_sheet(sheet_name: str, first_cell: object) -> Optional[MonthSheetOption]:
    if isinstance(first_cell, datetime):
        year = first_cell.year
        month = first_cell.month
        month_name = _month_name(month)
        return MonthSheetOption(year, month, month_name, sheet_name, f"{month_name} {year}")

    normalized = sheet_name.replace("_", " ")
    parts = normalized.split()
    if len(parts) != 2:
        return None

    month_name, year_text = parts
    month = MONTHS_RU.get(month_name)
    if month is None:
        return None
    try:
        year = int(year_text)
    except ValueError:
        return None

    return MonthSheetOption(year, month, month_name, sheet_name, f"{month_name} {year}")


def _month_name(month: int) -> str:
    for name, number in MONTHS_RU.items():
        if number == month:
            return name
    return str(month)


def _parse_day(value: object) -> Optional[int]:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _parse_int(value: object) -> Optional[int]:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def _normalize_name(name: str) -> str:
    return " ".join(name.split())
