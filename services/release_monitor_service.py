import logging
import copy
import gzip
import hashlib
import json
import html
import os
import re
import threading
import time
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from uuid import uuid4

import requests
from openpyxl import load_workbook

from config import DASHBOARD_CACHE_TTL, OPLOT_VALUES, TOKENS
from services.jira_service import get_jira_domain_and_token
from services.jira_oplot_issue_service import create_oplot_release_issue
from services.release_artifact_service import (
    classify_artifact_entry,
    extract_artifact_ke_id,
    extract_artifact_url,
    extract_distribution_version,
    flatten_artifact_candidates,
    is_ai_agent_release_context,
    select_distribution_artifact,
)
from services.template_constructor_service import (
    build_runtime_template_catalog,
    is_ai_agents_template_category,
    select_template_by_summary,
)


FINAL_RELEASE_STATUS = "\u0423\u0441\u0442\u0430\u043d\u043e\u0432\u043b\u0435\u043d \u043d\u0430 \u041f\u0420\u041e\u041c"
CANCELLED_RELEASE_STATUS = "\u041e\u0442\u043c\u0435\u043d\u0435\u043d\u043e"
READY_FOR_PROM_STATUS = "\u0413\u043e\u0442\u043e\u0432 \u043a \u0443\u0441\u0442\u0430\u043d\u043e\u0432\u043a\u0435 \u043d\u0430 \u041f\u0420\u041e\u041c"
FINAL_RELEASE_STATUSES = (
    FINAL_RELEASE_STATUS,
    CANCELLED_RELEASE_STATUS,
)
PRE_FINAL_RELEASE_STATUSES = (
    "\u0423\u0441\u0442\u0430\u043d\u043e\u0432\u043a\u0430 \u043d\u0430 \u041f\u0420\u041e\u041c",
    READY_FOR_PROM_STATUS,
)
APPROVED_ROV_STATUSES = (
    "\u0423\u0442\u0432\u0435\u0440\u0436\u0434\u0435\u043d",
    "\u0423\u0442\u0432\u0435\u0440\u0436\u0434\u0451\u043d",
    "\u0423\u0442\u0432\u0435\u0440\u0436\u0434\u0435\u043d\u043e",
    "\u0423\u0442\u0432\u0435\u0440\u0436\u0434\u0451\u043d\u043e",
)
RELEASE_PREFIXES = ("EMRM", "SMECLM", "SMECSC", "HELPERAI", "AIGAS", "DRMMMB")
RELEASE_ISSUE_TYPE = "Release 2.0"
ROV_ISSUE_TYPE = "Introduction Order"
QUICK_REFRESH_DAYS = 9
AUTO_FULL_REFRESH_ENABLED = False
AUTO_FULL_REFRESH_HOUR = 6
AUTO_REFRESH_CHECK_INTERVAL = 300
AUTO_INCREMENTAL_REFRESH_ENABLED = True
AUTO_INCREMENTAL_REFRESH_INTERVAL_SECONDS = 180
AUTO_INCREMENTAL_REFRESH_LOOKBACK_MINUTES = 60
AUTO_INCREMENTAL_REFRESH_CHECK_INTERVAL = 30
RELEASE_OPERATIONAL_DAY_START_HOUR = 3
AI_AGENTS_SYSTEM_NAME = "AI-\u0410\u0433\u0435\u043d\u0442\u044b"
RELEASE_MONITOR_TRACE_ENABLED = False
RELIABLE_FULL_REFRESH_MODE = "reliable_full"
RELIABLE_SEARCH_RETRY_STATUS_CODES = {429, 500, 502, 503, 504}
FULL_SEARCH_MAX_ATTEMPTS = int(os.getenv("RELEASE_MONITOR_FULL_SEARCH_MAX_ATTEMPTS", "3"))
RELIABLE_SEARCH_MAX_ATTEMPTS = int(os.getenv("RELEASE_MONITOR_RELIABLE_SEARCH_MAX_ATTEMPTS", "6"))
RELIABLE_SEARCH_RETRY_DELAY_SECONDS = float(os.getenv("RELEASE_MONITOR_RELIABLE_SEARCH_RETRY_DELAY_SECONDS", "3"))
FULL_REFRESH_DEADLINE_SECONDS = int(os.getenv("RELEASE_MONITOR_FULL_REFRESH_DEADLINE_SECONDS", "300"))
RELIABLE_FULL_REFRESH_DEADLINE_SECONDS = int(
    os.getenv("RELEASE_MONITOR_RELIABLE_FULL_REFRESH_DEADLINE_SECONDS", "900")
)
AUTO_INCREMENTAL_SEARCH_MAX_ATTEMPTS = int(os.getenv("RELEASE_MONITOR_AUTO_INCREMENTAL_SEARCH_MAX_ATTEMPTS", "3"))
AUTO_INCREMENTAL_SEARCH_RETRY_DELAY_SECONDS = float(os.getenv("RELEASE_MONITOR_AUTO_INCREMENTAL_SEARCH_RETRY_DELAY_SECONDS", "2"))
SNAPSHOT_DIR = Path(__file__).resolve().parent.parent / "cache"
SNAPSHOT_FILE = SNAPSHOT_DIR / "release_monitor_snapshot.json"
LAST_GOOD_SNAPSHOT_FILE = SNAPSHOT_DIR / "release_monitor_last_good.json"
CANDIDATE_SNAPSHOT_FILE = SNAPSHOT_DIR / "release_monitor_candidate.json"
SNAPSHOT_ARCHIVES_DIR = SNAPSHOT_DIR / "release_monitor_archives"
MAX_GOOD_SNAPSHOT_ARCHIVES = 5
MANUAL_RELEASES_FILE = SNAPSHOT_DIR / "release_monitor_manual_releases.json"
MANUAL_OVERRIDES_FILE = SNAPSHOT_DIR / "release_monitor_manual_overrides.json"
REVIEWERS_FILE = SNAPSHOT_DIR / "release_monitor_reviewers.json"
ORDER_FILE = SNAPSHOT_DIR / "release_monitor_order.json"
DUTY_SCHEDULE_FILE = SNAPSHOT_DIR / "release_monitor_duty_schedule.json"
DATE_OVERRIDES_FILE = SNAPSHOT_DIR / "release_monitor_date_overrides.json"
ZNI_FILE = SNAPSHOT_DIR / "release_monitor_zni.json"
WORK_MARKS_FILE = SNAPSHOT_DIR / "release_monitor_work_marks.json"
ATTEMPTS_FILE = SNAPSHOT_DIR / "release_monitor_attempts.json"
REVISION_FILE = SNAPSHOT_DIR / "release_monitor_revision.txt"
CONFLUENCE_DELTA_BASE = "https://confluence.delta.sbrf.ru"
JIRA_DELTA_BASE = "https://jira.delta.sbrf.ru"
RELEASE_TYPE_VALUES = {"release", "hotfix", "reroll", "technical"}
RELEASE_TYPE_ALIASES = {
    "": "",
    "release": "release",
    "rel": "release",
    "\u0440\u0435\u043b\u0438\u0437": "release",
    "hotfix": "hotfix",
    "fix": "hotfix",
    "\u0445\u043e\u0442\u0444\u0438\u043a\u0441": "hotfix",
    "reroll": "reroll",
    "re-roll": "reroll",
    "\u043f\u0435\u0440\u0435\u0440\u0430\u0441\u043a\u0430\u0442\u043a\u0430": "reroll",
    "emergency": "technical",
    "\u0430\u0432\u0430\u0440\u0438\u0439\u043d\u044b\u0439": "technical",
    "\u0430\u0432\u0430\u0440\u0438\u0439\u043d\u043e\u0435": "technical",
    "technical": "technical",
    "tech": "technical",
    "\u0442\u0435\u0445\u043d\u0438\u0447\u0435\u0441\u043a\u0438\u0439": "technical",
    "custom": "release",
    "\u0434\u0440\u0443\u0433\u043e\u0435": "release",
}
RELEASE_TYPE_ROW_LABELS = {
    "release": "(\u0420\u0435\u043b\u0438\u0437)",
    "hotfix": "(\u0425\u043e\u0442\u0444\u0438\u043a\u0441)",
    "reroll": "(\u041f\u0435\u0440\u0435\u0440\u0430\u0441\u043a\u0430\u0442\u043a\u0430)",
    "technical": "(\u0422\u0435\u0445\u043d\u0438\u0447\u0435\u0441\u043a\u0438\u0439)",
}
MANUAL_OVERRIDE_SCALAR_FIELDS = (
    "release_type",
    "release_summary",
    "release_key",
    "rov_key",
    "release_url",
    "rov_url",
    "release_status",
    "rov_status",
    "ke_id",
    "ke",
    "release_version",
    "release_dist_url",
    "system_name",
    "zni_key",
    "zni_url",
)
MANUAL_OVERRIDE_DICT_FIELDS = ("display_fields", "doc_fields", "confluence_fields")
MANUAL_OVERRIDE_FIELDS = MANUAL_OVERRIDE_SCALAR_FIELDS + MANUAL_OVERRIDE_DICT_FIELDS
BASE_OVERRIDE_FIELDS = MANUAL_OVERRIDE_FIELDS
MANUAL_RELEASE_SCALAR_FIELDS = MANUAL_OVERRIDE_SCALAR_FIELDS + ("deployment_start", "deployment_end", "year")
MANUAL_RELEASE_FIELDS = MANUAL_RELEASE_SCALAR_FIELDS + MANUAL_OVERRIDE_DICT_FIELDS
RELEASE_VERSION_PATTERN = re.compile(r"[DP]-\d+(?:\.\d+){2}(?:[.-][A-Za-z0-9_]+)+")
ARTIFACT_URL_PATTERN = re.compile(r"(?:https?://)?[A-Za-z0-9.-]+\.[A-Za-z]{2,}/[^\s\"'<>)]+")
DUTY_HARD_EXCLUDED_STATUSES = {"ДД", "ВД", "ВР", "ХД", "ХР"}
DUTY_RESERVE_STATUSES = {"ДР"}
DUTY_ABSENCE_KEYWORDS = ("отпуск", "отгул", "больн")

MONTH_NAME_MAP = {
    "\u044f\u043d\u0432\u0430\u0440": 1,
    "\u0444\u0435\u0432\u0440\u0430\u043b": 2,
    "\u043c\u0430\u0440\u0442": 3,
    "\u0430\u043f\u0440\u0435\u043b": 4,
    "\u043c\u0430\u0439": 5,
    "\u0438\u044e\u043d": 6,
    "\u0438\u044e\u043b": 7,
    "\u0430\u0432\u0433\u0443\u0441\u0442": 8,
    "\u0441\u0435\u043d\u0442\u044f\u0431": 9,
    "\u043e\u043a\u0442\u044f\u0431": 10,
    "\u043d\u043e\u044f\u0431": 11,
    "\u0434\u0435\u043a\u0430\u0431": 12,
}

FIELD_FALLBACKS = {
    "planned_prom_start": None,
    "planned_prom_end": "customfield_18606",
    "system_info": "customfield_22400",
    "ke_object": "customfield_18300",
    "release_distributive": "customfield_21710",
    "delta_release_distributive": "customfield_27011",
    "rov_start": None,
    "rov_end": None,
}

FIELD_ALIASES = {
    "planned_prom_start": (
        "\u043d\u0430\u0447\u0430\u043b\u043e \u0432\u043d\u0435\u0434\u0440\u0435\u043d\u0438\u044f",
        "\u0434\u0430\u0442\u0430 \u043d\u0430\u0447\u0430\u043b\u0430 \u0432\u043d\u0435\u0434\u0440\u0435\u043d\u0438\u044f",
        "\u043d\u0430\u0447\u0430\u043b\u043e \u0432\u043d\u0435\u0434\u0440\u0435\u043d\u0438\u044f \u043f\u043b\u0430\u043d",
    ),
    "planned_prom_end": (
        "\u0434\u0430\u0442\u0430 \u0437\u0430\u0432\u0435\u0440\u0448\u0435\u043d\u0438\u044f \u0443\u0441\u0442\u0430\u043d\u043e\u0432\u043a\u0438 \u0432 \u043f\u0440\u043e\u043c",
        "\u0434\u0430\u0442\u0430 \u0443\u0441\u0442\u0430\u043d\u043e\u0432\u043a\u0438 \u0432 \u043f\u0440\u043e\u043c",
    ),
    "system_info": (
        "\u0438\u0442-\u0443\u0441\u043b\u0443\u0433\u0430",
    ),
    "ke_object": (
        "\u043a\u044d",
    ),
    "release_distributive": (
        "\u043a\u044d \u0434\u0438\u0441\u0442\u0440\u0438\u0431\u0443\u0442\u0438\u0432\u0430",
        "\u043a\u044d \u0434\u0438\u0441\u0442\u0440\u0438\u0431\u0443\u0442\u0438\u0432\u043e\u0432",
    ),
    "rov_start": (
        "\u0434\u0430\u0442\u0430/\u0432\u0440\u0435\u043c\u044f \u043d\u0430\u0447\u0430\u043b\u0430 \u0440\u0430\u0431\u043e\u0442 \u043f\u043e \u0432\u043d\u0435\u0434\u0440\u0435\u043d\u0438\u044e",
        "\u0434\u0430\u0442\u0430 \u0432\u0440\u0435\u043c\u044f \u043d\u0430\u0447\u0430\u043b\u0430 \u0440\u0430\u0431\u043e\u0442 \u043f\u043e \u0432\u043d\u0435\u0434\u0440\u0435\u043d\u0438\u044e",
        "\u043d\u0430\u0447\u0430\u043b\u043e \u0440\u0430\u0431\u043e\u0442 \u043f\u043e \u0432\u043d\u0435\u0434\u0440\u0435\u043d\u0438\u044e",
    ),
    "rov_end": (
        "\u0434\u0430\u0442\u0430/\u0432\u0440\u0435\u043c\u044f \u043e\u043a\u043e\u043d\u0447\u0430\u043d\u0438\u044f \u0440\u0430\u0431\u043e\u0442 \u043f\u043e \u0432\u043d\u0435\u0434\u0440\u0435\u043d\u0438\u044e",
        "\u0434\u0430\u0442\u0430 \u0432\u0440\u0435\u043c\u044f \u043e\u043a\u043e\u043d\u0447\u0430\u043d\u0438\u044f \u0440\u0430\u0431\u043e\u0442 \u043f\u043e \u0432\u043d\u0435\u0434\u0440\u0435\u043d\u0438\u044e",
        "\u043e\u043a\u043e\u043d\u0447\u0430\u043d\u0438\u0435 \u0440\u0430\u0431\u043e\u0442 \u043f\u043e \u0432\u043d\u0435\u0434\u0440\u0435\u043d\u0438\u044e",
    ),
}

_cache_lock = threading.RLock()
_cached_data = None
_last_cache_update = None
_manual_order_cache = None
_manual_overrides_legacy_migration_checked = False
_manual_override_unknown_fields_warned = set()
_field_map_cache = {}
_refresh_thread = None
_auto_incremental_thread = None
_last_auto_incremental_refresh_at = None
_auto_incremental_status = {
    "state": "idle",
    "last_started_at": None,
    "last_finished_at": None,
    "last_changed": False,
    "last_error": None,
    "jira_checks_total": 0,
    "jira_checks_success": 0,
    "jira_checks_failed": 0,
}
_scheduler_thread = None
_scheduler_started = False
_snapshot_recovery_checked = False
_snapshot_requires_display_migration = False
_refresh_status = {
    "state": "idle",
    "message": "",
    "started_at": None,
    "finished_at": None,
    "error": None,
    "mode": None,
    "trigger": None,
}


class ReleaseMonitorSourceError(RuntimeError):
    def __init__(self, message, candidate=None):
        super().__init__(message)
        self.candidate = candidate


class ReleaseMonitorCandidateRejected(RuntimeError):
    def __init__(self, message, candidate=None, validation_report=None):
        super().__init__(message)
        self.candidate = candidate
        self.validation_report = validation_report or {}


class ReleaseMonitorAssignmentConflict(RuntimeError):
    def __init__(self, message, assignment=None):
        super().__init__(message)
        self.assignment = assignment or {}


def _rm_trace(tag, event, *, started_at=None, level=logging.INFO, **details):
    if not RELEASE_MONITOR_TRACE_ENABLED:
        return
    fields = []
    if started_at is not None:
        fields.append(f"elapsed_ms={int((time.monotonic() - started_at) * 1000)}")
    for key, value in sorted(details.items()):
        fields.append(f"{key}={value}")
    suffix = f" {' '.join(fields)}" if fields else ""
    logging.log(level, "[%s] event=%s%s", tag, event, suffix)


def _build_empty_release_monitor_payload():
    current_year = datetime.now().year
    previous_year = current_year - 1
    return {
        "items": [],
        "manual_overrides": {},
        "summary": {
            "total": 0,
            "non_final": 0,
            "overdue": 0,
            "today": 0,
            "pre_final": 0,
            "final": 0,
            "cancelled": 0,
            "by_status": {},
            "by_year": {
                str(current_year): {
                    "total": 0,
                    "non_final": 0,
                    "overdue": 0,
                    "today": 0,
                    "pre_final": 0,
                    "final": 0,
                    "cancelled": 0,
                },
                str(previous_year): {
                    "total": 0,
                    "non_final": 0,
                    "overdue": 0,
                    "today": 0,
                    "pre_final": 0,
                    "final": 0,
                    "cancelled": 0,
                },
            },
        },
        "meta": {
            "final_status": FINAL_RELEASE_STATUS,
            "cancelled_status": CANCELLED_RELEASE_STATUS,
            "final_statuses": list(FINAL_RELEASE_STATUSES),
            "pre_final_statuses": list(PRE_FINAL_RELEASE_STATUSES),
            "prefixes": list(RELEASE_PREFIXES),
            "years": [current_year, previous_year],
            "current_year": current_year,
            "last_updated": None,
            "last_full_sync": None,
            "last_quick_sync": None,
            "last_auto_incremental_sync": None,
            "last_confluence_sync": None,
            "last_duty_schedule_upload": None,
            "last_sync_mode": None,
            "quick_refresh_days": QUICK_REFRESH_DAYS,
            "auto_full_refresh_enabled": AUTO_FULL_REFRESH_ENABLED,
            "auto_full_refresh_hour": AUTO_FULL_REFRESH_HOUR,
            "auto_incremental_refresh_enabled": AUTO_INCREMENTAL_REFRESH_ENABLED,
            "auto_incremental_refresh_interval_seconds": AUTO_INCREMENTAL_REFRESH_INTERVAL_SECONDS,
            "auto_incremental_refresh_lookback_minutes": AUTO_INCREMENTAL_REFRESH_LOOKBACK_MINUTES,
            "data_revision": _read_data_revision(),
            "is_cached": False,
        },
    }


def _ensure_snapshot_dir():
    os.makedirs(SNAPSHOT_DIR, exist_ok=True)


def _prepare_atomic_bytes(file_path, payload):
    file_path = Path(file_path)
    file_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = file_path.with_name(f".{file_path.name}.{uuid4().hex}.tmp")
    try:
        with temp_path.open("wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        return temp_path
    except Exception:
        if temp_path.exists():
            temp_path.unlink()
        raise


def _atomic_write_bytes(file_path, payload):
    file_path = Path(file_path)
    temp_path = _prepare_atomic_bytes(file_path, payload)
    try:
        os.replace(temp_path, file_path)
    finally:
        if temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


def _atomic_write_text(file_path, text):
    _atomic_write_bytes(file_path, str(text).encode("utf-8"))


def _atomic_write_json(file_path, payload):
    serialized = json.dumps(payload, ensure_ascii=False, indent=2)
    json.loads(serialized)
    _atomic_write_text(file_path, serialized)


def _read_json_file(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))


def _load_json_payload(file_path, *, log_errors=True):
    file_path = Path(file_path)
    if not file_path.exists():
        return None
    try:
        payload = _read_json_file(file_path)
        return payload if isinstance(payload, dict) else None
    except Exception as exc:
        if log_errors:
            logging.warning("Release monitor: failed to load JSON file %s: %s", file_path, exc)
        return None


def _is_valid_snapshot_payload(payload):
    return isinstance(payload, dict) and isinstance(payload.get("items"), list)


def _parse_snapshot_timestamp(value):
    value = str(value or "").strip()
    if not value:
        return 0
    for fmt in (
        "%Y-%m-%dT%H:%M:%S.%fZ",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S",
        "%d.%m.%Y %H:%M:%S",
    ):
        try:
            return int(datetime.strptime(value, fmt).timestamp() * 1000)
        except ValueError:
            continue
    return 0


def _snapshot_revision_value(payload, file_path=None):
    meta = dict((payload or {}).get("meta") or {})
    raw_revision = str(meta.get("accepted_revision") or "").strip()
    try:
        return int(raw_revision)
    except (TypeError, ValueError):
        pass

    timestamp_value = (
        _parse_snapshot_timestamp(meta.get("accepted_at"))
        or _parse_snapshot_timestamp(meta.get("last_updated"))
    )
    if timestamp_value:
        return timestamp_value
    try:
        return int(Path(file_path).stat().st_mtime * 1000) if file_path else 0
    except OSError:
        return 0


def _snapshot_payload_hash(payload):
    comparable = copy.deepcopy(payload or {})
    meta = dict(comparable.get("meta") or {}) if isinstance(comparable, dict) else {}
    for field_name in (
        "accepted_revision",
        "accepted_at",
        "data_revision",
        "last_updated",
        "last_full_sync",
        "last_sync_mode",
    ):
        meta.pop(field_name, None)
    if isinstance(comparable, dict):
        comparable["meta"] = meta
    raw = json.dumps(comparable, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _load_archive_snapshot(file_path):
    try:
        with gzip.open(file_path, "rt", encoding="utf-8") as handle:
            payload = json.load(handle)
        return payload if _is_valid_snapshot_payload(payload) else None
    except Exception as exc:
        logging.warning("Release monitor: failed to load archive snapshot %s: %s", file_path, exc)
        return None


def _new_accepted_revision(*payloads):
    highest = int(time.time() * 1000)
    for payload in payloads:
        highest = max(highest, _snapshot_revision_value(payload))
    return str(highest + 1)


def _utc_now_iso():
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _prepare_accepted_snapshot(payload, accepted_revision=None, accepted_at=None):
    prepared = copy.deepcopy(payload or {})
    prepared.setdefault("items", [])
    meta = dict(prepared.get("meta") or {})
    revision = str(accepted_revision or meta.get("accepted_revision") or "").strip()
    if not revision:
        revision = _new_accepted_revision(prepared)
    meta["accepted_revision"] = revision
    meta["accepted_at"] = accepted_at or meta.get("accepted_at") or _utc_now_iso()
    meta["data_revision"] = revision
    meta["is_cached"] = True
    prepared["meta"] = meta
    return prepared


def _recover_snapshot_storage(*, force=False):
    global _snapshot_recovery_checked, _snapshot_requires_display_migration

    trace_started_at = time.monotonic()
    if _snapshot_recovery_checked and not force:
        _rm_trace(
            "RM_RECOVERY",
            "skip_already_checked",
            started_at=trace_started_at,
            level=logging.DEBUG,
        )
        return

    _rm_trace("RM_RECOVERY", "start", force=force)
    _ensure_snapshot_dir()
    active = _load_json_payload(SNAPSHOT_FILE)
    last_good = _load_json_payload(LAST_GOOD_SNAPSHOT_FILE)
    active_valid = _is_valid_snapshot_payload(active)
    last_good_valid = _is_valid_snapshot_payload(last_good)
    selected = None
    repair_pair = False

    if active_valid and last_good_valid:
        active_revision = _snapshot_revision_value(active, SNAPSHOT_FILE)
        last_good_revision = _snapshot_revision_value(last_good, LAST_GOOD_SNAPSHOT_FILE)
        active_accepted = str((active.get("meta") or {}).get("accepted_revision") or "").strip()
        last_good_accepted = str((last_good.get("meta") or {}).get("accepted_revision") or "").strip()
        if active_accepted and active_accepted == last_good_accepted:
            selected = active
        elif active_revision != last_good_revision or active_accepted != last_good_accepted:
            selected = active if active_revision >= last_good_revision else last_good
            repair_pair = True
            logging.warning(
                "Release monitor: active/last-good revisions differ (%s/%s); using newer snapshot",
                active_accepted or active_revision,
                last_good_accepted or last_good_revision,
            )
        else:
            selected = active
    elif active_valid:
        selected = active
        repair_pair = True
        logging.warning("Release monitor: last-good snapshot is missing or invalid; restoring it from active")
    elif last_good_valid:
        selected = last_good
        repair_pair = True
        logging.warning("Release monitor: active snapshot is missing or invalid; restoring it from last-good")
    else:
        archive_paths = sorted(
            SNAPSHOT_ARCHIVES_DIR.glob("snapshot_*.json.gz") if SNAPSHOT_ARCHIVES_DIR.exists() else [],
            key=lambda path: (path.stat().st_mtime, path.name),
            reverse=True,
        )
        for archive_path in archive_paths:
            selected = _load_archive_snapshot(archive_path)
            if selected is not None:
                repair_pair = True
                logging.warning(
                    "Release monitor: active and last-good are invalid; recovered from %s",
                    archive_path.name,
                )
                break

    if selected is not None:
        accepted_revision = str((selected.get("meta") or {}).get("accepted_revision") or "").strip()
        if not accepted_revision:
            _snapshot_requires_display_migration = True
            logging.warning("Release monitor: migrating legacy snapshot to accepted_revision model")
            if not active_valid or selected is not active:
                _atomic_write_json(SNAPSHOT_FILE, selected)
            _snapshot_recovery_checked = True
            _rm_trace(
                "RM_RECOVERY",
                "legacy_migration_required",
                started_at=trace_started_at,
                active_valid=active_valid,
                items=_count_payload_items(selected),
                last_good_valid=last_good_valid,
            )
            return

        if repair_pair:
            _atomic_write_json(LAST_GOOD_SNAPSHOT_FILE, selected)
            _atomic_write_json(SNAPSHOT_FILE, selected)
            _atomic_write_text(REVISION_FILE, (selected.get("meta") or {}).get("data_revision") or "")

    _snapshot_recovery_checked = True
    _rm_trace(
        "RM_RECOVERY",
        "complete",
        started_at=trace_started_at,
        active_valid=active_valid,
        items=_count_payload_items(selected),
        last_good_valid=last_good_valid,
        repaired=repair_pair,
        revision=(selected.get("meta") or {}).get("accepted_revision") if selected else "",
    )


def _load_state_json(file_path, default=None):
    file_path = Path(file_path)
    if not file_path.exists():
        return default

    try:
        return _read_json_file(file_path)
    except Exception as exc:
        logging.warning("Release monitor: failed to load state file %s: %s", file_path, exc)
    return default


def _write_state_json(file_path, payload):
    file_path = Path(file_path)
    _ensure_snapshot_dir()

    text = json.dumps(payload, ensure_ascii=False, indent=2)
    json.loads(text)
    file_path.write_text(text, encoding="utf-8")


def _mark_release_monitor_state_changed():
    revision = _touch_release_monitor_revision()
    if isinstance(_cached_data, dict):
        _cached_data.setdefault("meta", {})["data_revision"] = revision
    return revision


def _count_payload_items(payload):
    if not isinstance(payload, dict):
        return 0
    items = payload.get("items", [])
    return len(items) if isinstance(items, list) else 0


def _new_data_revision():
    return str(int(time.time() * 1000))


def _read_data_revision():
    try:
        if REVISION_FILE.exists():
            revision = REVISION_FILE.read_text(encoding="utf-8").strip()
            if revision:
                return revision
    except Exception as exc:
        logging.warning("Release monitor: failed to read revision: %s", exc)
    return ""


def _touch_release_monitor_revision():
    revision = _new_data_revision()
    try:
        _ensure_snapshot_dir()
        _atomic_write_text(REVISION_FILE, revision)
    except Exception as exc:
        logging.warning("Release monitor: failed to save revision: %s", exc)
    return revision


def _append_revision_meta(meta):
    meta = dict(meta or {})
    revision = str(meta.get("data_revision") or "").strip() or _read_data_revision()
    if not revision:
        revision = str(_get_snapshot_mtime() or "")
    meta["data_revision"] = revision
    return meta


def _get_auto_incremental_status():
    status = dict(_auto_incremental_status)
    status["enabled"] = AUTO_INCREMENTAL_REFRESH_ENABLED
    status["interval_seconds"] = AUTO_INCREMENTAL_REFRESH_INTERVAL_SECONDS
    status["lookback_minutes"] = AUTO_INCREMENTAL_REFRESH_LOOKBACK_MINUTES
    status["running"] = bool(_auto_incremental_thread and _auto_incremental_thread.is_alive())
    return status


def _update_auto_incremental_status(**updates):
    with _cache_lock:
        _auto_incremental_status.update(updates)


def _append_auto_incremental_meta(meta):
    meta = dict(meta or {})
    meta["auto_full_refresh_enabled"] = AUTO_FULL_REFRESH_ENABLED
    meta["auto_full_refresh_hour"] = AUTO_FULL_REFRESH_HOUR
    meta["auto_incremental_status"] = _get_auto_incremental_status()
    return meta


def is_release_monitor_refreshing():
    with _cache_lock:
        return bool(_refresh_thread and _refresh_thread.is_alive()) or _refresh_status.get("state") == "refreshing"


def ensure_release_monitor_not_refreshing():
    if is_release_monitor_refreshing():
        raise RuntimeError("Идет обновление релизов. Дождитесь завершения и обновите страницу.")


def _strip_manual_release_overrides_for_snapshot(payload):
    payload_to_save = dict(payload or {})
    payload_to_save.pop("manual_overrides", None)

    items_to_save = []
    for item in payload_to_save.get("items") or []:
        if not isinstance(item, dict):
            continue
        row_key = str(item.get("row_key") or "").strip()
        if item.get("manual_release") or item.get("source") == "manual" or row_key.startswith("manual::"):
            continue
        clean_item = dict(item)

        for target_field in BASE_OVERRIDE_FIELDS:
            base_field = _base_field_name(target_field)
            if base_field in clean_item:
                clean_item[target_field] = clean_item.get(base_field) or ""

        if isinstance(clean_item.get("base_release_name_lines"), list):
            clean_item["release_name_lines"] = list(clean_item.get("base_release_name_lines") or [])

        manual_fields = {
            "has_manual_release_override",
            "manual_clear_zni",
            "manual_overridden_fields",
        }
        manual_fields.update(_manual_field_name(field) for field in MANUAL_OVERRIDE_FIELDS)
        for manual_field in manual_fields:
            clean_item.pop(manual_field, None)

        items_to_save.append(clean_item)

    payload_to_save["items"] = items_to_save
    return payload_to_save


def _save_snapshot_to_disk(payload, bump_revision=True):
    try:
        _ensure_snapshot_dir()
        current_disk_payload = _load_snapshot_from_disk()
        incoming_items = _count_payload_items(payload)
        existing_items = _count_payload_items(current_disk_payload)

        if incoming_items == 0 and existing_items > 0:
            logging.warning(
                "Release monitor: skipped overwriting non-empty snapshot with empty payload (existing_items=%s)",
                existing_items,
            )
            return

        if isinstance(payload, dict):
            payload.setdefault("meta", {})
            if bump_revision:
                payload["meta"]["data_revision"] = _touch_release_monitor_revision()
            else:
                payload["meta"] = _append_revision_meta(payload.get("meta", {}))

        _atomic_write_json(SNAPSHOT_FILE, payload)
    except Exception as exc:
        logging.warning("Release monitor: failed to save snapshot to disk: %s", exc)


def _load_snapshot_from_disk():
    _recover_snapshot_storage()
    payload = _load_json_payload(SNAPSHOT_FILE)
    return payload if _is_valid_snapshot_payload(payload) else None


def _save_candidate_diagnostic(candidate, validation_report=None, *, state="candidate", error=""):
    diagnostic = copy.deepcopy(candidate or {})
    diagnostic.setdefault("items", [])
    diagnostic["candidate_state"] = state
    diagnostic["saved_at"] = _utc_now_iso()
    diagnostic["validation_report"] = copy.deepcopy(validation_report or {})
    if error:
        diagnostic["error"] = str(error)
    try:
        _atomic_write_json(CANDIDATE_SNAPSHOT_FILE, diagnostic)
    except Exception as exc:
        logging.warning("Release monitor: failed to save candidate diagnostic: %s", exc)


def _archive_timestamp_for_snapshot(payload):
    meta = dict((payload or {}).get("meta") or {})
    timestamp_ms = (
        _parse_snapshot_timestamp(meta.get("accepted_at"))
        or _parse_snapshot_timestamp(meta.get("last_updated"))
        or int(time.time() * 1000)
    )
    return datetime.fromtimestamp(timestamp_ms / 1000).strftime("%Y%m%d_%H%M%S")


def _latest_archive_hash():
    if not SNAPSHOT_ARCHIVES_DIR.exists():
        return ""
    archive_paths = sorted(
        SNAPSHOT_ARCHIVES_DIR.glob("snapshot_*.json.gz"),
        key=lambda path: (path.stat().st_mtime, path.name),
        reverse=True,
    )
    for archive_path in archive_paths:
        payload = _load_archive_snapshot(archive_path)
        if payload is not None:
            return _snapshot_payload_hash(payload)
    return ""


def _archive_previous_good_snapshot(payload):
    if not _is_valid_snapshot_payload(payload):
        return None
    if _snapshot_payload_hash(payload) == _latest_archive_hash():
        logging.info("Release monitor: previous good snapshot already matches latest archive")
        return None

    SNAPSHOT_ARCHIVES_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = _archive_timestamp_for_snapshot(payload)
    archive_path = SNAPSHOT_ARCHIVES_DIR / f"snapshot_{timestamp}.json.gz"
    suffix = 2
    while archive_path.exists():
        archive_path = SNAPSHOT_ARCHIVES_DIR / f"snapshot_{timestamp}_{suffix}.json.gz"
        suffix += 1

    raw = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    compressed = gzip.compress(raw, compresslevel=6, mtime=0)
    _atomic_write_bytes(archive_path, compressed)
    logging.info(
        "Release monitor: archived previous good snapshot %s, items=%s",
        archive_path.name,
        _count_payload_items(payload),
    )
    return archive_path


def _rotate_snapshot_archives():
    if not SNAPSHOT_ARCHIVES_DIR.exists():
        return
    archives = sorted(
        SNAPSHOT_ARCHIVES_DIR.glob("snapshot_*.json.gz"),
        key=lambda path: (path.stat().st_mtime, path.name),
        reverse=True,
    )
    for archive_path in archives[MAX_GOOD_SNAPSHOT_ARCHIVES:]:
        try:
            archive_path.unlink()
            logging.info("Release monitor: removed old snapshot archive %s", archive_path.name)
        except OSError as exc:
            logging.warning("Release monitor: failed to remove archive %s: %s", archive_path, exc)


def _commit_accepted_snapshot(payload, *, mode):
    global _snapshot_recovery_checked

    trace_started_at = time.monotonic()
    _rm_trace(
        "RM_COMMIT",
        "start",
        items=_count_payload_items(payload),
        mode=mode,
    )
    if not _is_valid_snapshot_payload(payload):
        raise ValueError("Accepted release monitor snapshot has invalid payload")

    _recover_snapshot_storage()
    previous_active = _load_json_payload(SNAPSHOT_FILE)
    previous_last_good = _load_json_payload(LAST_GOOD_SNAPSHOT_FILE)
    previous_good = (
        previous_active
        if _is_valid_snapshot_payload(previous_active)
        else previous_last_good
        if _is_valid_snapshot_payload(previous_last_good)
        else None
    )
    accepted_revision = _new_accepted_revision(previous_active, previous_last_good, payload)
    accepted_at = _utc_now_iso()
    prepared = _prepare_accepted_snapshot(
        payload,
        accepted_revision=accepted_revision,
        accepted_at=accepted_at,
    )

    created_archive = None
    if mode == RELIABLE_FULL_REFRESH_MODE and previous_good is not None:
        created_archive = _archive_previous_good_snapshot(previous_good)

    serialized = json.dumps(prepared, ensure_ascii=False, indent=2).encode("utf-8")
    last_good_temp = None
    active_temp = None
    try:
        last_good_temp = _prepare_atomic_bytes(LAST_GOOD_SNAPSHOT_FILE, serialized)
        active_temp = _prepare_atomic_bytes(SNAPSHOT_FILE, serialized)
        os.replace(last_good_temp, LAST_GOOD_SNAPSHOT_FILE)
        os.replace(active_temp, SNAPSHOT_FILE)
        _atomic_write_text(REVISION_FILE, accepted_revision)
    except Exception:
        if created_archive and created_archive.exists():
            try:
                created_archive.unlink()
            except OSError:
                logging.warning(
                    "Release monitor: failed to remove archive after unsuccessful commit: %s",
                    created_archive,
                )
        _snapshot_recovery_checked = False
        try:
            _recover_snapshot_storage(force=True)
        except Exception:
            logging.exception("Release monitor: immediate snapshot reconciliation failed")
        raise
    finally:
        for temp_path in (last_good_temp, active_temp):
            if temp_path and temp_path.exists():
                try:
                    temp_path.unlink()
                except OSError:
                    pass

    if mode == RELIABLE_FULL_REFRESH_MODE:
        _rotate_snapshot_archives()

    _snapshot_recovery_checked = True
    _rm_trace(
        "RM_COMMIT",
        "complete",
        started_at=trace_started_at,
        archived=bool(created_archive),
        items=_count_payload_items(prepared),
        mode=mode,
        revision=accepted_revision,
    )
    return prepared


def _get_snapshot_mtime():
    try:
        if SNAPSHOT_FILE.exists():
            return SNAPSHOT_FILE.stat().st_mtime
    except Exception:
        logging.exception("Release monitor: failed to read snapshot mtime")
    return None


def _reload_snapshot_from_disk_if_newer():
    global _cached_data, _last_cache_update

    disk_mtime = _get_snapshot_mtime()
    if disk_mtime is None:
        return

    if _cached_data is not None and _last_cache_update is not None and disk_mtime <= _last_cache_update:
        return

    disk_payload = _load_snapshot_from_disk()
    if disk_payload is not None:
        _cached_data = _hydrate_release_monitor_payload(disk_payload)
        _last_cache_update = disk_mtime
        _migrate_cached_display_snapshot_pair_if_needed(disk_payload)


def _normalize_reviewer_assignments_payload(payload):
    normalized = {}
    if not isinstance(payload, dict):
        return normalized

    for key, value in payload.items():
        release_key = str(key or "").strip()
        if not release_key:
            continue
        if isinstance(value, dict):
            raw_responsibles = value.get("responsibles", [])
            if not isinstance(raw_responsibles, list):
                raw_responsibles = [raw_responsibles] if raw_responsibles else []
            normalized[release_key] = {
                "reviewer": str(value.get("reviewer", "") or "").strip(),
                "reviewer_source": str(value.get("reviewer_source") or "").strip(),
                "reviewer_date": str(value.get("reviewer_date", "") or "").strip(),
                "zni_reviewer": str(value.get("zni_reviewer", "") or "").strip(),
                "checker": str(value.get("checker", "") or "").strip(),
                "responsibles": [
                    str(item or "").strip()
                    for item in raw_responsibles
                    if str(item or "").strip()
                ],
            }
        elif value:
            normalized[release_key] = {
                "reviewer": str(value).strip(),
                "reviewer_source": "manual",
                "reviewer_date": "",
                "zni_reviewer": "",
                "checker": "",
                "responsibles": [],
            }
    return normalized


def _load_reviewer_assignments():
    payload = _load_state_json(REVIEWERS_FILE, default={})
    return _normalize_reviewer_assignments_payload(payload)


def _save_reviewer_assignments(assignments):
    try:
        normalized = _normalize_reviewer_assignments_payload(assignments)
        _write_state_json(REVIEWERS_FILE, normalized)
    except Exception as exc:
        logging.warning("Release monitor: failed to save reviewer assignments: %s", exc)
        raise


def _normalize_zni_assignments(payload):
    normalized = {}
    if not isinstance(payload, dict):
        return normalized

    for key, value in payload.items():
        row_key = str(key or "").strip()
        if not row_key:
            continue
        if isinstance(value, dict):
            zni_key = str(value.get("key") or value.get("zni_key") or "").strip()
            if zni_key:
                normalized[row_key] = {
                    "key": zni_key,
                    "url": str(value.get("url") or "").strip(),
                    "summary": str(value.get("summary") or "").strip(),
                    "created_at": str(value.get("created_at") or "").strip(),
                    "assignee": str(value.get("assignee") or "").strip(),
                    "reporter": str(value.get("reporter") or "").strip(),
                }
        elif value:
            normalized[row_key] = {
                "key": str(value).strip(),
                "url": "",
                "summary": "",
                "created_at": "",
                "assignee": "",
                "reporter": "",
            }
    return normalized


def _normalize_rollout_note_flags(payload):
    normalized = {}
    if not isinstance(payload, dict):
        return normalized

    for key, value in payload.items():
        row_key = str(key or "").strip()
        if not row_key or not isinstance(value, dict):
            continue
        raw_level = str(value.get("rollout_notes_level") or value.get("level") or "").strip().lower()
        if not raw_level and bool(value.get("has_rollout_notes")):
            raw_level = "warning"
        if raw_level not in {"success", "warning", "danger", "none"}:
            continue
        normalized[row_key] = {
            "has_rollout_notes": raw_level != "none",
            "rollout_notes_level": raw_level,
            "updated_at": str(value.get("updated_at") or "").strip(),
        }
    return normalized


def _normalize_zni_payload_meta(payload):
    return dict(payload) if isinstance(payload, dict) else {}


def _load_zni_payload():
    payload = _load_state_json(ZNI_FILE, default=None)
    if payload is None:
        return {"issues": {}, "flags": {}, "meta": {}}

    if not isinstance(payload, dict):
        return {
            "issues": {},
            "flags": {},
            "meta": {},
        }

    if "issues" in payload or "flags" in payload:
        return {
            "issues": _normalize_zni_assignments(payload.get("issues") or {}),
            "flags": _normalize_rollout_note_flags(payload.get("flags") or {}),
            "meta": _normalize_zni_payload_meta(payload.get("meta") or {}),
        }

    return {
        "issues": _normalize_zni_assignments(payload),
        "flags": {},
        "meta": {},
    }


def _save_zni_payload(payload):
    try:
        normalized_payload = {
            "issues": _normalize_zni_assignments((payload or {}).get("issues") or {}),
            "flags": _normalize_rollout_note_flags((payload or {}).get("flags") or {}),
            "meta": _normalize_zni_payload_meta((payload or {}).get("meta") or {}),
        }
        _write_state_json(ZNI_FILE, normalized_payload)
    except Exception as exc:
        logging.warning("Release monitor: failed to save ZNI payload: %s", exc)
        raise


def _normalize_work_marks(payload):
    normalized = {}
    if not isinstance(payload, dict):
        return normalized

    for row_key, value in payload.items():
        normalized_row_key = str(row_key or "").strip()
        if not normalized_row_key:
            continue

        if isinstance(value, dict):
            mark = str(value.get("mark") or "").strip()
            updated_at = str(value.get("updated_at") or "").strip()
        else:
            mark = str(value or "").strip()
            updated_at = ""

        if mark:
            normalized[normalized_row_key] = {
                "mark": mark,
                "updated_at": updated_at,
            }
    return normalized


def _load_work_marks():
    return _normalize_work_marks(_load_state_json(WORK_MARKS_FILE, default={}) or {})


def _save_work_marks(marks):
    try:
        _write_state_json(WORK_MARKS_FILE, _normalize_work_marks(marks))
    except Exception as exc:
        logging.warning("Release monitor: failed to save work marks payload: %s", exc)
        raise


def _apply_work_marks(items):
    marks = _load_work_marks()
    for item in items or []:
        if not isinstance(item, dict):
            continue
        row_key = _get_assignment_key_for_item(item)
        mark_payload = marks.get(row_key) or {}
        mark = str(mark_payload.get("mark") or "").strip()
        item["work_mark"] = mark
        item["work_mark_updated_at"] = str(mark_payload.get("updated_at") or "").strip() if mark else ""
    return items


def _clear_release_work_mark(row_key):
    row_key = str(row_key or "").strip()
    if not row_key:
        return False

    marks = _load_work_marks()
    if row_key not in marks:
        return False

    marks.pop(row_key, None)
    _save_work_marks(marks)
    return True


def _load_manual_overrides_payload():
    global _manual_overrides_legacy_migration_checked

    payload = _load_state_json(MANUAL_OVERRIDES_FILE, default=None)
    if isinstance(payload, dict):
        _manual_overrides_legacy_migration_checked = True
        return _normalize_manual_release_overrides(payload)

    if _manual_overrides_legacy_migration_checked:
        return {}

    _manual_overrides_legacy_migration_checked = True
    legacy_payload = _load_snapshot_from_disk() or {}
    legacy_overrides = _normalize_manual_release_overrides(legacy_payload.get("manual_overrides") or {})
    return legacy_overrides


def _save_manual_overrides_payload(overrides):
    try:
        _write_state_json(MANUAL_OVERRIDES_FILE, _normalize_manual_release_overrides(overrides))
    except Exception as exc:
        logging.warning("Release monitor: failed to save manual overrides payload: %s", exc)
        raise


def normalize_release_type(value, default=""):
    raw_value = str(value or "").strip().lower()
    normalized = RELEASE_TYPE_ALIASES.get(raw_value, raw_value)
    if normalized in RELEASE_TYPE_VALUES:
        return normalized
    return default


def _release_row_label_from_type(release_type, fallback=""):
    normalized_type = normalize_release_type(release_type, default="release")
    return RELEASE_TYPE_ROW_LABELS.get(normalized_type) or fallback or RELEASE_TYPE_ROW_LABELS["release"]


def _sync_release_type_fields(item):
    if not isinstance(item, dict):
        return item

    release_type = normalize_release_type(
        item.get("release_type"),
        default=derive_release_type_from_jira(item),
    ) or "release"
    item["release_type"] = release_type
    item["row_label"] = _release_row_label_from_type(release_type, item.get("row_label"))
    item["is_reroll"] = release_type == "reroll"
    item["is_hotfix"] = release_type == "hotfix"
    item["release_name_lines"] = _build_release_name_lines(
        item.get("release_summary"),
        _get_release_name_ke_line(item),
        row_label=str(item.get("row_label") or RELEASE_TYPE_ROW_LABELS["release"]),
    )
    return item


def derive_release_type_from_jira(item):
    if not isinstance(item, dict):
        return "release"

    existing_type = normalize_release_type(item.get("release_type"))
    if existing_type:
        return existing_type
    if item.get("is_reroll"):
        return "reroll"

    row_label = _normalize_text(item.get("row_label"))
    summary = _normalize_text(item.get("release_summary"))
    release_version = str(item.get("release_version") or "").strip().upper()
    if "перераскат" in row_label or "перераскат" in summary:
        return "reroll"
    if "хотфикс" in row_label or "hotfix" in summary or release_version.startswith("P-"):
        return "hotfix"
    if "авар" in row_label or "авар" in summary or "emergency" in summary:
        return "technical"
    if "техн" in row_label or "техн" in summary or "technical" in summary:
        return "technical"
    return "release"


def _base_field_name(field_name):
    return f"base_{field_name}"


def _manual_field_name(field_name):
    return f"manual_{field_name}"


def _normalize_manual_dict_field(value):
    if not isinstance(value, dict):
        return {}
    normalized = {}
    for key, field_value in value.items():
        field_key = str(key or "").strip()
        if not field_key:
            continue
        if isinstance(field_value, (dict, list)):
            normalized[field_key] = field_value
        else:
            normalized[field_key] = str(field_value or "").strip()
    return normalized


def _normalize_manual_scalar_field(field_name, value):
    raw_value = str(value or "").strip()
    if not raw_value:
        return ""
    if field_name == "release_type":
        return normalize_release_type(raw_value)
    if field_name in {"release_dist_url"}:
        return _normalize_artifact_url(raw_value)
    if field_name == "ke":
        return _format_ke_id(raw_value) if re.fullmatch(r"(?:CI)?\d+", raw_value, re.IGNORECASE) else raw_value
    return raw_value


def _warn_unknown_manual_override_fields(fields):
    unknown_fields = sorted(set(fields or []) - set(MANUAL_OVERRIDE_FIELDS) - {
        "clear_zni",
        "updated_at",
        "updated_by",
        *(_base_field_name(field) for field in BASE_OVERRIDE_FIELDS),
    })
    for field_name in unknown_fields:
        if field_name in _manual_override_unknown_fields_warned:
            continue
        _manual_override_unknown_fields_warned.add(field_name)
        logging.warning("Release monitor: ignored unknown manual override field: %s", field_name)


def _validate_manual_release_payload(value):
    if not isinstance(value, dict):
        return ["manual release payload must be an object"]

    errors = []
    release_key = str(value.get("release_key") or "").strip()
    release_summary = str(value.get("release_summary") or "").strip()
    if not release_key and not release_summary:
        errors.append("release_key is required")

    start_dt = _parse_release_monitor_date(value.get("deployment_start") or value.get("deployment_start_iso"))
    end_dt = _parse_release_monitor_date(value.get("deployment_end") or value.get("deployment_end_iso"))
    if start_dt and end_dt and start_dt > end_dt:
        errors.append("deployment_start must be before deployment_end")

    if value.get("row_key") and not str(value.get("row_key")).startswith("manual::"):
        errors.append("manual row_key must start with manual::")
    if value.get("release_type") and not normalize_release_type(value.get("release_type")):
        errors.append("release_type is invalid")

    return errors


def _build_jira_issue_url_from_key(issue_key, fallback_url=""):
    issue_key = str(issue_key or "").strip().upper()
    fallback_url = str(fallback_url or "").strip()
    if fallback_url:
        return fallback_url
    if not issue_key:
        return ""
    try:
        domain, _ = get_jira_domain_and_token(issue_key)
    except Exception:
        domain = JIRA_DELTA_BASE
    return f"{str(domain or JIRA_DELTA_BASE).rstrip('/')}/browse/{issue_key}"


def _normalize_manual_release(row_key, value):
    if not isinstance(value, dict):
        return {}

    row_key = str(value.get("row_key") or row_key or "").strip()
    if not row_key.startswith("manual::"):
        return {}

    validation_errors = _validate_manual_release_payload({**value, "row_key": row_key})
    if validation_errors:
        logging.warning(
            "Release monitor: ignored invalid manual release %s: %s",
            row_key,
            "; ".join(validation_errors),
        )
        return {}

    start_dt = _parse_release_monitor_date(value.get("deployment_start_iso") or value.get("deployment_start"))
    end_dt = _parse_release_monitor_date(value.get("deployment_end_iso") or value.get("deployment_end")) or start_dt
    sort_dt = start_dt or end_dt
    year = int(value.get("year") or (sort_dt.year if sort_dt else datetime.now().year))
    release_type = normalize_release_type(value.get("release_type"), default="release")
    row_label = _release_row_label_from_type(release_type)
    release_key = str(value.get("release_key") or "").strip()
    release_summary = str(value.get("release_summary") or "").strip() or release_key
    system_name = str(value.get("system_name") or "").strip()
    ke_id = str(value.get("ke_id") or "").strip()
    ke_line = f"{system_name}({ke_id})" if system_name and ke_id else (system_name or ke_id)
    created_at = str(value.get("created_at") or "").strip()
    updated_at = str(value.get("updated_at") or "").strip()

    release_url = _build_jira_issue_url_from_key(release_key, value.get("release_url"))
    rov_key = str(value.get("rov_key") or "").strip()
    rov_url = _build_jira_issue_url_from_key(rov_key, value.get("rov_url"))

    manual_item = {
        "row_key": row_key,
        "manual_release": True,
        "source": "manual",
        "year": year,
        "release_number": "",
        "release_type": release_type,
        "base_release_type": release_type,
        "release_key": release_key,
        "release_url": release_url,
        "release_status": str(value.get("release_status") or "").strip(),
        "release_summary": release_summary,
        "release_name_lines": _build_release_name_lines(release_summary, ke_line, row_label=row_label),
        "base_release_summary": release_summary,
        "base_release_name_lines": _build_release_name_lines(release_summary, ke_line, row_label=row_label),
        "row_label": row_label,
        "rov_key": rov_key,
        "rov_url": rov_url,
        "rov_status": str(value.get("rov_status") or "").strip(),
        "has_rov": bool(rov_key),
        "ke_id": ke_id,
        "ke_name": system_name,
        "ke": _normalize_manual_scalar_field("ke", value.get("ke")),
        "release_version": str(value.get("release_version") or "").strip(),
        "release_dist_url": _normalize_artifact_url(str(value.get("release_dist_url") or "").strip()),
        "system_name": system_name,
        "zni_key": str(value.get("zni_key") or "").strip(),
        "zni_url": str(value.get("zni_url") or "").strip(),
        "has_rollout_notes": False,
        "rollout_notes_level": "",
        "deployment_start": _format_release_monitor_date(start_dt) if start_dt else "",
        "deployment_start_iso": start_dt.isoformat() if start_dt else "",
        "deployment_end": _format_release_monitor_date(end_dt) if end_dt else "",
        "deployment_end_iso": end_dt.isoformat() if end_dt else "",
        "source_deployment_start": _format_release_monitor_date(start_dt) if start_dt else "",
        "source_deployment_start_iso": start_dt.isoformat() if start_dt else "",
        "source_deployment_end": _format_release_monitor_date(end_dt) if end_dt else "",
        "source_deployment_end_iso": end_dt.isoformat() if end_dt else "",
        "psi_owner": "",
        "psi_responsibles": [],
        "psi_checker": "",
        "row_state": "planned",
        "is_final": False,
        "is_cancelled": False,
        "is_non_final": True,
        "is_pre_final": False,
        "is_ready_for_prom": False,
        "is_overdue": False,
        "is_today": False,
        "days_overdue": 0,
        "waits_for_rov": False,
        "is_unnumbered": False,
        "is_natural_unnumbered": False,
        "is_force_unnumbered": False,
        "sort_date": sort_dt.isoformat() if sort_dt else "",
        "created_sort_date": created_at,
        "created": created_at,
        "updated_at": updated_at,
        "updated_by": str(value.get("updated_by") or "").strip(),
    }

    for field_name in MANUAL_OVERRIDE_DICT_FIELDS:
        normalized_value = _normalize_manual_dict_field(value.get(field_name))
        if normalized_value:
            manual_item[field_name] = normalized_value

    for field_name in BASE_OVERRIDE_FIELDS:
        base_field = _base_field_name(field_name)
        if base_field not in manual_item:
            manual_item[base_field] = manual_item.get(field_name, {})

    return manual_item


def _normalize_manual_releases_payload(payload):
    if payload is None or not isinstance(payload, dict):
        return {"items": {}, "meta": {}}

    raw_items = payload.get("items") if isinstance(payload.get("items"), dict) else payload
    normalized_items = {}
    for row_key, value in (raw_items or {}).items():
        normalized_item = _normalize_manual_release(row_key, value)
        if normalized_item:
            normalized_items[normalized_item["row_key"]] = normalized_item

    meta = dict(payload.get("meta") or {}) if isinstance(payload.get("meta"), dict) else {}
    return {"items": normalized_items, "meta": meta}


def _load_manual_releases_payload():
    payload = _load_state_json(MANUAL_RELEASES_FILE, default=None)
    return _normalize_manual_releases_payload(payload)


def _save_manual_releases_payload(payload):
    normalized_payload = _normalize_manual_releases_payload(payload)
    _write_state_json(MANUAL_RELEASES_FILE, normalized_payload)
    return normalized_payload


def _normalize_duplicate_issue_key(value):
    return str(value or "").strip().upper()


def _is_exact_manual_release_duplicate(manual_release, item):
    release_key = _normalize_duplicate_issue_key(manual_release.get("release_key"))
    item_release_key = _normalize_duplicate_issue_key(item.get("release_key"))
    if not release_key or release_key != item_release_key:
        return False

    rov_key = _normalize_duplicate_issue_key(manual_release.get("rov_key"))
    item_rov_key = _normalize_duplicate_issue_key(item.get("rov_key"))
    return rov_key == item_rov_key


def _find_exact_manual_release_duplicate(manual_release, items=None, *, include_manual=False, ignore_row_key=""):
    ignore_row_key = str(ignore_row_key or manual_release.get("row_key") or "").strip()
    for item in items or []:
        if not isinstance(item, dict):
            continue
        item_row_key = _get_assignment_key_for_item(item)
        if ignore_row_key and item_row_key == ignore_row_key:
            continue
        if item.get("manual_release") and not include_manual:
            continue
        if _is_exact_manual_release_duplicate(manual_release, item):
            return item
    return None


def _find_manual_release_display_duplicate(manual_release, items=None, *, ignore_row_key=""):
    exact_duplicate = _find_exact_manual_release_duplicate(
        manual_release,
        items,
        include_manual=False,
        ignore_row_key=ignore_row_key,
    )
    if exact_duplicate:
        return exact_duplicate

    release_key = _normalize_duplicate_issue_key(manual_release.get("release_key"))
    rov_key = _normalize_duplicate_issue_key(manual_release.get("rov_key"))
    if not release_key or rov_key:
        return None

    release_candidates = []
    for item in items or []:
        if not isinstance(item, dict) or item.get("manual_release"):
            continue
        if ignore_row_key and _get_assignment_key_for_item(item) == ignore_row_key:
            continue
        if _normalize_duplicate_issue_key(item.get("release_key")) == release_key:
            release_candidates.append(item)
    return release_candidates[0] if len(release_candidates) == 1 else None


def _find_jira_release_candidates_for_manual_release(manual_release, jira_items=None):
    if not isinstance(manual_release, dict):
        return []

    release_key = _normalize_duplicate_issue_key(manual_release.get("release_key"))
    rov_key = _normalize_duplicate_issue_key(manual_release.get("rov_key"))
    if not release_key:
        return []

    candidates = []
    for item in jira_items or []:
        if not isinstance(item, dict) or item.get("manual_release"):
            continue
        item_release_key = _normalize_duplicate_issue_key(item.get("release_key"))
        if item_release_key != release_key:
            continue
        item_rov_key = _normalize_duplicate_issue_key(item.get("rov_key"))
        match_type = "exact" if rov_key and item_rov_key == rov_key else "release_key"
        if not rov_key and not item_rov_key:
            match_type = "exact"
        candidates.append({
            "row_key": _get_assignment_key_for_item(item),
            "release_key": str(item.get("release_key") or "").strip(),
            "rov_key": str(item.get("rov_key") or "").strip(),
            "release_status": str(item.get("release_status") or "").strip(),
            "rov_status": str(item.get("rov_status") or "").strip(),
            "deployment_start": str(item.get("deployment_start") or "").strip(),
            "deployment_end": str(item.get("deployment_end") or "").strip(),
            "release_summary": str(item.get("release_summary") or "").strip(),
            "match_type": match_type,
        })
    return candidates


def _manual_release_has_value(item, field_name):
    value = item.get(field_name)
    if isinstance(value, dict):
        return bool(value)
    return bool(str(value or "").strip())


def _apply_manual_release_jira_reconciliation(item, candidates):
    if not isinstance(item, dict):
        return item

    candidates = [candidate for candidate in candidates or [] if isinstance(candidate, dict)]
    item["jira_reconcile_candidates"] = candidates
    item["jira_reconcile_available"] = bool(candidates)
    item["jira_reconcile_action_required"] = False
    item["jira_reconcile_candidate_count"] = len(candidates)
    item["jira_reconcile_suggested_fields"] = {}
    item["jira_reconcile_matched_row_key"] = ""

    if not candidates:
        return item

    exact_candidates = [candidate for candidate in candidates if candidate.get("match_type") == "exact"]
    selected = exact_candidates[0] if exact_candidates else candidates[0]
    item["jira_reconcile_matched_row_key"] = str(selected.get("row_key") or "")
    item["jira_reconcile_action_required"] = True

    suggested_fields = {}
    for field_name in (
        "release_summary",
        "rov_key",
        "release_status",
        "rov_status",
        "deployment_start",
        "deployment_end",
    ):
        candidate_value = str(selected.get(field_name) or "").strip()
        if candidate_value and not _manual_release_has_value(item, field_name):
            suggested_fields[field_name] = candidate_value

    item["jira_reconcile_suggested_fields"] = suggested_fields
    return item


def _find_manual_release_duplicate_warnings(manual_release, jira_items=None, *, include_manual=False, ignore_row_key=""):
    if not isinstance(manual_release, dict):
        return []

    release_key = _normalize_duplicate_issue_key(manual_release.get("release_key"))
    rov_key = _normalize_duplicate_issue_key(manual_release.get("rov_key"))
    if not release_key and not rov_key:
        return []

    warnings = []
    for item in jira_items or []:
        if not isinstance(item, dict):
            continue
        item_row_key = _get_assignment_key_for_item(item)
        if ignore_row_key and item_row_key == ignore_row_key:
            continue
        is_manual_item = bool(item.get("manual_release"))
        if is_manual_item and not include_manual:
            continue

        item_release_key = _normalize_duplicate_issue_key(item.get("release_key"))
        item_rov_key = _normalize_duplicate_issue_key(item.get("rov_key"))
        if release_key and release_key == item_release_key:
            exact_duplicate = rov_key == item_rov_key
            missing_rov_duplicate = bool(not rov_key and item_rov_key)
            is_blocking = bool(exact_duplicate or missing_rov_duplicate)
            warnings.append({
                "type": "exact_duplicate" if exact_duplicate else "release_key_without_rov" if missing_rov_duplicate else "release_key",
                "severity": "error" if is_blocking else "warning",
                "blocking": is_blocking,
                "message": (
                    f"Manual release duplicates existing row {release_key}"
                    if is_blocking
                    else f"Manual release has the same release key {release_key}"
                ),
                "row_key": item_row_key,
                "release_key": item_release_key,
                "rov_key": item_rov_key,
                "source": "manual" if is_manual_item else "jira",
            })
        if rov_key and rov_key == item_rov_key:
            warnings.append({
                "type": "rov_key",
                "severity": "warning",
                "blocking": False,
                "message": f"Manual release has the same ROV {rov_key}",
                "row_key": item_row_key,
                "release_key": item_release_key,
                "rov_key": item_rov_key,
                "source": "manual" if is_manual_item else "jira",
            })
    return warnings


def _apply_manual_releases(items):
    manual_releases = _load_manual_releases_payload().get("items", {})
    if not manual_releases:
        return items

    jira_items = [item for item in items if isinstance(item, dict) and not item.get("manual_release")]
    existing_keys = {_get_assignment_key_for_item(item) for item in items if isinstance(item, dict)}
    for row_key, manual_item in manual_releases.items():
        if row_key in existing_keys:
            continue
        item_to_append = dict(manual_item)
        reconciliation_candidates = _find_jira_release_candidates_for_manual_release(item_to_append, jira_items)
        _apply_manual_release_jira_reconciliation(item_to_append, reconciliation_candidates)
        duplicate_item = _find_manual_release_display_duplicate(item_to_append, jira_items)
        if duplicate_item:
            item_to_append["jira_duplicate_detected"] = True
            item_to_append["jira_duplicate_row_key"] = _get_assignment_key_for_item(duplicate_item)
            item_to_append["jira_duplicate_release_key"] = str(duplicate_item.get("release_key") or "").strip()
            item_to_append["jira_duplicate_rov_key"] = str(duplicate_item.get("rov_key") or "").strip()
            item_to_append["jira_duplicate_action_required"] = True
        else:
            item_to_append["jira_duplicate_detected"] = False
            item_to_append["jira_duplicate_row_key"] = ""
            item_to_append["jira_duplicate_release_key"] = ""
            item_to_append["jira_duplicate_rov_key"] = ""
            item_to_append["jira_duplicate_action_required"] = False
        items.append(item_to_append)
        existing_keys.add(row_key)
    return items


def _apply_manual_duplicate_reconciliation(items):
    jira_items = [item for item in items if isinstance(item, dict) and not item.get("manual_release")]
    if not jira_items:
        return items

    for manual_item in (item for item in items if isinstance(item, dict) and item.get("manual_release")):
        duplicate_item = _find_manual_release_display_duplicate(manual_item, jira_items)
        if not duplicate_item:
            manual_item["jira_duplicate_detected"] = False
            manual_item["manual_hidden_by_jira_duplicate"] = False
            continue

        duplicate_row_key = _get_assignment_key_for_item(duplicate_item)
        manual_item["jira_duplicate_detected"] = True
        manual_item["manual_hidden_by_jira_duplicate"] = True
        manual_item["jira_duplicate_row_key"] = duplicate_row_key
        manual_item["jira_duplicate_release_key"] = str(duplicate_item.get("release_key") or "").strip()
        manual_item["jira_duplicate_rov_key"] = str(duplicate_item.get("rov_key") or "").strip()
        manual_item["jira_duplicate_action_required"] = True

        duplicate_item["manual_duplicate_detected"] = True
        duplicate_item["manual_duplicate_row_key"] = _get_assignment_key_for_item(manual_item)

        if not _has_release_responsible(duplicate_item) and _has_release_responsible(manual_item):
            duplicate_item["psi_responsibles"] = list(manual_item.get("psi_responsibles") or [])
        if not str(duplicate_item.get("psi_owner") or "").strip() and str(manual_item.get("psi_owner") or "").strip():
            duplicate_item["psi_owner"] = manual_item.get("psi_owner", "")
            duplicate_item["psi_owner_source"] = manual_item.get("psi_owner_source", "")
            duplicate_item["psi_owner_date"] = manual_item.get("psi_owner_date", "")
            duplicate_item["psi_zni_reviewer"] = manual_item.get("psi_zni_reviewer", "")
        if not str(duplicate_item.get("psi_checker") or "").strip() and str(manual_item.get("psi_checker") or "").strip():
            duplicate_item["psi_checker"] = manual_item.get("psi_checker", "")

        if not str(duplicate_item.get("zni_key") or "").strip() and str(manual_item.get("zni_key") or "").strip():
            duplicate_item["zni_key"] = manual_item.get("zni_key", "")
            duplicate_item["zni_url"] = manual_item.get("zni_url", "")
        if not str(duplicate_item.get("rollout_notes_level") or "").strip() and str(manual_item.get("rollout_notes_level") or "").strip():
            duplicate_item["rollout_notes_level"] = manual_item.get("rollout_notes_level", "")
            duplicate_item["has_rollout_notes"] = bool(manual_item.get("has_rollout_notes"))
    return items


def _load_zni_assignments():
    return _load_zni_payload().get("issues", {})


def _save_zni_assignments(assignments):
    payload = _load_zni_payload()
    payload["issues"] = _normalize_zni_assignments(assignments)
    _save_zni_payload(payload)


def _remove_zni_assignment_for_keys(*keys):
    normalized_keys = [
        str(key or "").strip()
        for key in keys
        if str(key or "").strip()
    ]
    if not normalized_keys:
        return False

    assignments = _load_zni_assignments()
    changed = False
    for key in normalized_keys:
        if key in assignments:
            assignments.pop(key, None)
            changed = True
    if changed:
        _save_zni_assignments(assignments)
    return changed


def _clear_cached_zni_fields_for_keys(*keys):
    if _cached_data is None:
        return

    normalized_keys = {
        str(key or "").strip()
        for key in keys
        if str(key or "").strip()
    }
    if not normalized_keys:
        return

    for item in _cached_data.get("items") or []:
        assignment_key = _get_assignment_key_for_item(item)
        release_key = str(item.get("release_key") or "").strip()
        if assignment_key not in normalized_keys and release_key not in normalized_keys:
            continue
        for field_name in (
            "zni_key",
            "zni_url",
            "base_zni_key",
            "base_zni_url",
            "manual_zni_key",
            "manual_zni_url",
        ):
            item[field_name] = ""
        item["manual_clear_zni"] = False


def _load_rollout_note_flags():
    return _load_zni_payload().get("flags", {})


def _save_rollout_note_flags(flags):
    payload = _load_zni_payload()
    payload["flags"] = _normalize_rollout_note_flags(flags)
    _save_zni_payload(payload)


def _normalize_manual_release_override(value):
    if not isinstance(value, dict):
        return {}

    _warn_unknown_manual_override_fields(value.keys())
    normalized = {}
    for field_name in MANUAL_OVERRIDE_SCALAR_FIELDS:
        normalized_value = _normalize_manual_scalar_field(field_name, value.get(field_name))
        if normalized_value:
            normalized[field_name] = normalized_value

    for field_name in MANUAL_OVERRIDE_DICT_FIELDS:
        normalized_value = _normalize_manual_dict_field(value.get(field_name))
        if normalized_value:
            normalized[field_name] = normalized_value

    for field_name in BASE_OVERRIDE_FIELDS:
        base_field = _base_field_name(field_name)
        if field_name in MANUAL_OVERRIDE_DICT_FIELDS:
            normalized_value = _normalize_manual_dict_field(value.get(base_field))
        else:
            normalized_value = _normalize_manual_scalar_field(field_name, value.get(base_field))
        if normalized_value:
            normalized[base_field] = normalized_value

    clear_zni = bool(value.get("clear_zni"))
    if clear_zni:
        normalized["clear_zni"] = True

    updated_at = str(value.get("updated_at") or "").strip()
    if updated_at:
        normalized["updated_at"] = updated_at
    updated_by = str(value.get("updated_by") or "").strip()
    if updated_by:
        normalized["updated_by"] = updated_by

    return normalized


def _normalize_manual_release_overrides(payload):
    normalized = {}
    if not isinstance(payload, dict):
        return normalized

    for key, value in payload.items():
        row_key = str(key or "").strip()
        normalized_value = _normalize_manual_release_override(value)
        if row_key and normalized_value:
            normalized[row_key] = normalized_value
    return normalized


def _load_manual_release_overrides():
    return _load_manual_overrides_payload()


def _save_manual_release_overrides(overrides):
    normalized_overrides = _normalize_manual_release_overrides(overrides)
    _save_manual_overrides_payload(normalized_overrides)
    return normalized_overrides


def _load_manual_order():
    global _manual_order_cache

    def _normalize_group_order(value):
        if isinstance(value, dict):
            normalized_buckets = {}
            for bucket, row_keys in value.items():
                bucket_key = str(bucket or "").strip()
                if not bucket_key or not isinstance(row_keys, list):
                    continue
                normalized_keys = []
                for row_key in row_keys:
                    row_key = str(row_key or "").strip()
                    if row_key and row_key not in normalized_keys:
                        normalized_keys.append(row_key)
                if normalized_keys:
                    normalized_buckets[bucket_key] = normalized_keys
            return normalized_buckets

        return [
            str(item or "").strip()
            for item in (value or [])
            if str(item or "").strip()
        ]

    if isinstance(_manual_order_cache, dict):
        return _manual_order_cache

    payload = _load_state_json(ORDER_FILE, default=None)
    if not isinstance(payload, dict):
        _manual_order_cache = {}
        return {}

    normalized = {}
    for year, value in payload.items():
        year_key = str(year or "").strip()
        if not year_key or not isinstance(value, dict):
            continue

        normalized[year_key] = {
            "waiting": _normalize_group_order(value.get("waiting")),
            "numbered": _normalize_group_order(value.get("numbered")),
            "force_unnumbered": [
                str(item or "").strip()
                for item in (value.get("force_unnumbered") or [])
                if str(item or "").strip()
            ],
            "force_numbered": [
                str(item or "").strip()
                for item in (value.get("force_numbered") or [])
                if str(item or "").strip()
            ],
        }
    _manual_order_cache = normalized
    return normalized


def _save_manual_order(order_payload):
    global _manual_order_cache

    try:
        payload = dict(order_payload or {})
        _write_state_json(ORDER_FILE, payload)
        _manual_order_cache = payload
    except Exception as exc:
        logging.warning("Release monitor: failed to save manual order: %s", exc)
        raise


def _remove_row_from_manual_order(row_key):
    row_key = str(row_key or "").strip()
    if not row_key:
        return

    manual_order = _load_manual_order()
    changed = False
    for year_payload in manual_order.values():
        if not isinstance(year_payload, dict):
            continue
        for group_name in ("waiting", "numbered", "force_unnumbered", "force_numbered"):
            values = year_payload.get(group_name)
            if isinstance(values, dict):
                next_values = {}
                for bucket, bucket_values in values.items():
                    if not isinstance(bucket_values, list):
                        continue
                    filtered_values = [
                        value
                        for value in bucket_values
                        if not _manual_order_row_matches_release(str(value or "").strip(), row_key)
                    ]
                    if len(filtered_values) != len(bucket_values):
                        changed = True
                    if filtered_values:
                        next_values[bucket] = filtered_values
                if next_values != values:
                    year_payload[group_name] = next_values
                continue

            if isinstance(values, list):
                filtered_values = [
                    value
                    for value in values
                    if not _manual_order_row_matches_release(str(value or "").strip(), row_key)
                ]
                if len(filtered_values) != len(values):
                    year_payload[group_name] = filtered_values
                    changed = True

    if changed:
        _save_manual_order(manual_order)


def _manual_order_row_matches_release(stored_row_key, release_key):
    stored_row_key = str(stored_row_key or "").strip()
    release_key = str(release_key or "").strip()
    if not stored_row_key or not release_key:
        return False
    return stored_row_key == release_key or stored_row_key.startswith(f"{release_key}::")


def _get_release_order_bucket(item):
    raw_value = str(
        item.get("deployment_start_iso")
        or item.get("deployment_start")
        or item.get("sort_date")
        or "no-date"
    ).strip()
    parsed = _parse_release_monitor_date(raw_value)
    if parsed:
        return parsed.strftime("%Y-%m-%d")
    return raw_value[:10] if raw_value else "no-date"


def _build_empty_duty_schedule_payload():
    return {
        "dates": {},
        "availability": {},
        "months": [],
        "files": [],
        "last_upload": None,
    }


def _normalize_duty_availability_payload(value):
    normalized = {}
    if not isinstance(value, dict):
        return normalized

    for date_key, people in value.items():
        normalized_date = str(date_key or "").strip()
        if not normalized_date or not isinstance(people, dict):
            continue

        normalized_people = {}
        for person, info in people.items():
            person_name = str(person or "").strip()
            if not person_name:
                continue

            if isinstance(info, dict):
                status = str(info.get("status") or "").strip()
                availability = str(info.get("availability") or "").strip()
                reason = str(info.get("reason") or "").strip()
            else:
                status = str(info or "").strip()
                availability, reason = _classify_duty_status(status)

            if not status:
                continue
            if not availability:
                availability, reason = _classify_duty_status(status)
            normalized_people[person_name] = {
                "status": status,
                "availability": availability,
                "reason": reason,
            }

        if normalized_people:
            normalized[normalized_date] = normalized_people

    return normalized


def _load_duty_schedule_payload():
    payload = _load_state_json(DUTY_SCHEDULE_FILE, default=None)
    if not isinstance(payload, dict):
        return _build_empty_duty_schedule_payload()

    dates = {
        str(date_key).strip(): str(reviewer).strip()
        for date_key, reviewer in (payload.get("dates") or {}).items()
        if str(date_key).strip() and str(reviewer).strip()
    }
    availability = _normalize_duty_availability_payload(payload.get("availability") or {})
    months = [
        str(month_label).strip()
        for month_label in (payload.get("months") or [])
        if str(month_label).strip()
    ]
    files = [
        dict(file_info)
        for file_info in (payload.get("files") or [])
        if isinstance(file_info, dict)
    ]
    return {
        "dates": dates,
        "availability": availability,
        "months": list(dict.fromkeys(months)),
        "files": files,
        "last_upload": str(payload.get("last_upload") or "").strip() or None,
    }


def _save_duty_schedule_payload(payload):
    try:
        _write_state_json(DUTY_SCHEDULE_FILE, payload or _build_empty_duty_schedule_payload())
    except Exception as exc:
        logging.warning("Release monitor: failed to save duty schedules: %s", exc)
        raise


def _load_date_overrides():
    payload = _load_state_json(DATE_OVERRIDES_FILE, default=None)
    if not isinstance(payload, dict):
        return {}

    normalized = {}
    for row_key, value in payload.items():
        key = str(row_key or "").strip()
        if not key or not isinstance(value, dict):
            continue

        start_value = str(value.get("start", "") or "").strip()
        end_value = str(value.get("end", "") or "").strip()
        if not start_value and not end_value:
            continue

        normalized[key] = {
            "start": start_value,
            "end": end_value,
            "updated_at": str(value.get("updated_at", "") or "").strip(),
        }
    return normalized


def _save_date_overrides(payload):
    try:
        _write_state_json(DATE_OVERRIDES_FILE, payload or {})
    except Exception as exc:
        logging.warning("Release monitor: failed to save date overrides: %s", exc)
        raise


def _normalize_release_attempt_outcomes(payload):
    normalized = {}
    if not isinstance(payload, dict):
        return normalized

    for row_key, value in payload.items():
        key = str(row_key or "").strip()
        if not key or not isinstance(value, dict):
            continue
        state = str(value.get("state") or "").strip().lower()
        if state != "deferred":
            continue
        normalized[key] = {
            "state": "deferred",
            "release_key": str(value.get("release_key") or "").strip(),
            "rov_key": str(value.get("rov_key") or "").strip(),
            "detected_at": str(value.get("detected_at") or "").strip(),
            "updated_at": str(value.get("updated_at") or "").strip(),
        }
    return normalized


def _load_release_attempt_outcomes():
    payload = _load_state_json(ATTEMPTS_FILE, default={})
    return _normalize_release_attempt_outcomes(payload)


def _save_release_attempt_outcomes(payload):
    try:
        _write_state_json(ATTEMPTS_FILE, _normalize_release_attempt_outcomes(payload))
    except Exception as exc:
        logging.warning("Release monitor: failed to save release attempt outcomes: %s", exc)
        raise


def _parse_release_monitor_date(value):
    if not value:
        return None

    parsed = _parse_jira_date(str(value))
    if parsed:
        return parsed

    for fmt in ("%d.%m.%Y", "%d.%m.%Y %H:%M", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(str(value), fmt)
        except ValueError:
            continue
    return None


def _format_release_monitor_date(dt_value):
    return dt_value.strftime("%d.%m.%Y") if dt_value else ""


def _is_release_window_expired(end_dt, now_dt=None):
    if not end_dt:
        return False
    now_dt = now_dt or datetime.now()
    if end_dt.time() == datetime.min.time():
        return end_dt.date() < now_dt.date()
    return end_dt < now_dt


def _release_days_overdue(end_dt, now_dt=None):
    if not end_dt or not _is_release_window_expired(end_dt, now_dt):
        return 0
    now_dt = now_dt or datetime.now()
    return max((now_dt.date() - end_dt.date()).days, 0)


def _normalize_rule_status(value):
    return _normalize_text(value).replace("\u0451", "\u0435")


def _is_approved_rov_status(status):
    normalized = _normalize_rule_status(status)
    approved_statuses = {_normalize_rule_status(value) for value in APPROVED_ROV_STATUSES}
    return normalized in approved_statuses


def _is_cancelled_rov_status(status):
    normalized = _normalize_rule_status(status)
    return normalized in {"отменен", "отменено", "cancelled", "canceled"}


def _is_final_release_status(status):
    return _normalize_rule_status(status) == _normalize_rule_status(FINAL_RELEASE_STATUS)


def _is_cancelled_reroll_rov(item):
    if not isinstance(item, dict):
        return False
    return bool(
        item.get("is_reroll")
        and _is_final_release_status(item.get("release_status"))
        and _is_cancelled_rov_status(item.get("rov_status"))
    )


def _is_confirmed_deployment_attempt(item):
    if not isinstance(item, dict):
        return False
    if item.get("is_pre_final") or item.get("is_ready_for_prom"):
        return True
    return _is_approved_rov_status(item.get("rov_status"))


def _get_release_operational_day_bounds(now_dt=None):
    now_dt = now_dt or datetime.now()
    start_dt = now_dt.replace(
        hour=RELEASE_OPERATIONAL_DAY_START_HOUR,
        minute=0,
        second=0,
        microsecond=0,
    )
    if now_dt < start_dt:
        start_dt -= timedelta(days=1)
    return start_dt, start_dt + timedelta(days=1)


def _is_release_window_in_operational_day(start_dt, end_dt=None, now_dt=None):
    window_start, window_end = _get_release_operational_day_bounds(now_dt)
    release_start = start_dt or end_dt
    release_end = end_dt or start_dt
    if not release_start:
        return False
    if release_end and release_end < release_start:
        release_start, release_end = release_end, release_start
    if release_end == release_start and release_start.time().replace(microsecond=0) == datetime.min.time():
        return release_start.date() == window_start.date()
    if release_end == release_start:
        return bool(window_start <= release_start < window_end)
    return bool(release_start < window_end and release_end > window_start)


def _get_final_manual_date_row_state(start_dt, end_dt, now_dt=None):
    now_dt = now_dt or datetime.now()
    window_end_dt = end_dt or start_dt
    if _is_release_window_expired(window_end_dt, now_dt):
        return "final"

    if _is_release_window_in_operational_day(start_dt, end_dt, now_dt):
        return "today"
    return "planned"


def _apply_active_reroll_schedule_state(item, start_dt=None, end_dt=None, now_dt=None):
    if not item.get("is_reroll"):
        return False

    now_dt = now_dt or datetime.now()
    start_dt = start_dt or _parse_release_monitor_date(
        item.get("deployment_start_iso")
        or item.get("source_deployment_start_iso")
        or item.get("deployment_start")
        or item.get("source_deployment_start")
    )
    end_dt = end_dt or _parse_release_monitor_date(
        item.get("deployment_end_iso")
        or item.get("source_deployment_end_iso")
        or item.get("deployment_end")
        or item.get("source_deployment_end")
    )
    window_end_dt = end_dt or start_dt
    if not window_end_dt or _is_release_window_expired(window_end_dt, now_dt):
        return False

    is_today = _is_release_window_in_operational_day(start_dt, end_dt, now_dt)

    item["is_final"] = False
    item["is_non_final"] = True
    item["is_pre_final"] = False
    item["is_ready_for_prom"] = False
    item["is_overdue"] = False
    item["is_today"] = is_today
    item["days_overdue"] = 0
    item["row_state"] = "today" if is_today else "planned"
    return True


def _resolve_effective_release_date(base_dt, manual_dt):
    return manual_dt or base_dt


def _release_dates_match(left_dt, right_dt):
    if not left_dt or not right_dt:
        return False
    return _format_release_monitor_date(left_dt) == _format_release_monitor_date(right_dt)


def _release_base_covers_manual_date(base_dt, manual_dt):
    return _release_dates_match(base_dt, manual_dt)


def _extract_schedule_month_year(sheet_title):
    normalized_title = _normalize_text(sheet_title)
    year_match = re.search(r"(20\d{2})", normalized_title)
    if not year_match:
        return None, None

    month = None
    for month_token, month_number in MONTH_NAME_MAP.items():
        if month_token in normalized_title:
            month = month_number
            break

    if month is None:
        return None, None

    return month, int(year_match.group(1))


def _coerce_schedule_day(value):
    if isinstance(value, int):
        return value if 1 <= value <= 31 else None
    if isinstance(value, float) and value.is_integer():
        day_value = int(value)
        return day_value if 1 <= day_value <= 31 else None

    text_value = str(value or "").strip()
    if text_value.isdigit():
        day_value = int(text_value)
        return day_value if 1 <= day_value <= 31 else None
    return None


def _detect_schedule_day_columns(worksheet):
    max_scan_row = min(12, worksheet.max_row)
    max_scan_col = min(50, worksheet.max_column)

    for row_index in range(1, max_scan_row + 1):
        detected = []
        for col_index in range(2, max_scan_col + 1):
            day_value = _coerce_schedule_day(worksheet.cell(row_index, col_index).value)
            if day_value is not None:
                detected.append((col_index, day_value))

        if len(detected) >= 20:
            return row_index, detected

    return None, []


def _classify_duty_status(status):
    raw_status = str(status or "").strip()
    if not raw_status:
        return "", ""

    normalized_status = _normalize_text(raw_status)
    upper_status = raw_status.upper()
    if upper_status in DUTY_HARD_EXCLUDED_STATUSES:
        return "excluded", {
            "ДД": "Дневной дежурный",
            "ВД": "Вечерний дежурный",
            "ВР": "Вечерний резервный дежурный",
            "ХД": "Хабаровская смена",
            "ХР": "Хабаровский резерв",
        }.get(upper_status, "Дежурство")
    if upper_status in DUTY_RESERVE_STATUSES:
        return "reserve", "Дневной резервный ответственный"
    if any(keyword in normalized_status for keyword in DUTY_ABSENCE_KEYWORDS):
        return "excluded", raw_status
    if "празд" in normalized_status:
        return "excluded", raw_status
    return "available", ""


def _parse_duty_schedule_sheet(worksheet, filename):
    month, year = _extract_schedule_month_year(worksheet.title)
    if not month or not year:
        return {"dates": {}, "months": [], "warnings": []}

    day_row_index, day_columns = _detect_schedule_day_columns(worksheet)
    if not day_row_index or not day_columns:
        return {"dates": {}, "months": [], "warnings": []}

    daily_candidates = defaultdict(list)
    availability_dates = defaultdict(dict)
    max_scan_row = min(worksheet.max_row, day_row_index + 40)

    for row_index in range(day_row_index + 1, max_scan_row + 1):
        raw_name = str(worksheet.cell(row_index, 1).value or "").strip()
        if not raw_name:
            continue

        matched_name = _match_oplot_name(raw_name)
        for col_index, day_number in day_columns:
            raw_marker = str(worksheet.cell(row_index, col_index).value or "").strip()
            marker = _normalize_text(raw_marker)
            if marker != "\u0432\u0434":
                if matched_name and raw_marker:
                    try:
                        day_date = datetime(year, month, day_number).date()
                    except ValueError:
                        continue
                    availability, reason = _classify_duty_status(raw_marker)
                    if availability in {"excluded", "reserve"}:
                        availability_dates[day_date.isoformat()][matched_name] = {
                            "status": raw_marker,
                            "availability": availability,
                            "reason": reason,
                        }
                continue

            if not matched_name:
                daily_candidates[day_number].append({"name": raw_name, "matched": ""})
            else:
                daily_candidates[day_number].append({"name": raw_name, "matched": matched_name})
                try:
                    day_date = datetime(year, month, day_number).date()
                except ValueError:
                    continue
                availability, reason = _classify_duty_status(raw_marker)
                availability_dates[day_date.isoformat()][matched_name] = {
                    "status": raw_marker,
                    "availability": availability,
                    "reason": reason,
                }

    parsed_dates = {}
    warnings = []
    month_label = f"{year:04d}-{month:02d}"

    for day_number, entries in daily_candidates.items():
        try:
            day_date = datetime(year, month, day_number).date()
        except ValueError:
            continue

        if day_date.weekday() >= 5:
            continue

        matched_names = [entry["matched"] for entry in entries if entry.get("matched")]
        unique_names = list(dict.fromkeys(matched_names))

        if len(unique_names) == 1:
            parsed_dates[day_date.isoformat()] = unique_names[0]
            continue

        if not unique_names:
            source_names = ", ".join(entry.get("name", "") for entry in entries if entry.get("name"))
            warnings.append(
                f"{worksheet.title}: не удалось сопоставить ВД за {day_date.strftime('%d.%m.%Y')} ({source_names})"
            )
        else:
            warnings.append(
                f"{worksheet.title}: найдено несколько ВД за {day_date.strftime('%d.%m.%Y')} ({', '.join(unique_names)})"
            )

    if not parsed_dates and not availability_dates:
        return {"dates": {}, "availability": {}, "months": [], "warnings": warnings}

    return {
        "dates": parsed_dates,
        "availability": dict(availability_dates),
        "months": [month_label],
        "warnings": warnings,
    }


def _parse_duty_schedule_workbook(file_bytes, filename):
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    merged_dates = {}
    merged_availability = {}
    parsed_months = []
    warnings = []

    for worksheet in workbook.worksheets:
        parsed_sheet = _parse_duty_schedule_sheet(worksheet, filename)
        merged_dates.update(parsed_sheet.get("dates", {}))
        for date_key, people in (parsed_sheet.get("availability") or {}).items():
            merged_availability.setdefault(date_key, {}).update(people or {})
        parsed_months.extend(parsed_sheet.get("months", []))
        warnings.extend(parsed_sheet.get("warnings", []))

    if not merged_dates and not merged_availability:
        raise ValueError(f"\u0412 \u0444\u0430\u0439\u043b\u0435 {filename} \u043d\u0435 \u0443\u0434\u0430\u043b\u043e\u0441\u044c \u043d\u0430\u0439\u0442\u0438 \u043b\u0438\u0441\u0442\u044b \u0441 \u0433\u0440\u0430\u0444\u0438\u043a\u043e\u043c \u0434\u0435\u0436\u0443\u0440\u0441\u0442\u0432")

    return {
        "dates": merged_dates,
        "availability": merged_availability,
        "months": list(dict.fromkeys(parsed_months)),
        "warnings": warnings,
        "files": [{
            "name": filename,
            "uploaded_at": _format_timestamp(),
            "months": list(dict.fromkeys(parsed_months)),
        }],
    }


def _merge_duty_schedule_payload(base_payload, incoming_payload):
    merged = _build_empty_duty_schedule_payload()
    merged["dates"] = dict((base_payload or {}).get("dates") or {})
    merged["dates"].update((incoming_payload or {}).get("dates") or {})
    merged["availability"] = _normalize_duty_availability_payload((base_payload or {}).get("availability") or {})
    for date_key, people in _normalize_duty_availability_payload((incoming_payload or {}).get("availability") or {}).items():
        merged["availability"].setdefault(date_key, {}).update(people)
    merged["months"] = list(
        dict.fromkeys(
            list((base_payload or {}).get("months") or [])
            + list((incoming_payload or {}).get("months") or [])
        )
    )
    merged["files"] = list((base_payload or {}).get("files") or []) + list((incoming_payload or {}).get("files") or [])
    merged["last_upload"] = (
        (incoming_payload or {}).get("last_upload")
        or (base_payload or {}).get("last_upload")
    )
    return merged


def _is_explicit_manual_reviewer_assignment(assignment):
    if not isinstance(assignment, dict):
        return False
    return (
        str(assignment.get("reviewer_source") or "").strip() in {"manual", "manual_text"}
        and str(assignment.get("reviewer_date") or "").strip() == "manual"
    )


def _apply_duty_schedule_assignments(items, persist=False, force=False, debug_limit=0):
    assignments = _load_reviewer_assignments()
    duty_payload = _load_duty_schedule_payload()
    duty_dates = duty_payload.get("dates") or {}
    changed = False
    applied_count = 0
    debug_rows = []

    for item in items:
        assignment_key = _get_assignment_key_for_item(item)
        current_assignment = dict(assignments.get(assignment_key) or {})
        current_reviewer = str(current_assignment.get("reviewer") or "").strip()
        reviewer_source = str(current_assignment.get("reviewer_source") or "").strip()
        if not force and current_reviewer and _is_explicit_manual_reviewer_assignment(current_assignment):
            if debug_limit and len(debug_rows) < debug_limit:
                debug_rows.append({
                    "row_key": assignment_key,
                    "release_key": item.get("release_key", ""),
                    "date": "",
                    "previous": current_reviewer,
                    "scheduled": "",
                    "result": current_reviewer,
                    "reason": "manual",
                })
            continue

        deployment_dt = _parse_release_monitor_date(item.get("deployment_start_iso") or item.get("deployment_start"))
        if not deployment_dt or deployment_dt.date().weekday() >= 5:
            if force or current_reviewer or reviewer_source == "duty_schedule":
                current_assignment["reviewer"] = ""
                current_assignment["reviewer_source"] = ""
                current_assignment["reviewer_date"] = ""
                current_assignment["checker"] = str(current_assignment.get("checker", "") or "").strip()
                raw_responsibles = current_assignment.get("responsibles") or []
                if not isinstance(raw_responsibles, list):
                    raw_responsibles = [raw_responsibles] if raw_responsibles else []
                current_assignment["responsibles"] = [
                    str(value or "").strip()
                    for value in raw_responsibles
                    if str(value or "").strip()
                ]
                if current_assignment["checker"] or current_assignment["responsibles"]:
                    assignments[assignment_key] = current_assignment
                else:
                    assignments.pop(assignment_key, None)
                item["psi_owner"] = ""
                item["psi_owner_source"] = ""
                item["psi_owner_date"] = ""
                item["psi_checker"] = current_assignment["checker"]
                item["psi_responsibles"] = list(current_assignment["responsibles"])
                changed = True
            if debug_limit and len(debug_rows) < debug_limit:
                debug_rows.append({
                    "row_key": assignment_key,
                    "release_key": item.get("release_key", ""),
                    "date": deployment_dt.date().isoformat() if deployment_dt else "",
                    "previous": current_reviewer,
                    "scheduled": "",
                    "result": "",
                    "reason": "no-date-or-weekend",
                })
            continue

        deployment_date = deployment_dt.date()
        reviewer_name = str(duty_dates.get(deployment_date.isoformat()) or "").strip()
        if reviewer_name and reviewer_name not in OPLOT_VALUES:
            reviewer_name = _match_oplot_name(reviewer_name)
        reviewer_name = reviewer_name or ""

        if (
            current_reviewer == reviewer_name
            and reviewer_source == "duty_schedule"
            and str(current_assignment.get("reviewer_date") or "") == deployment_date.isoformat()
            and not force
        ):
            item["psi_owner"] = reviewer_name
            item["psi_owner_source"] = "duty_schedule"
            item["psi_owner_date"] = deployment_date.isoformat()
            continue

        current_assignment["reviewer"] = reviewer_name
        current_assignment["reviewer_source"] = "duty_schedule" if reviewer_name else ""
        current_assignment["reviewer_date"] = deployment_date.isoformat() if reviewer_name else ""
        current_assignment["checker"] = str(current_assignment.get("checker", "") or "").strip()
        raw_responsibles = current_assignment.get("responsibles") or []
        if not isinstance(raw_responsibles, list):
            raw_responsibles = [raw_responsibles] if raw_responsibles else []
        current_assignment["responsibles"] = [
            str(value or "").strip()
            for value in raw_responsibles
            if str(value or "").strip()
        ]
        if current_assignment["reviewer"] or current_assignment["checker"] or current_assignment["responsibles"]:
            assignments[assignment_key] = current_assignment
        else:
            assignments.pop(assignment_key, None)

        item["psi_owner"] = reviewer_name
        item["psi_owner_source"] = current_assignment["reviewer_source"]
        item["psi_owner_date"] = current_assignment["reviewer_date"]
        item["psi_checker"] = current_assignment["checker"]
        item["psi_responsibles"] = list(current_assignment["responsibles"])
        changed = True
        if reviewer_name:
            applied_count += 1
        if debug_limit and len(debug_rows) < debug_limit:
            debug_rows.append({
                "row_key": assignment_key,
                "release_key": item.get("release_key", ""),
                "date": deployment_date.isoformat(),
                "previous": current_reviewer,
                "scheduled": reviewer_name,
                "result": reviewer_name,
                "reason": "updated" if reviewer_name else "no-duty-for-date",
            })

    if changed and persist:
        _save_reviewer_assignments(assignments)

    if debug_limit:
        return {
            "applied_count": applied_count,
            "debug_rows": debug_rows,
        }
    return applied_count


def _get_scheduled_duty_reviewer_for_item(item):
    deployment_dt = _parse_release_monitor_date(item.get("deployment_start_iso") or item.get("deployment_start"))
    if not deployment_dt or deployment_dt.date().weekday() >= 5:
        return ""

    duty_payload = _load_duty_schedule_payload()
    duty_dates = duty_payload.get("dates") or {}
    reviewer_name = str(duty_dates.get(deployment_dt.date().isoformat()) or "").strip()
    if reviewer_name and reviewer_name not in OPLOT_VALUES:
        reviewer_name = _match_oplot_name(reviewer_name)
    return reviewer_name or ""


def _prepare_item_for_zni_creation(item):
    zni_item = dict(item or {})
    if str(zni_item.get("psi_owner_source") or "").strip() != "manual_text":
        return zni_item

    zni_reviewer = str(zni_item.get("psi_zni_reviewer") or "").strip()
    if zni_reviewer and zni_reviewer not in OPLOT_VALUES:
        zni_reviewer = _match_oplot_name(zni_reviewer)

    if not zni_reviewer:
        zni_reviewer = _get_scheduled_duty_reviewer_for_item(zni_item)

    if not zni_reviewer:
        raise ValueError(
            "Для создания ЗНИ не удалось определить дежурного ОПЛОТ. "
            "Если в строке указан устанавливающий, сначала выберите дежурного в таблице, "
            "а затем переключите поле обратно в режим устанавливающего."
        )

    zni_item["manual_installer"] = str(zni_item.get("psi_owner") or "").strip()
    zni_item["psi_owner"] = zni_reviewer
    zni_item["psi_owner_source"] = "manual"
    return zni_item


def _append_duty_schedule_meta(meta):
    duty_payload = _load_duty_schedule_payload()
    meta["last_duty_schedule_upload"] = duty_payload.get("last_upload")
    meta["duty_schedule_months"] = list(duty_payload.get("months") or [])
    meta["duty_schedule_files"] = list(duty_payload.get("files") or [])
    meta["work_mark_suggested_participants"] = _collect_work_mark_suggested_participants(duty_payload)
    return meta


def _collect_work_mark_suggested_participants(duty_payload=None):
    duty_payload = duty_payload if isinstance(duty_payload, dict) else _load_duty_schedule_payload()
    week_start, week_end = _get_current_week_bounds()
    dates = duty_payload.get("dates") or {}
    availability = duty_payload.get("availability") or {}
    participants = []

    def add_participant(name):
        raw_name = str(name or "").strip()
        if not raw_name:
            return
        matched_name = raw_name if raw_name in OPLOT_VALUES else _match_oplot_name(raw_name)
        if matched_name and matched_name not in participants:
            participants.append(matched_name)

    current_date = week_start
    while current_date <= week_end:
        date_key = current_date.isoformat()
        add_participant(dates.get(date_key))
        for person, info in (availability.get(date_key) or {}).items():
            status = str((info or {}).get("status") or "").strip().upper()
            reason = str((info or {}).get("reason") or "").strip()
            if status == "ВР" or reason == "Вечерний резервный дежурный":
                add_participant(person)
        current_date += timedelta(days=1)

    return participants


def _apply_date_overrides(items):
    overrides = _load_date_overrides()
    overrides_changed = False
    for item in items:
        row_key = _get_assignment_key_for_item(item)
        override = overrides.get(row_key)
        if not override:
            source_start_dt = _parse_release_monitor_date(
                item.get("source_deployment_start_iso") or item.get("source_deployment_start")
            )
            source_end_dt = _parse_release_monitor_date(
                item.get("source_deployment_end_iso") or item.get("source_deployment_end")
            )
            source_sort_dt = source_start_dt or source_end_dt
            if source_start_dt:
                item["deployment_start"] = _format_release_monitor_date(source_start_dt)
                item["deployment_start_iso"] = source_start_dt.isoformat()
            else:
                item["deployment_start"] = ""
                item["deployment_start_iso"] = ""
            if source_end_dt:
                item["deployment_end"] = _format_release_monitor_date(source_end_dt)
                item["deployment_end_iso"] = source_end_dt.isoformat()
            else:
                item["deployment_end"] = ""
                item["deployment_end_iso"] = ""
            item["sort_date"] = source_sort_dt.isoformat() if source_sort_dt else ""
            if item.get("is_non_final"):
                now_dt = datetime.now()
                item["is_overdue"] = bool(
                    _is_confirmed_deployment_attempt(item)
                    and _is_release_window_expired(source_end_dt, now_dt)
                )
                item["is_today"] = _is_release_window_in_operational_day(
                    source_start_dt,
                    source_end_dt,
                    now_dt,
                )
                item["days_overdue"] = _release_days_overdue(source_end_dt, now_dt) if item.get("is_overdue") else 0
                if item.get("is_overdue"):
                    item["row_state"] = "overdue"
                elif item.get("is_today"):
                    item["row_state"] = "today"
                else:
                    item["row_state"] = "planned"
            item["has_manual_date_override"] = False
            item["manual_deployment_start"] = ""
            item["manual_deployment_end"] = ""
            continue

        manual_start_dt = _parse_release_monitor_date(override.get("start"))
        manual_end_dt = _parse_release_monitor_date(override.get("end"))
        base_start_dt = _parse_release_monitor_date(
            item.get("source_deployment_start_iso")
            or item.get("source_deployment_start")
        )
        base_end_dt = _parse_release_monitor_date(
            item.get("source_deployment_end_iso")
            or item.get("source_deployment_end")
        )

        manual_start_matches_base = _release_base_covers_manual_date(base_start_dt, manual_start_dt)
        manual_end_matches_base = _release_base_covers_manual_date(base_end_dt, manual_end_dt)
        active_manual_start = manual_start_dt if not manual_start_matches_base else None
        active_manual_end = manual_end_dt if not manual_end_matches_base else None

        if manual_start_matches_base or manual_end_matches_base:
            cleaned_override = dict(override)
            if manual_start_matches_base:
                cleaned_override["start"] = ""
            if manual_end_matches_base:
                cleaned_override["end"] = ""

            if cleaned_override.get("start") or cleaned_override.get("end"):
                overrides[row_key] = cleaned_override
                override = cleaned_override
            else:
                overrides.pop(row_key, None)
                override = {}
            overrides_changed = True

        effective_start_dt = _resolve_effective_release_date(base_start_dt, active_manual_start)
        effective_end_dt = _resolve_effective_release_date(base_end_dt, active_manual_end)

        if effective_start_dt:
            item["deployment_start"] = _format_release_monitor_date(effective_start_dt)
            item["deployment_start_iso"] = effective_start_dt.isoformat()
            item["sort_date"] = effective_start_dt.isoformat()
        if effective_end_dt:
            item["deployment_end"] = _format_release_monitor_date(effective_end_dt)
            item["deployment_end_iso"] = effective_end_dt.isoformat()

        row_is_non_final = bool(item.get("is_non_final"))
        if row_is_non_final:
            now_dt = datetime.now()
            item["is_overdue"] = bool(
                _is_confirmed_deployment_attempt(item)
                and _is_release_window_expired(effective_end_dt, now_dt)
            )
            item["is_today"] = _is_release_window_in_operational_day(
                effective_start_dt,
                effective_end_dt,
                now_dt,
            )
            item["days_overdue"] = _release_days_overdue(effective_end_dt, now_dt) if item.get("is_overdue") else 0

            if item.get("is_overdue"):
                item["row_state"] = "overdue"
            elif item.get("is_today"):
                item["row_state"] = "today"
            else:
                item["row_state"] = "planned"
        elif item.get("is_final") and (active_manual_start or active_manual_end):
            item["row_state"] = _get_final_manual_date_row_state(
                active_manual_start or effective_start_dt,
                active_manual_end or active_manual_start or effective_end_dt,
            )

        item["has_manual_date_override"] = bool(active_manual_start or active_manual_end)
        item["manual_deployment_start"] = override.get("start", "") if active_manual_start else ""
        item["manual_deployment_end"] = override.get("end", "") if active_manual_end else ""

    if overrides_changed:
        _save_date_overrides(overrides)
    return items


def _apply_reviewer_assignments(items):
    assignments = _load_reviewer_assignments()
    for item in items:
        assignment_key = _get_assignment_key_for_item(item)
        release_assignment = assignments.get(assignment_key) or assignments.get(item.get("release_key"), {})
        item["psi_owner"] = release_assignment.get("reviewer", "")
        item["psi_owner_source"] = release_assignment.get("reviewer_source", "")
        item["psi_owner_date"] = release_assignment.get("reviewer_date", "")
        item["psi_zni_reviewer"] = release_assignment.get("zni_reviewer", "")
        item["psi_checker"] = release_assignment.get("checker", "")
        item["psi_responsibles"] = list(release_assignment.get("responsibles", []))
    return items


def _get_current_week_bounds(reference_dt=None):
    reference_dt = reference_dt or datetime.now()
    week_start = reference_dt.date() - timedelta(days=reference_dt.weekday())
    week_end = week_start + timedelta(days=6)
    return week_start, week_end


def _get_release_start_date(item):
    deployment_dt = _parse_release_monitor_date(
        item.get("deployment_start_iso")
        or item.get("deployment_start")
        or item.get("sort_date")
    )
    return deployment_dt.date() if deployment_dt else None


def _has_release_responsible(item):
    responsibles = item.get("psi_responsibles") or []
    if not isinstance(responsibles, list):
        responsibles = [responsibles] if responsibles else []
    return any(str(value or "").strip() for value in responsibles)


def _is_release_assignment_relevant_for_week(item, week_start=None, week_end=None):
    week_start, week_end = (week_start, week_end) if week_start and week_end else _get_current_week_bounds()
    start_date = _get_release_start_date(item)
    if not start_date or start_date < week_start or start_date > week_end:
        return False
    if item.get("is_cancelled"):
        return False
    return True


def _apply_week_control_flags(items):
    week_start, week_end = _get_current_week_bounds()
    for item in items:
        is_week_release = _is_release_assignment_relevant_for_week(item, week_start, week_end)
        missing_responsible = bool(is_week_release and not _has_release_responsible(item))
        item["is_current_week_assignment_scope"] = is_week_release
        item["is_missing_week_responsible"] = missing_responsible
    return items


def _apply_release_week_buckets(items):
    week_start, week_end = _get_current_week_bounds()
    next_week_end = week_end + timedelta(days=7)

    for item in items:
        bucket = ""
        label = ""
        start_date = _get_release_start_date(item)

        if start_date and not item.get("is_cancelled"):
            if week_start <= start_date <= week_end:
                bucket = "current"
            elif week_end < start_date <= next_week_end:
                bucket = "next"
                label = "Следующая неделя"
            elif start_date > next_week_end:
                bucket = "future"
                label = "Дальше недели"

        item["week_bucket"] = bucket
        item["week_bucket_label"] = label

    return items


def _collect_week_candidate_availability(week_start=None, week_end=None, target_dates=None):
    week_start, week_end = (week_start, week_end) if week_start and week_end else _get_current_week_bounds()
    duty_payload = _load_duty_schedule_payload()
    availability_by_date = duty_payload.get("availability") or {}
    if target_dates is not None:
        target_dates = {
            value for value in target_dates
            if value and week_start <= value <= week_end
        }

    candidates = {
        name: {
            "name": name,
            "availability": "available",
            "reasons": [],
            "statuses": [],
            "availability_by_date": {},
            "available_dates": [],
            "reserve_dates": [],
            "excluded_dates": [],
            "is_date_dependent": False,
        }
        for name in OPLOT_VALUES
    }

    dates_to_check = sorted(target_dates) if target_dates else []
    if not dates_to_check:
        current_date = week_start
        while current_date <= week_end:
            dates_to_check.append(current_date)
            current_date += timedelta(days=1)

    for current_date in dates_to_check:
        day_people = availability_by_date.get(current_date.isoformat()) or {}
        for matched_name, entry in candidates.items():
            info = day_people.get(matched_name) or {}
            status = str((info or {}).get("status") or "").strip()
            availability = str((info or {}).get("availability") or "").strip()
            reason = str((info or {}).get("reason") or "").strip() or status
            if not availability:
                availability, reason = _classify_duty_status(status) if status else ("available", "")
            if availability not in {"available", "reserve", "excluded"}:
                availability = "available"

            date_key = current_date.isoformat()
            entry["availability_by_date"][date_key] = {
                "date": date_key,
                "status": status,
                "availability": availability,
                "reason": reason,
            }
            entry[f"{availability}_dates"].append(date_key)
            if status:
                entry["statuses"].append({
                    "date": date_key,
                    "status": status,
                    "availability": availability,
                    "reason": reason,
                })

    for entry in candidates.values():
        if entry["available_dates"]:
            entry["availability"] = "available"
        elif entry["reserve_dates"]:
            entry["availability"] = "reserve"
        else:
            entry["availability"] = "excluded"

        observed_states = {
            info.get("availability")
            for info in entry["availability_by_date"].values()
            if info.get("availability")
        }
        entry["is_date_dependent"] = len(observed_states) > 1
        entry["reasons"] = _summarize_duty_restrictions(entry["statuses"])

    grouped = {"available": [], "reserve": [], "excluded": []}
    for entry in candidates.values():
        grouped[entry["availability"]].append(entry)
    for values in grouped.values():
        values.sort(key=lambda item: item["name"])
    return grouped


def _summarize_duty_restrictions(statuses):
    normalized = []
    for status_info in statuses or []:
        date_value = _parse_release_monitor_date(status_info.get("date"))
        availability = str(status_info.get("availability") or "").strip()
        status = str(status_info.get("status") or "").strip()
        reason = str(status_info.get("reason") or "").strip()
        if not date_value or availability == "available":
            continue
        label = (status or reason or "ограничение") if availability == "reserve" else (reason or status or "ограничение")
        if availability == "reserve" and "резерв" not in _normalize_text(label):
            label = f"{label} (резерв)"
        normalized.append({
            "date": date_value.date(),
            "availability": availability,
            "label": label,
        })

    normalized.sort(key=lambda item: item["date"])
    grouped = []
    for item in normalized:
        if (
            grouped
            and grouped[-1]["availability"] == item["availability"]
            and grouped[-1]["label"] == item["label"]
            and item["date"] == grouped[-1]["end"] + timedelta(days=1)
        ):
            grouped[-1]["end"] = item["date"]
        else:
            grouped.append({
                "start": item["date"],
                "end": item["date"],
                "availability": item["availability"],
                "label": item["label"],
            })

    summaries = []
    for group in grouped:
        start_date = group["start"]
        end_date = group["end"]
        if start_date == end_date:
            date_label = start_date.strftime("%d.%m")
        elif start_date.month == end_date.month:
            date_label = f"{start_date.strftime('%d')}–{end_date.strftime('%d.%m')}"
        else:
            date_label = f"{start_date.strftime('%d.%m')}–{end_date.strftime('%d.%m')}"
        summaries.append(f"{date_label}: {group['label']}")
    return summaries


def _candidate_availability_for_release_date(candidate_groups, release_date):
    date_key = release_date.isoformat() if release_date else ""
    result = {}
    for group_name in ("available", "reserve", "excluded"):
        for candidate in (candidate_groups or {}).get(group_name, []):
            candidate_name = str(candidate.get("name") or "").strip()
            if not candidate_name:
                continue
            date_info = dict((candidate.get("availability_by_date") or {}).get(date_key) or {})
            availability = str(date_info.get("availability") or "").strip()
            if availability not in {"available", "reserve", "excluded"}:
                availability = str(candidate.get("availability") or group_name or "available").strip()
            result[candidate_name] = {
                "date": date_key,
                "availability": availability,
                "status": str(date_info.get("status") or "").strip(),
                "reason": str(date_info.get("reason") or "").strip(),
            }
    return result


def _apply_release_status_consistency(items):
    final_status = _normalize_text(FINAL_RELEASE_STATUS)
    cancelled_status = _normalize_text(CANCELLED_RELEASE_STATUS)
    pre_final_statuses = {_normalize_text(status) for status in PRE_FINAL_RELEASE_STATUSES}

    for item in items:
        normalized_status = _normalize_text(item.get("release_status"))
        is_final_status = normalized_status == final_status
        is_cancelled_status = normalized_status == cancelled_status

        if is_final_status:
            if _apply_active_reroll_schedule_state(item):
                item["is_cancelled"] = False
                continue
            item["is_final"] = True
            item["is_cancelled"] = False
            item["is_non_final"] = False
            item["is_pre_final"] = False
            item["is_ready_for_prom"] = False
            item["is_overdue"] = False
            item["is_today"] = False
            item["days_overdue"] = 0
            item["row_state"] = "final"
            continue

        if is_cancelled_status:
            item["is_final"] = False
            item["is_cancelled"] = True
            item["is_non_final"] = False
            item["is_pre_final"] = False
            item["is_ready_for_prom"] = False
            item["is_overdue"] = False
            item["is_today"] = False
            item["days_overdue"] = 0
            item["row_state"] = "cancelled"
            continue

        item["is_final"] = False
        item["is_cancelled"] = False
        item["is_non_final"] = True
        item["is_pre_final"] = normalized_status in pre_final_statuses

    return items


def _apply_release_attempt_outcomes(items):
    outcomes = _load_release_attempt_outcomes()
    changed = False
    now_text = _format_timestamp()

    for item in items:
        row_key = _get_assignment_key_for_item(item)
        if not row_key:
            continue

        has_confirmed_attempt = _is_confirmed_deployment_attempt(item)
        if (
            row_key in outcomes
            and not item.get("is_final")
            and not item.get("is_cancelled")
            and (
                not item.get("has_rov")
                or not item.get("is_overdue")
                or not has_confirmed_attempt
            )
        ):
            outcomes.pop(row_key, None)
            changed = True
            continue

        if (
            item.get("has_rov")
            and not item.get("is_final")
            and not item.get("is_cancelled")
            and item.get("is_overdue")
            and has_confirmed_attempt
        ):
            current = dict(outcomes.get(row_key) or {})
            outcomes[row_key] = {
                "state": "deferred",
                "release_key": str(item.get("release_key") or "").strip(),
                "rov_key": str(item.get("rov_key") or "").strip(),
                "detected_at": current.get("detected_at") or now_text,
                "updated_at": now_text,
            }
            changed = True

    stale_successful_attempt_keys = set()
    final_items_by_release = {}
    successful_final_release_keys = set()
    for item in items:
        if not item.get("has_rov") or not item.get("is_final"):
            continue
        release_key = str(item.get("release_key") or "").strip()
        row_key = _get_assignment_key_for_item(item)
        if not release_key or not row_key:
            continue
        final_items_by_release.setdefault(release_key, []).append(item)

    for release_items in final_items_by_release.values():
        latest_item = max(
            release_items,
            key=lambda item: (
                _sort_datetime_value(item, "sort_date"),
                _sort_datetime_value(item, "deployment_start_iso"),
                str(item.get("rov_key") or ""),
            ),
        )
        latest_key = _get_assignment_key_for_item(latest_item)
        if latest_key in outcomes:
            stale_successful_attempt_keys.add(latest_key)
        if latest_key:
            release_key = str(latest_item.get("release_key") or "").strip()
            if release_key:
                successful_final_release_keys.add(release_key)

    for row_key in stale_successful_attempt_keys:
        outcomes.pop(row_key, None)
        changed = True

    if changed:
        _save_release_attempt_outcomes(outcomes)

    deferred_keys = set(outcomes)
    for item in items:
        row_key = _get_assignment_key_for_item(item)
        if row_key not in deferred_keys:
            item["is_deferred_attempt"] = False
            item["is_deferred_resolved"] = False
            continue

        release_key = str(item.get("release_key") or "").strip()
        is_resolved_history = release_key in successful_final_release_keys
        item["is_deferred_attempt"] = True
        item["is_deferred_resolved"] = is_resolved_history
        item["is_final"] = False
        item["is_non_final"] = not is_resolved_history
        item["is_pre_final"] = False
        item["is_ready_for_prom"] = False
        item["is_overdue"] = not is_resolved_history
        item["is_today"] = False
        item["days_overdue"] = (
            0
            if is_resolved_history
            else _release_days_overdue(
                _parse_release_monitor_date(item.get("deployment_end_iso") or item.get("deployment_end"))
            )
        )
        item["row_state"] = "overdue"

    return items


def _apply_zni_assignments(items):
    assignments = _load_zni_assignments()
    flags = _load_rollout_note_flags()
    for item in items:
        assignment_key = _get_assignment_key_for_item(item)
        zni_assignment = assignments.get(assignment_key) or assignments.get(item.get("release_key"), {})
        zni_key = str(zni_assignment.get("key") or "").strip() if isinstance(zni_assignment, dict) else ""
        zni_url = str(zni_assignment.get("url") or "").strip() if isinstance(zni_assignment, dict) else ""
        item["zni_key"] = zni_key
        item["zni_url"] = zni_url
        if not str(item.get("base_zni_key") or "").strip():
            item["base_zni_key"] = zni_key
        if not str(item.get("base_zni_url") or "").strip():
            item["base_zni_url"] = zni_url
        rollout_flag_key = assignment_key if assignment_key in flags else item.get("release_key")
        rollout_note = flags.get(rollout_flag_key) or {}
        rollout_level = ""
        if isinstance(rollout_note, dict) and (
            rollout_note.get("has_rollout_notes") or str(rollout_note.get("rollout_notes_level") or "").strip()
        ):
            rollout_level = str(rollout_note.get("rollout_notes_level") or "warning").strip().lower()
            if rollout_level not in {"success", "warning", "danger", "none"}:
                rollout_level = "warning"
        item["rollout_notes_level"] = rollout_level
        item["has_rollout_notes"] = bool(rollout_level and rollout_level != "none")
    return items


def _get_release_name_ke_line(item):
    has_manual_ke_line = bool(
        item.get("manual_release")
        or item.get("manual_system_name")
        or item.get("manual_ke_id")
    )
    if has_manual_ke_line:
        system_name = str(item.get("system_name") or item.get("ke_name") or "").strip()
        ke_id = str(item.get("ke_id") or "").strip()
        if system_name and ke_id:
            return f"{system_name}({ke_id})"
        if system_name or ke_id:
            return system_name or ke_id

    lines = item.get("base_release_name_lines") or item.get("release_name_lines") or []
    if isinstance(lines, list) and len(lines) > 1:
        line_value = str(lines[1] or "").strip()
        if line_value:
            return line_value

    ke_name = str(item.get("ke_name") or "").strip()
    ke_id = str(item.get("ke_id") or "").strip()
    if ke_name and ke_id:
        return f"{ke_name}({ke_id})"
    return ke_name or ""


def _get_release_domain_from_url(url_value):
    url_value = str(url_value or "").strip()
    if not url_value:
        return ""
    if "/browse/" in url_value:
        return url_value.split("/browse/", 1)[0].rstrip("/")
    return url_value.rsplit("/", 1)[0].rstrip("/")


def _build_delta_issue_url(issue_key):
    issue_key = str(issue_key or "").strip()
    return f"{JIRA_DELTA_BASE}/browse/{issue_key}" if issue_key else ""


def _resolve_manual_zni_url(issue_key, url_value=""):
    url_value = str(url_value or "").strip()
    if not url_value or "/browse/" in url_value:
        return _build_delta_issue_url(issue_key)
    return url_value


def _apply_manual_release_overrides(items, overrides=None):
    overrides = _normalize_manual_release_overrides(overrides) if overrides is not None else _load_manual_release_overrides()
    for item in items:
        assignment_key = _get_assignment_key_for_item(item)
        override = overrides.get(assignment_key) or overrides.get(item.get("release_key"), {})
        if not isinstance(override, dict):
            override = {}

        derived_release_type = derive_release_type_from_jira(item)
        if not str(item.get("base_release_type") or "").strip():
            item["base_release_type"] = override.get("base_release_type") or derived_release_type
        if not str(item.get("release_type") or "").strip():
            item["release_type"] = item.get("base_release_type") or derived_release_type

        for field_name in BASE_OVERRIDE_FIELDS:
            base_field = _base_field_name(field_name)
            if field_name in MANUAL_OVERRIDE_DICT_FIELDS:
                if not isinstance(item.get(base_field), dict):
                    item[base_field] = (
                        _normalize_manual_dict_field(override.get(base_field))
                        or _normalize_manual_dict_field(item.get(field_name))
                    )
                continue

            if str(item.get(base_field) or "").strip():
                continue
            base_value = _normalize_manual_scalar_field(field_name, override.get(base_field))
            if not base_value:
                base_value = _normalize_manual_scalar_field(field_name, item.get(field_name))
            if field_name == "release_type" and not base_value:
                base_value = derived_release_type
            item[base_field] = base_value

        if not isinstance(item.get("base_release_name_lines"), list) or not item.get("base_release_name_lines"):
            item["base_release_name_lines"] = list(item.get("release_name_lines") or [])

        for field_name in BASE_OVERRIDE_FIELDS:
            base_field = _base_field_name(field_name)
            if field_name in MANUAL_OVERRIDE_DICT_FIELDS:
                item[field_name] = dict(item.get(base_field) or {})
            else:
                item[field_name] = str(item.get(base_field) or "").strip()

        base_name_lines = item.get("base_release_name_lines")
        if isinstance(base_name_lines, list) and base_name_lines:
            item["release_name_lines"] = list(base_name_lines)

        for field_name in MANUAL_OVERRIDE_FIELDS:
            item[_manual_field_name(field_name)] = "" if field_name in MANUAL_OVERRIDE_SCALAR_FIELDS else {}
        item["manual_clear_zni"] = False
        item["manual_overridden_fields"] = []

        if not override:
            item["has_manual_release_override"] = False
            item["has_rov"] = bool(str(item.get("rov_key") or "").strip())
            _sync_release_type_fields(item)
            continue

        changed_fields = []
        for field_name in MANUAL_OVERRIDE_SCALAR_FIELDS:
            manual_value = _normalize_manual_scalar_field(field_name, override.get(field_name))
            if not manual_value:
                continue
            if field_name == "zni_url" and not override.get("zni_key"):
                continue
            item[field_name] = manual_value
            item[_manual_field_name(field_name)] = manual_value
            changed_fields.append(field_name)

        for field_name in MANUAL_OVERRIDE_DICT_FIELDS:
            manual_value = _normalize_manual_dict_field(override.get(field_name))
            if not manual_value:
                continue
            item[field_name] = manual_value
            item[_manual_field_name(field_name)] = manual_value
            changed_fields.append(field_name)

        clear_zni = bool(override.get("clear_zni"))
        if clear_zni:
            item["zni_key"] = ""
            item["zni_url"] = ""
            changed_fields.append("clear_zni")
        elif item.get("manual_zni_key"):
            item["zni_url"] = _resolve_manual_zni_url(item.get("manual_zni_key"), item.get("manual_zni_url"))

        item["manual_clear_zni"] = clear_zni
        item["has_rov"] = bool(str(item.get("rov_key") or "").strip())
        item["manual_overridden_fields"] = sorted(set(changed_fields))
        item["has_manual_release_override"] = bool(changed_fields)
        _sync_release_type_fields(item)
    return items


def _get_assignment_key_for_item(item):
    if not isinstance(item, dict):
        return ""
    if item.get("row_key"):
        return str(item.get("row_key"))

    release_key = str(item.get("release_key") or "").strip()
    rov_key = str(item.get("rov_key") or "").strip()
    if release_key:
        return f"{release_key}::{rov_key or 'no-rov'}"
    return ""


def _normalize_text(value):
    return str(value or "").strip().lower().replace("С‘", "Рµ")


def _parse_jira_date(value):
    if not value:
        return None

    for fmt in (
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    ):
        try:
            parsed = datetime.strptime(value, fmt)
            if parsed.tzinfo is not None:
                return parsed.astimezone().replace(tzinfo=None)
            return parsed
        except ValueError:
            continue
    return None


def _normalize_cell_text(value):
    value = (value or "").replace("\xa0", " ")
    value = re.sub(r"\s*\n\s*", "\n", value)
    value = re.sub(r"[ \t]+", " ", value)
    return value.strip()


class _ConfluenceTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.tables = []
        self._table_depth = 0
        self._current_table = None
        self._current_row = None
        self._current_cell = None

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._table_depth += 1
            if self._table_depth == 1:
                self._current_table = []
            return

        if self._table_depth != 1:
            return

        if tag == "tr":
            self._current_row = []
        elif tag in {"td", "th"}:
            self._current_cell = {"text_parts": [], "links": []}
        elif tag == "a" and self._current_cell is not None:
            href = dict(attrs).get("href")
            if href:
                self._current_cell["links"].append(href)
        elif tag in {"br", "p", "div", "li"} and self._current_cell is not None:
            self._current_cell["text_parts"].append("\n")

    def handle_data(self, data):
        if self._table_depth == 1 and self._current_cell is not None:
            self._current_cell["text_parts"].append(data)

    def handle_endtag(self, tag):
        if tag == "table":
            if self._table_depth == 1 and self._current_table is not None:
                self.tables.append(self._current_table)
                self._current_table = None
            self._table_depth = max(0, self._table_depth - 1)
            return

        if self._table_depth != 1:
            return

        if tag in {"td", "th"} and self._current_row is not None and self._current_cell is not None:
            self._current_row.append(
                {
                    "text": _normalize_cell_text("".join(self._current_cell["text_parts"])),
                    "links": list(self._current_cell["links"]),
                }
            )
            self._current_cell = None
        elif tag == "tr" and self._current_table is not None and self._current_row is not None:
            if self._current_row:
                self._current_table.append(self._current_row)
            self._current_row = None


def _extract_issue_key_from_cell(cell):
    cell = cell or {}
    for href in cell.get("links", []):
        match = re.search(r"/browse/([A-Z]+-\d+)", href or "")
        if match:
            return match.group(1)

    match = re.search(r"([A-Z]+-\d+)", cell.get("text", "") or "")
    return match.group(1) if match else ""


def _match_oplot_name(raw_name):
    normalized_raw = _normalize_text(raw_name).replace(".", "")
    if not normalized_raw:
        return ""

    exact_map = {
        _normalize_text(option).replace(".", ""): option
        for option in OPLOT_VALUES
    }
    if normalized_raw in exact_map:
        return exact_map[normalized_raw]

    surname_match = re.match(r"^([\u0430-\u044f\u0451a-z-]+)\s+([\u0430-\u044f\u0451a-z])", normalized_raw, re.IGNORECASE)
    if not surname_match:
        return ""

    surname = surname_match.group(1)
    first_initial = surname_match.group(2)
    candidates = []
    for option in OPLOT_VALUES:
        normalized_option = _normalize_text(option).replace(".", "")
        option_match = re.match(r"^([\u0430-\u044f\u0451a-z-]+)\s+([\u0430-\u044f\u0451a-z])", normalized_option, re.IGNORECASE)
        if option_match and option_match.group(1) == surname and option_match.group(2) == first_initial:
            candidates.append(option)

    return candidates[0] if len(candidates) == 1 else ""


def _parse_confluence_assignment_cell(cell_text):
    responsibles = []
    checker_lines = []
    mode = "responsibles"

    lines = [
        line.strip(" /")
        for line in re.split(r"[\n\r]+", cell_text or "")
        if line.strip(" /")
    ]

    for line in lines:
        normalized_line = _normalize_text(line)

        if normalized_line.startswith("\u043f\u0440\u043e\u0432\u0435\u0440\u044f\u0435\u0442"):
            mode = "checker"
            remainder = line.split(":", 1)[1].strip() if ":" in line else ""
            if remainder and "\u043e\u0442\u0441\u0443\u0442\u0441\u0442\u0432" not in _normalize_text(remainder):
                checker_lines.append(remainder)
            continue

        if normalized_line.startswith("\u0443\u0441\u0442\u0430\u043d\u0430\u0432\u043b\u0438\u0432\u0430\u0435\u0442"):
            mode = "ignore"
            continue

        if "\u043f\u0440\u043e\u0432\u0435\u0440\u043a\u0438 \u043e\u0442\u0441\u0443\u0442\u0441\u0442\u0432" in normalized_line:
            checker_lines = []
            mode = "ignore"
            continue

        if mode == "responsibles":
            for part in [chunk.strip(" /") for chunk in line.split("/")]:
                if part:
                    matched_name = _match_oplot_name(part)
                    if matched_name and matched_name not in responsibles:
                        responsibles.append(matched_name)
        elif mode == "checker":
            checker_lines.append(line.strip())

    checker = " ".join(checker_lines).strip()
    return responsibles, checker


def _find_release_assignment_table(storage_html):
    parser = _ConfluenceTableParser()
    parser.feed(storage_html or "")

    for table in parser.tables:
        if not table:
            continue
        headers = [_normalize_text(cell.get("text", "")) for cell in table[0]]
        if (
            any("id \u0440\u0435\u043b\u0438\u0437\u0430" in header for header in headers)
            and any("id \u0440\u0430\u0441\u043f\u043e\u0440\u044f\u0436\u0435\u043d\u0438\u044f" in header for header in headers)
            and any("\u043e\u0442\u0432\u0435\u0442\u0441\u0442\u0432\u0435\u043d\u043d\u044b\u0439" in header for header in headers)
        ):
            return table
    return []


def _get_confluence_release_page_id(year):
    # Публикуем только в одну заранее зафиксированную страницу релизов.
    return "18369778404"


def _fetch_confluence_release_page(year):
    page_id = _get_confluence_release_page_id(year)
    token = str(TOKENS.get("confluence_delta_token", "") or "").strip()
    if not page_id:
        raise ValueError(f"Не настроен pageId Confluence для {year} года")
    if not token:
        raise ValueError("Не настроен token доступа к Confluence")

    url = f"{CONFLUENCE_DELTA_BASE}/rest/api/content/{page_id}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    params = {"expand": "body.storage,version,title"}
    response = requests.get(url, headers=headers, params=params, verify=False, timeout=60)
    if not response.ok:
        detail = _extract_confluence_error_detail(response)
        raise ValueError(f"Confluence GET failed ({response.status_code}): {detail}")
    data = response.json()
    return {
        "page_id": page_id,
        "title": str(data.get("title") or "").strip(),
        "version": int(((data.get("version") or {}).get("number")) or 0),
        "storage_html": (((data.get("body") or {}).get("storage") or {}).get("value") or ""),
        "raw": data,
    }


def _extract_confluence_error_detail(response):
    try:
        payload = response.json()
        if isinstance(payload, dict):
            for key in ("message", "errorMessage", "reason"):
                value = str(payload.get(key) or "").strip()
                if value:
                    return value
            errors = payload.get("errors")
            if isinstance(errors, dict) and errors:
                return "; ".join(f"{key}: {value}" for key, value in errors.items())
            if isinstance(errors, list) and errors:
                return "; ".join(str(item) for item in errors if str(item).strip())
            return str(payload)[:1000]
    except Exception:
        pass

    body = str(getattr(response, "text", "") or "").strip()
    if body:
        return body[:1000]
    return str(getattr(response, "reason", "") or "empty response")


def _render_confluence_release_link(url, text):
    label = html.escape(str(text or "").strip()) or html.escape("—")
    href = str(url or "").strip()
    if not href:
        return label
    return f'<a href="{html.escape(href, quote=True)}">{label}</a>'


def _render_confluence_release_name_lines(item):
    lines = [str(line or "").strip() for line in (item.get("release_name_lines") or []) if str(line or "").strip()]
    if not lines:
        fallback = str(item.get("release_summary") or "").strip()
        if fallback:
            lines = [fallback]
    if not lines:
        return html.escape("—")
    return "".join(f"<div>{html.escape(line)}</div>" for line in lines)


def _render_confluence_release_name_cell(item):
    lines = [str(line or "").strip() for line in (item.get("release_name_lines") or []) if str(line or "").strip()]
    if not lines:
        fallback = str(item.get("release_summary") or "").strip()
        if fallback:
            lines = [fallback]

    build_version = str(item.get("release_version") or "").strip()
    build_url = str(item.get("release_dist_url") or "").strip()
    build_line = ""
    if build_version:
        build_label = html.escape(build_version)
        if build_url:
            build_line = (
                '<div class="release-name-line">'
                f'сборка: <a href="{html.escape(build_url, quote=True)}">{build_label}</a>'
                '</div>'
            )
        else:
            build_line = f'<div class="release-name-line">сборка: {build_label}</div>'

    if not lines and not build_line:
        return html.escape("—")

    rendered_lines = "".join(f"<div>{html.escape(line)}</div>" for line in lines)
    return rendered_lines + build_line


def _render_confluence_release_assignment_cell(item):
    responsibles = [
        str(value or "").strip()
        for value in (item.get("psi_responsibles") or [])
        if str(value or "").strip()
    ]
    checker = str(item.get("psi_checker") or "").strip()

    cell_lines = []
    if responsibles:
        cell_lines.append(" / ".join(html.escape(name) for name in responsibles))
    else:
        cell_lines.append(html.escape("Отсутствует"))

    if checker:
        cell_lines.append(f"Проверяет:<br/>{html.escape(checker)}")
    else:
        cell_lines.append("Проверяет:<br/>Отсутствует")

    return "<br/>".join(cell_lines)


def _parse_release_monitor_date_value(value):
    raw_value = str(value or "").strip()
    if not raw_value:
        return None

    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw_value, fmt)
        except ValueError:
            continue

    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw_value[:19], fmt)
        except ValueError:
            continue

    return None


def _is_release_far_future(item):
    deployment_date = (
        _parse_release_monitor_date_value(item.get("deployment_start"))
        or _parse_release_monitor_date_value(item.get("deployment_start_iso"))
        or _parse_release_monitor_date_value(item.get("source_deployment_start"))
        or _parse_release_monitor_date_value(item.get("source_deployment_start_iso"))
    )
    if not deployment_date:
        return False

    today = datetime.now().date()
    threshold = today + timedelta(days=10)
    return deployment_date.date() > threshold


def _render_confluence_release_row(item):
    row_bg = ""
    rollout_level = str(item.get("rollout_notes_level") or ("warning" if item.get("has_rollout_notes") else "")).strip().lower()
    if rollout_level == "success":
        row_bg = "#eaf7ef"
    elif rollout_level == "danger":
        row_bg = "#fdebec"
    elif rollout_level == "warning":
        row_bg = "#fff7db"
    elif rollout_level == "none":
        row_bg = ""
    row_style = f' style="background-color: {row_bg};"' if row_bg else ""

    row_number = str(item.get("release_number") or "").strip() or "—"
    zni_key = str(item.get("zni_key") or "").strip()
    ke_value = str(item.get("ke") or "").strip() or "—"
    release_key = str(item.get("release_key") or "").strip()
    rov_key = str(item.get("rov_key") or "").strip()
    start_date = str(item.get("deployment_start") or "").strip() or "—"
    end_date = str(item.get("deployment_end") or "").strip() or "—"

    cells = [
        f"<td>{html.escape(row_number)}</td>",
        f"<td>{_render_confluence_release_name_cell(item)}</td>",
        f"<td>{_render_confluence_release_link(item.get('zni_url'), zni_key) if zni_key else html.escape('—')}</td>",
        f"<td>{html.escape(ke_value)}</td>",
        f"<td>{_render_confluence_release_link(item.get('release_url'), release_key)}</td>",
        f"<td>{_render_confluence_release_link(item.get('rov_url'), rov_key) if rov_key else html.escape('—')}</td>",
        f"<td>{html.escape(start_date)}</td>",
        f"<td>{html.escape(end_date)}</td>",
        f"<td>{_render_confluence_release_assignment_cell(item)}</td>",
    ]
    return f"<tr{row_style}>{''.join(cells)}</tr>"


def _build_confluence_release_table(items, year):
    rows = []
    for item in items or []:
        if int(item.get("year", 0) or 0) != int(year or 0):
            continue
        if item.get("is_unnumbered"):
            continue
        if _is_release_far_future(item):
            continue
        rows.append(_render_confluence_release_row(item))

    table_rows = "".join(rows)
    return f"""
<table data-layout="wide" style="width: 100%; border-collapse: collapse; table-layout: fixed;">
    <colgroup>
        <col style="width: 5%;" />
        <col style="width: 28%;" />
        <col style="width: 9%;" />
        <col style="width: 10%;" />
        <col style="width: 12%;" />
        <col style="width: 12%;" />
        <col style="width: 8%;" />
        <col style="width: 8%;" />
        <col style="width: 18%;" />
    </colgroup>
    <thead>
        <tr>
            <th style="border: 1px solid #d9e2f2; background: #f6f8fc; padding: 8px 6px; text-align: left;">№ релиза</th>
            <th style="border: 1px solid #d9e2f2; background: #f6f8fc; padding: 8px 6px; text-align: left;">Название релиза</th>
            <th style="border: 1px solid #d9e2f2; background: #f6f8fc; padding: 8px 6px; text-align: left;">№ ЗНИ</th>
            <th style="border: 1px solid #d9e2f2; background: #f6f8fc; padding: 8px 6px; text-align: left;">КЭ</th>
            <th style="border: 1px solid #d9e2f2; background: #f6f8fc; padding: 8px 6px; text-align: left;">ID релиза</th>
            <th style="border: 1px solid #d9e2f2; background: #f6f8fc; padding: 8px 6px; text-align: left;">ID распоряжения</th>
            <th style="border: 1px solid #d9e2f2; background: #f6f8fc; padding: 8px 6px; text-align: left;">Дата начала внедрения</th>
            <th style="border: 1px solid #d9e2f2; background: #f6f8fc; padding: 8px 6px; text-align: left;">Дата окончания внедрения</th>
            <th style="border: 1px solid #d9e2f2; background: #f6f8fc; padding: 8px 6px; text-align: left;">Ответственный за ПСИ/Проверки</th>
        </tr>
    </thead>
    <tbody>
        {table_rows}
    </tbody>
</table>
""".strip()


def _replace_release_monitor_table_in_storage(storage_html, replacement_table_html):
    if not storage_html:
        return replacement_table_html, False

    table_pattern = re.compile(r"<table\b[^>]*>.*?</table>", re.IGNORECASE | re.DOTALL)
    for match in table_pattern.finditer(storage_html):
        if _find_release_assignment_table(match.group(0)):
            updated_html = storage_html[:match.start()] + replacement_table_html + storage_html[match.end():]
            return updated_html, True
    return storage_html, False


def _load_confluence_release_assignments(year):
    page_id = "18369778404"
    token = str(TOKENS.get("confluence_delta_token", "") or "").strip()
    if not page_id:
        raise ValueError(f"РќРµ РЅР°СЃС‚СЂРѕРµРЅ pageId Confluence РґР»СЏ {year} РіРѕРґР°")
    if not token:
        raise ValueError("РќРµ РЅР°СЃС‚СЂРѕРµРЅ С‚РѕРєРµРЅ РґРѕСЃС‚СѓРїР° Рє Confluence")

    url = f"{CONFLUENCE_DELTA_BASE}/rest/api/content/{page_id}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    params = {"expand": "body.storage,version"}
    response = requests.get(url, headers=headers, params=params, verify=False, timeout=60)
    response.raise_for_status()
    data = response.json()
    storage_html = (((data.get("body") or {}).get("storage") or {}).get("value") or "")
    table = _find_release_assignment_table(storage_html)
    if not table:
        raise ValueError("РќРµ СѓРґР°Р»РѕСЃСЊ РЅР°Р№С‚Рё С‚Р°Р±Р»РёС†Сѓ СЂРµР»РёР·РѕРІ РЅР° СЃС‚СЂР°РЅРёС†Рµ Confluence")

    assignments = {}
    for row in table[1:]:
        if len(row) < 9:
            continue
        release_key = _extract_issue_key_from_cell(row[4])
        rov_key = _extract_issue_key_from_cell(row[5])
        if not release_key:
            continue

        responsibles, checker = _parse_confluence_assignment_cell(row[8].get("text", ""))
        if not responsibles and not checker:
            continue

        row_key = f"{release_key}::{rov_key or 'no-rov'}"
        assignments[row_key] = {
            "release_key": release_key,
            "rov_key": rov_key,
            "responsibles": responsibles,
            "checker": checker,
        }

    return assignments


def _format_timestamp(dt=None):
    return (dt or datetime.now()).strftime("%d.%m.%Y %H:%M:%S")


def _extract_field_value(raw_value):
    if raw_value is None:
        return None
    if isinstance(raw_value, list):
        values = []
        for item in raw_value:
            if isinstance(item, dict):
                values.append(item.get("value") or item.get("name") or item.get("id"))
            else:
                values.append(str(item))
        return ", ".join(str(v) for v in values if v)
    if isinstance(raw_value, dict):
        return raw_value.get("value") or raw_value.get("name") or raw_value.get("id")
    return raw_value


def _first_list_item(raw_value):
    if isinstance(raw_value, list) and raw_value:
        return raw_value[0]
    if isinstance(raw_value, dict):
        return raw_value
    return None


def _format_ke_id(raw_ke_id):
    if not raw_ke_id:
        return ""

    digits = re.sub(r"\D", "", str(raw_ke_id))
    if not digits:
        return ""
    return f"CI{digits.zfill(8)}"


def _extract_version(dist_item):
    if not dist_item:
        return ""

    if isinstance(dist_item, dict):
        for key in ("version", "buildVersion"):
            value = dist_item.get(key)
            if value:
                return str(value)

        for key in ("value", "url"):
            value = dist_item.get(key)
            if not value:
                continue
            match = RELEASE_VERSION_PATTERN.search(str(value))
            if match:
                return match.group(0)
    else:
        match = RELEASE_VERSION_PATTERN.search(str(dist_item))
        if match:
            return match.group(0)

    return ""


def _iter_nested_values(value):
    if isinstance(value, dict):
        yield value
        for nested_value in value.values():
            yield from _iter_nested_values(nested_value)
    elif isinstance(value, list):
        for nested_item in value:
            yield from _iter_nested_values(nested_item)
    elif value is not None:
        yield value


def _extract_nested_version(dist_item):
    version = extract_distribution_version(dist_item)
    if version:
        return version

    for value in _iter_nested_values(dist_item):
        if isinstance(value, dict):
            for key in ("version", "buildVersion"):
                raw_value = value.get(key)
                if raw_value and RELEASE_VERSION_PATTERN.search(str(raw_value)):
                    return str(raw_value)
        else:
            match = RELEASE_VERSION_PATTERN.search(str(value))
            if match:
                return match.group(0)
    return _extract_version(dist_item)


def _extract_dist_url(dist_item):
    if not dist_item:
        return ""

    candidate_values = []
    if isinstance(dist_item, dict):
        for key in ("url", "value", "downloadUrl", "artifactUrl", "link"):
            value = dist_item.get(key)
            if value:
                candidate_values.append(str(value))
    else:
        candidate_values.append(str(dist_item))

    for value in candidate_values:
        match = ARTIFACT_URL_PATTERN.search(value)
        if match:
            return _normalize_artifact_url(match.group(0).rstrip('",)'))
        if value.startswith("http://") or value.startswith("https://"):
            return value

    return ""


def _normalize_artifact_url(url_value):
    url_value = str(url_value or "").strip()
    if not url_value:
        return ""
    if url_value.startswith(("http://", "https://")):
        return url_value
    return f"https://{url_value}"


def _extract_nested_dist_url(dist_item):
    artifact_url = extract_artifact_url(dist_item)
    if artifact_url:
        return _normalize_artifact_url(artifact_url)

    for value in _iter_nested_values(dist_item):
        if isinstance(value, dict):
            for key in ("url", "downloadUrl", "artifactUrl", "link", "value"):
                raw_value = value.get(key)
                if not raw_value:
                    continue
                match = ARTIFACT_URL_PATTERN.search(str(raw_value))
                if match:
                    return _normalize_artifact_url(match.group(0).rstrip('",)'))
                if str(raw_value).startswith(("http://", "https://")):
                    return str(raw_value)
        else:
            match = ARTIFACT_URL_PATTERN.search(str(value))
            if match:
                return _normalize_artifact_url(match.group(0).rstrip('",)'))
    return _extract_dist_url(dist_item)


def _extract_ke_object(fields, resolved_fields):
    raw_ke_object = fields.get(resolved_fields["ke_object"])
    item = _first_list_item(raw_ke_object)
    if isinstance(item, dict):
        return {
            "id": item.get("id") or item.get("smId") or "",
            "name": item.get("value") or item.get("name") or "",
        }

    raw_delta_object = fields.get(resolved_fields["delta_release_distributive"])
    item = _first_list_item(raw_delta_object)
    if isinstance(item, dict):
        parent_ci = item.get("PARENT_CI") or item.get("id") or ""
        return {
            "id": parent_ci[2:].lstrip("0") if str(parent_ci).startswith("CI") else str(parent_ci),
            "name": item.get("value") or item.get("name") or "",
        }

    return {"id": "", "name": ""}


def _build_ai_agent_release_context(fields, resolved_fields, summary="", system_info_text="", ke_object=None):
    context = [summary, system_info_text, ke_object or {}]
    raw_ke_object = fields.get(resolved_fields["ke_object"])
    if raw_ke_object:
        context.append(raw_ke_object)

    for link in fields.get("issuelinks") or []:
        linked_issue = link.get("inwardIssue") or link.get("outwardIssue") or {}
        linked_fields = linked_issue.get("fields") or {}
        linked_summary = linked_fields.get("summary")
        if linked_summary:
            context.append(linked_summary)

    for subtask in fields.get("subtasks") or []:
        subtask_fields = subtask.get("fields") or {}
        subtask_summary = subtask_fields.get("summary")
        if subtask_summary:
            context.append(subtask_summary)

    return context


def _extract_release_dist(fields, resolved_fields, release_context=None):
    candidates = []
    for logical_name in ("release_distributive", "delta_release_distributive"):
        raw_dist = fields.get(resolved_fields[logical_name])
        if isinstance(raw_dist, list):
            candidates.extend(item for item in raw_dist if item)
        elif raw_dist:
            candidates.append(raw_dist)
    if candidates:
        if is_ai_agent_release_context(release_context):
            selected = select_distribution_artifact(
                candidates,
                allow_image_artifact=True,
                release_context=release_context,
            )
        else:
            selected = select_distribution_artifact(candidates, release_context=release_context)
        if selected:
            return selected

        flat_candidates = flatten_artifact_candidates(candidates)
        if any(classify_artifact_entry(item) == "image" for item in flat_candidates):
            logging.warning(
                "Release monitor distribution artifact not found: only image/unknown artifacts were provided"
            )
    return None


def _get_domain_groups():
    groups = {}
    for prefix in RELEASE_PREFIXES:
        domain, token = get_jira_domain_and_token(f"{prefix}-1")
        groups.setdefault((domain, token), []).append(prefix)
    return groups


def _fetch_field_name_map(domain, token):
    cache_key = domain
    cached = _field_map_cache.get(cache_key)
    if cached:
        return cached

    url = f"{domain}/rest/api/2/field"
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}

    try:
        response = requests.get(url, headers=headers, verify=False, timeout=30)
        response.raise_for_status()
        field_data = response.json()
        field_map = {item.get("id"): item.get("name", "") for item in field_data}
        _field_map_cache[cache_key] = field_map
        return field_map
    except Exception as exc:
        logging.warning("Release monitor: failed to fetch field map from %s: %s", domain, exc)
        field_map = {}
        _field_map_cache[cache_key] = field_map
        return field_map


def _resolve_field_ids(domain, token):
    field_name_map = _fetch_field_name_map(domain, token)
    resolved = {}

    for logical_name, fallback_id in FIELD_FALLBACKS.items():
        resolved[logical_name] = fallback_id
        aliases = FIELD_ALIASES.get(logical_name, ())
        if not aliases:
            continue

        normalized_aliases = [_normalize_text(alias) for alias in aliases if alias]

        for field_id, field_name in field_name_map.items():
            normalized_name = _normalize_text(field_name)
            if normalized_name in normalized_aliases:
                resolved[logical_name] = field_id
                break
        else:
            for field_id, field_name in field_name_map.items():
                normalized_name = _normalize_text(field_name)
                if any(alias in normalized_name for alias in normalized_aliases):
                    resolved[logical_name] = field_id
                    break

    return resolved


def _remaining_refresh_seconds(deadline_at):
    if deadline_at is None:
        return None
    remaining = float(deadline_at) - time.monotonic()
    if remaining <= 0:
        raise TimeoutError("Release monitor refresh deadline exceeded")
    return remaining


def _sleep_for_retry(delay_seconds, deadline_at):
    remaining = _remaining_refresh_seconds(deadline_at)
    delay = float(delay_seconds or 0)
    if remaining is not None:
        delay = min(delay, remaining)
    if delay > 0:
        time.sleep(delay)


def _execute_search(
    domain,
    token,
    jql,
    fields_to_load,
    *,
    retry_server_errors=False,
    retry_label="",
    retry_max_attempts=None,
    retry_delay_seconds=None,
    deadline_at=None,
    strict_pagination=False,
    source_ledger=None,
    source_kind="jira_search",
):
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    url = f"{domain}/rest/api/2/search"
    start_at = 0
    issues = []
    max_attempts = max(
        1,
        int(
            retry_max_attempts
            if retry_max_attempts is not None
            else (RELIABLE_SEARCH_MAX_ATTEMPTS if retry_server_errors else 1)
        ),
    )
    retry_delay = float(
        retry_delay_seconds
        if retry_delay_seconds is not None
        else RELIABLE_SEARCH_RETRY_DELAY_SECONDS
    )

    ledger_entry = {
        "kind": source_kind,
        "label": retry_label or jql,
        "domain": domain,
        "status": "running",
        "expected_total": None,
        "fetched_total": 0,
        "pages": 0,
    }

    try:
        while True:
            remaining = _remaining_refresh_seconds(deadline_at)
            params = {
                "jql": jql,
                "startAt": start_at,
                "maxResults": 100,
                "fields": ",".join(sorted(field for field in fields_to_load if field)),
            }
            attempt = 1
            while True:
                try:
                    request_timeout = 60 if remaining is None else max(1, min(60, int(remaining)))
                    response = requests.get(
                        url,
                        headers=headers,
                        params=params,
                        verify=False,
                        timeout=request_timeout,
                    )
                    if (
                        retry_server_errors
                        and response.status_code in RELIABLE_SEARCH_RETRY_STATUS_CODES
                        and attempt < max_attempts
                    ):
                        logging.warning(
                            "Release monitor refresh: Jira search failed with %s for %s, retry %s/%s",
                            response.status_code,
                            retry_label or jql,
                            attempt + 1,
                            max_attempts,
                        )
                        _sleep_for_retry(retry_delay, deadline_at)
                        attempt += 1
                        remaining = _remaining_refresh_seconds(deadline_at)
                        continue
                    response.raise_for_status()
                    break
                except (requests.ConnectionError, requests.Timeout) as exc:
                    if retry_server_errors and attempt < max_attempts:
                        logging.warning(
                            "Release monitor refresh: Jira search transient error for %s: %s, retry %s/%s",
                            retry_label or jql,
                            exc,
                            attempt + 1,
                            max_attempts,
                        )
                        _sleep_for_retry(retry_delay, deadline_at)
                        attempt += 1
                        remaining = _remaining_refresh_seconds(deadline_at)
                        continue
                    raise

            data = response.json()
            batch = data.get("issues", [])
            if not isinstance(batch, list):
                raise ValueError(f"Jira search returned invalid issues payload for {retry_label or jql}")
            try:
                expected_total = int(data.get("total", 0))
            except (TypeError, ValueError):
                raise ValueError(f"Jira search returned invalid total for {retry_label or jql}")

            if ledger_entry["expected_total"] is None:
                ledger_entry["expected_total"] = expected_total
            elif strict_pagination and expected_total != ledger_entry["expected_total"]:
                raise RuntimeError(
                    f"Jira pagination total changed for {retry_label or jql}: "
                    f"{ledger_entry['expected_total']} -> {expected_total}"
                )

            issues.extend(batch)
            ledger_entry["pages"] += 1
            ledger_entry["fetched_total"] = len(issues)
            if len(issues) >= expected_total:
                break
            if not batch:
                if strict_pagination:
                    raise RuntimeError(
                        f"Incomplete Jira pagination for {retry_label or jql}: "
                        f"fetched {len(issues)} of {expected_total}"
                    )
                break
            start_at += len(batch)

        if strict_pagination and len(issues) != int(ledger_entry["expected_total"] or 0):
            raise RuntimeError(
                f"Incomplete Jira pagination for {retry_label or jql}: "
                f"fetched {len(issues)} of {ledger_entry['expected_total']}"
            )
        ledger_entry["status"] = "success"
        return issues
    except Exception as exc:
        ledger_entry["status"] = "failed"
        ledger_entry["error"] = str(exc)
        raise
    finally:
        if isinstance(source_ledger, list):
            source_ledger.append(dict(ledger_entry))


def _execute_release_search(
    domain,
    token,
    prefix,
    year_from,
    fields_to_load,
    *,
    retry_server_errors=False,
    retry_max_attempts=None,
    retry_delay_seconds=None,
    deadline_at=None,
    strict_pagination=False,
    source_ledger=None,
):
    current_year = datetime.now().year
    jql = (
        f'project = {prefix} AND '
        f'issuetype = "{RELEASE_ISSUE_TYPE}" AND '
        f'created >= "{year_from}-01-01" AND '
        f'created < "{current_year + 1}-01-01" '
        f'ORDER BY created ASC, key ASC'
    )
    return _execute_search(
        domain,
        token,
        jql,
        fields_to_load,
        retry_server_errors=retry_server_errors,
        retry_label=f"{domain} {prefix} releases",
        retry_max_attempts=retry_max_attempts,
        retry_delay_seconds=retry_delay_seconds,
        deadline_at=deadline_at,
        strict_pagination=strict_pagination,
        source_ledger=source_ledger,
        source_kind="release_prefix",
    )


def _execute_quick_release_search(domain, token, prefix, updated_since, fields_to_load):
    current_year = datetime.now().year
    updated_since_str = updated_since.strftime("%Y-%m-%d")
    jql = (
        f'project = {prefix} AND '
        f'issuetype = "{RELEASE_ISSUE_TYPE}" AND '
        f'updated >= "{updated_since_str}" AND '
        f'created < "{current_year + 1}-01-01" '
        f'ORDER BY updated DESC, key ASC'
    )
    return _execute_search(domain, token, jql, fields_to_load)


def _execute_incremental_release_search(
    domain,
    token,
    prefix,
    lookback_minutes,
    fields_to_load,
    *,
    retry_server_errors=False,
    retry_max_attempts=None,
    retry_delay_seconds=None,
):
    current_year = datetime.now().year
    previous_year = current_year - 1
    lookback_minutes = max(1, int(lookback_minutes or AUTO_INCREMENTAL_REFRESH_LOOKBACK_MINUTES))
    jql = (
        f'project = {prefix} AND '
        f'issuetype = "{RELEASE_ISSUE_TYPE}" AND '
        f'updated >= -{lookback_minutes}m AND '
        f'created >= "{previous_year}-01-01" AND '
        f'created < "{current_year + 1}-01-01" '
        f'ORDER BY updated DESC, key ASC'
    )
    return _execute_search(
        domain,
        token,
        jql,
        fields_to_load,
        retry_server_errors=retry_server_errors,
        retry_label=f"{domain} {prefix} incremental releases",
        retry_max_attempts=retry_max_attempts,
        retry_delay_seconds=retry_delay_seconds,
    )


def _execute_incremental_rov_search(
    domain,
    token,
    prefix,
    lookback_minutes,
    fields_to_load,
    *,
    retry_server_errors=False,
    retry_max_attempts=None,
    retry_delay_seconds=None,
):
    current_year = datetime.now().year
    previous_year = current_year - 1
    lookback_minutes = max(1, int(lookback_minutes or AUTO_INCREMENTAL_REFRESH_LOOKBACK_MINUTES))
    jql = (
        f'project = {prefix} AND '
        f'issuetype = "{ROV_ISSUE_TYPE}" AND '
        f'updated >= -{lookback_minutes}m AND '
        f'created >= "{previous_year}-01-01" AND '
        f'created < "{current_year + 1}-01-01" '
        f'ORDER BY updated DESC, key ASC'
    )
    return _execute_search(
        domain,
        token,
        jql,
        fields_to_load,
        retry_server_errors=retry_server_errors,
        retry_label=f"{domain} {prefix} incremental ROV",
        retry_max_attempts=retry_max_attempts,
        retry_delay_seconds=retry_delay_seconds,
    )


def _execute_issue_keys_search(
    domain,
    token,
    issue_keys,
    fields_to_load,
    *,
    retry_server_errors=False,
    retry_max_attempts=None,
    retry_delay_seconds=None,
    deadline_at=None,
    strict_pagination=False,
    source_ledger=None,
):
    issues = []
    for offset in range(0, len(issue_keys), 50):
        batch_keys = issue_keys[offset: offset + 50]
        quoted = ", ".join(f'"{key}"' for key in batch_keys)
        jql = f"key in ({quoted})"
        issues.extend(
            _execute_search(
                domain,
                token,
                jql,
                fields_to_load,
                retry_server_errors=retry_server_errors,
                retry_label=f"{domain} issue batch {offset // 50 + 1}",
                retry_max_attempts=retry_max_attempts,
                retry_delay_seconds=retry_delay_seconds,
                deadline_at=deadline_at,
                strict_pagination=strict_pagination,
                source_ledger=source_ledger,
                source_kind="linked_rov_batch",
            )
        )
    return issues


def _execute_issue_get(
    domain,
    token,
    issue_key,
    fields_to_load,
    *,
    retry_server_errors=False,
    retry_max_attempts=None,
    retry_delay_seconds=None,
    retry_label="",
    deadline_at=None,
):
    issue_key = str(issue_key or "").strip()
    if not issue_key:
        return None

    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    url = f"{domain}/rest/api/2/issue/{issue_key}"
    params = {
        "fields": ",".join(sorted(field for field in fields_to_load if field)),
    }
    max_attempts = max(
        1,
        int(
            retry_max_attempts
            if retry_max_attempts is not None
            else (RELIABLE_SEARCH_MAX_ATTEMPTS if retry_server_errors else 1)
        ),
    )
    retry_delay = float(
        retry_delay_seconds
        if retry_delay_seconds is not None
        else RELIABLE_SEARCH_RETRY_DELAY_SECONDS
    )

    attempt = 1
    while True:
        try:
            remaining = _remaining_refresh_seconds(deadline_at)
            request_timeout = 60 if remaining is None else max(1, min(60, int(remaining)))
            response = requests.get(
                url,
                headers=headers,
                params=params,
                verify=False,
                timeout=request_timeout,
            )
            if (
                retry_server_errors
                and response.status_code in RELIABLE_SEARCH_RETRY_STATUS_CODES
                and attempt < max_attempts
            ):
                logging.warning(
                    "Release monitor: direct issue check failed with %s for %s, retry %s/%s",
                    response.status_code,
                    retry_label or issue_key,
                    attempt + 1,
                    max_attempts,
                )
                _sleep_for_retry(retry_delay, deadline_at)
                attempt += 1
                continue
            response.raise_for_status()
            return response.json()
        except (requests.ConnectionError, requests.Timeout) as exc:
            if retry_server_errors and attempt < max_attempts:
                logging.warning(
                    "Release monitor: direct issue check transient error for %s: %s, retry %s/%s",
                    retry_label or issue_key,
                    exc,
                    attempt + 1,
                    max_attempts,
                )
                _sleep_for_retry(retry_delay, deadline_at)
                attempt += 1
                continue
            raise


def _release_monitor_release_fields_to_load(resolved_fields):
    return {
        "key",
        "summary",
        "status",
        "resolution",
        "assignee",
        "reporter",
        "created",
        "updated",
        "issuetype",
        "priority",
        "issuelinks",
        resolved_fields["planned_prom_start"],
        resolved_fields["planned_prom_end"],
        resolved_fields["system_info"],
        resolved_fields["ke_object"],
        resolved_fields["release_distributive"],
        resolved_fields["delta_release_distributive"],
    }


def _release_monitor_rov_fields_to_load(resolved_fields):
    return {
        "key",
        "summary",
        "status",
        "created",
        "updated",
        "issuetype",
        "issuelinks",
        resolved_fields["rov_start"],
        resolved_fields["rov_end"],
    }


def _extract_release_io_keys(issue):
    keys = []
    for link in issue.get("fields", {}).get("issuelinks", []):
        link_type = link.get("type", {})
        linked_issues = []
        if link.get("inwardIssue"):
            linked_issues.append(link.get("inwardIssue"))
        if link.get("outwardIssue"):
            linked_issues.append(link.get("outwardIssue"))

        link_name = str(link_type.get("name") or "")
        inward_name = str(link_type.get("inward") or "")
        outward_name = str(link_type.get("outward") or "")
        is_release_io = (
            link_name == "ReleaseIO"
            or "Introduction Order" in inward_name
            or "Introduction Order" in outward_name
        )
        if not is_release_io:
            continue

        for linked_issue in linked_issues:
            key = linked_issue.get("key")
            if key:
                keys.append(key)

    def _sort_key(issue_key):
        match = re.search(r"-(\d+)$", issue_key or "")
        return int(match.group(1)) if match else -1

    return sorted(list(dict.fromkeys(keys)), key=_sort_key)


def _extract_release_io_key(issue):
    keys = _extract_release_io_keys(issue)
    return keys[-1] if keys else ""


def _issue_type_name(issue):
    return str(((issue.get("fields") or {}).get("issuetype") or {}).get("name") or "")


def _extract_linked_release_keys_from_rov(issue):
    release_keys = []
    for link in (issue.get("fields") or {}).get("issuelinks", []):
        linked_issues = []
        if link.get("inwardIssue"):
            linked_issues.append(link.get("inwardIssue"))
        if link.get("outwardIssue"):
            linked_issues.append(link.get("outwardIssue"))

        link_type = link.get("type", {})
        link_name = str(link_type.get("name") or "")
        inward_name = str(link_type.get("inward") or "")
        outward_name = str(link_type.get("outward") or "")
        is_release_io = (
            link_name == "ReleaseIO"
            or "Introduction Order" in inward_name
            or "Introduction Order" in outward_name
        )
        if not is_release_io:
            continue

        for linked_issue in linked_issues:
            key = str(linked_issue.get("key") or "").strip()
            if not key:
                continue
            linked_type = _issue_type_name(linked_issue)
            if linked_type and linked_type != RELEASE_ISSUE_TYPE:
                continue
            release_keys.append(key)

    return sorted(dict.fromkeys(release_keys))


def _split_release_monitor_row_key(row_key):
    raw_key = str(row_key or "").strip()
    if "::" not in raw_key:
        return "", ""
    release_key, rov_key = raw_key.split("::", 1)
    release_key = release_key.strip()
    rov_key = rov_key.strip()
    if not release_key or not rov_key or rov_key == "no-rov":
        return release_key, ""
    return release_key, rov_key


def _is_real_rov_key(value):
    raw_value = str(value or "").strip()
    return bool(raw_value and raw_value != "no-rov" and re.match(r"^[A-Z][A-Z0-9]*-\d+$", raw_value))


def _issue_key_sort_key(issue_key):
    match = re.search(r"-(\d+)$", str(issue_key or ""))
    return int(match.group(1)) if match else -1


def _collect_known_rov_links(base_items=None):
    links = defaultdict(dict)

    def _add_link(release_key, rov_key, *, source="", item=None):
        release_key = str(release_key or "").strip()
        rov_key = str(rov_key or "").strip()
        if not release_key or not _is_real_rov_key(rov_key):
            return
        entry = links[release_key].setdefault(
            rov_key,
            {
                "sources": set(),
                "item": None,
            },
        )
        if source:
            entry["sources"].add(source)
        if item is not None and entry.get("item") is None:
            entry["item"] = item

    for item in base_items or []:
        if not isinstance(item, dict):
            continue
        _add_link(item.get("release_key"), item.get("rov_key"), source="snapshot", item=dict(item))

    for row_key in (_load_reviewer_assignments() or {}).keys():
        release_key, rov_key = _split_release_monitor_row_key(row_key)
        _add_link(release_key, rov_key, source="reviewers")

    zni_payload = _load_zni_payload()
    for section_name in ("issues", "flags"):
        for row_key in (zni_payload.get(section_name) or {}).keys():
            release_key, rov_key = _split_release_monitor_row_key(row_key)
            _add_link(release_key, rov_key, source=f"zni.{section_name}")

    for row_key in (_load_release_attempt_outcomes() or {}).keys():
        release_key, rov_key = _split_release_monitor_row_key(row_key)
        _add_link(release_key, rov_key, source="attempts")

    return {
        release_key: {
            rov_key: {
                "sources": sorted(entry.get("sources") or []),
                "item": entry.get("item"),
            }
            for rov_key, entry in sorted(
                rov_entries.items(),
                key=lambda pair: _issue_key_sort_key(pair[0]),
            )
        }
        for release_key, rov_entries in links.items()
    }


def _build_cached_rov_record_from_item(item):
    if not isinstance(item, dict):
        return None
    rov_key = str(item.get("rov_key") or "").strip()
    if not _is_real_rov_key(rov_key):
        return None

    start_dt = _parse_jira_date(
        item.get("deployment_start_iso")
        or item.get("source_deployment_start_iso")
        or item.get("deployment_start")
        or item.get("source_deployment_start")
    )
    end_dt = _parse_jira_date(
        item.get("deployment_end_iso")
        or item.get("source_deployment_end_iso")
        or item.get("deployment_end")
        or item.get("source_deployment_end")
    )
    return {
        "key": rov_key,
        "summary": str(item.get("rov_summary") or ""),
        "status": str(item.get("rov_status") or ""),
        "issue_type": ROV_ISSUE_TYPE,
        "start_dt": start_dt,
        "end_dt": end_dt,
        "start": start_dt.strftime("%d.%m.%Y") if start_dt else str(item.get("deployment_start") or ""),
        "end": end_dt.strftime("%d.%m.%Y") if end_dt else str(item.get("deployment_end") or ""),
        "start_iso": start_dt.isoformat() if start_dt else str(item.get("deployment_start_iso") or ""),
        "end_iso": end_dt.isoformat() if end_dt else str(item.get("deployment_end_iso") or ""),
        "url": str(item.get("rov_url") or _build_jira_issue_url_from_key(rov_key)),
    }


def _inject_release_io_links(issue, rov_keys):
    cloned_issue = copy.deepcopy(issue or {})
    fields = cloned_issue.setdefault("fields", {})
    links = list(fields.get("issuelinks") or [])
    existing_keys = set(_extract_release_io_keys(cloned_issue))
    for rov_key in rov_keys:
        rov_key = str(rov_key or "").strip()
        if not _is_real_rov_key(rov_key) or rov_key in existing_keys:
            continue
        links.append(
            {
                "type": {
                    "name": "ReleaseIO",
                    "inward": "Introduction Order",
                    "outward": "Release",
                },
                "inwardIssue": {
                    "key": rov_key,
                    "fields": {
                        "issuetype": {"name": ROV_ISSUE_TYPE},
                    },
                },
            }
        )
        existing_keys.add(rov_key)
    fields["issuelinks"] = links
    return cloned_issue


def _restore_missing_rov_links_from_known_state(
    domain,
    token,
    release_issues_by_key,
    base_items,
    resolved_fields,
    release_fields_to_load,
    rov_fields_to_load,
    *,
    retry_kwargs=None,
    context_label="refresh",
):
    known_links = _collect_known_rov_links(base_items)
    if not known_links:
        return {}

    retry_kwargs = dict(retry_kwargs or {})
    restored_rov_records = {}
    for release_key, pair in list((release_issues_by_key or {}).items()):
        release_key = str(release_key or "").strip()
        if not release_key:
            continue

        prefix, search_issue = pair
        if _extract_release_io_keys(search_issue):
            continue

        known_rov_entries = known_links.get(release_key) or {}
        if not known_rov_entries:
            continue

        known_rov_keys = sorted(known_rov_entries.keys(), key=_issue_key_sort_key)
        logging.warning(
            "Release monitor: suspicious ROV link loss during %s for %s; known ROV keys=%s",
            context_label,
            release_key,
            ", ".join(known_rov_keys),
        )

        direct_release_issue = None
        direct_release_failed = False
        try:
            direct_release_issue = _execute_issue_get(
                domain,
                token,
                release_key,
                release_fields_to_load,
                retry_label=f"{domain} direct release {release_key}",
                **retry_kwargs,
            )
        except Exception as exc:
            direct_release_failed = True
            logging.warning(
                "Release monitor: direct release ROV recheck failed for %s during %s: %s",
                release_key,
                context_label,
                exc,
            )

        direct_rov_keys = _extract_release_io_keys(direct_release_issue or {})
        if direct_rov_keys:
            release_issues_by_key[release_key] = (prefix, direct_release_issue)
            for rov_key in direct_rov_keys:
                try:
                    rov_issue = _execute_issue_get(
                        domain,
                        token,
                        rov_key,
                        rov_fields_to_load,
                        retry_label=f"{domain} direct ROV {rov_key}",
                        **retry_kwargs,
                    )
                    restored_rov_records[rov_key] = _build_rov_record(rov_issue, domain, resolved_fields)
                except Exception as exc:
                    cached_record = _build_cached_rov_record_from_item(
                        (known_rov_entries.get(rov_key) or {}).get("item")
                    )
                    if cached_record:
                        restored_rov_records[rov_key] = cached_record
                    logging.warning(
                        "Release monitor: direct ROV load failed after release recheck for %s -> %s during %s: %s",
                        release_key,
                        rov_key,
                        context_label,
                        exc,
                    )
            logging.warning(
                "Release monitor: restored ROV links for %s from direct release issue API: %s",
                release_key,
                ", ".join(direct_rov_keys),
            )
            continue

        reverse_restored_keys = []
        reverse_check_failed = False
        for rov_key in known_rov_keys:
            try:
                rov_issue = _execute_issue_get(
                    domain,
                    token,
                    rov_key,
                    rov_fields_to_load,
                    retry_label=f"{domain} direct ROV {rov_key}",
                    **retry_kwargs,
                )
            except Exception as exc:
                reverse_check_failed = True
                logging.warning(
                    "Release monitor: direct ROV reverse recheck failed for %s -> %s during %s: %s",
                    release_key,
                    rov_key,
                    context_label,
                    exc,
                )
                continue

            linked_release_keys = _extract_linked_release_keys_from_rov(rov_issue or {})
            if release_key not in linked_release_keys:
                continue

            reverse_restored_keys.append(rov_key)
            restored_rov_records[rov_key] = _build_rov_record(rov_issue, domain, resolved_fields)

        if reverse_restored_keys:
            release_issues_by_key[release_key] = (
                prefix,
                _inject_release_io_links(direct_release_issue or search_issue, reverse_restored_keys),
            )
            logging.warning(
                "Release monitor: restored ROV links for %s from reverse ROV issue API: %s",
                release_key,
                ", ".join(reverse_restored_keys),
            )
            continue

        if direct_release_failed or reverse_check_failed:
            cached_restored_keys = []
            for rov_key in known_rov_keys:
                cached_record = _build_cached_rov_record_from_item(
                    (known_rov_entries.get(rov_key) or {}).get("item")
                )
                if not cached_record:
                    continue
                cached_restored_keys.append(rov_key)
                restored_rov_records[rov_key] = cached_record

            if cached_restored_keys:
                release_issues_by_key[release_key] = (
                    prefix,
                    _inject_release_io_links(direct_release_issue or search_issue, cached_restored_keys),
                )
                logging.warning(
                    "Release monitor: preserved cached ROV links for %s after technical recheck failure: %s",
                    release_key,
                    ", ".join(cached_restored_keys),
                )
                continue

        logging.warning(
            "Release monitor: confirmed missing ROV links for %s during %s after direct checks; known ROV keys=%s",
            release_key,
            context_label,
            ", ".join(known_rov_keys),
        )
        confirmed_issue = copy.deepcopy(direct_release_issue or search_issue or {})
        confirmed_issue["_release_monitor_confirmed_missing_rov"] = True
        release_issues_by_key[release_key] = (prefix, confirmed_issue)

    return restored_rov_records


def _sort_rov_records_for_release(rov_records):
    def _record_sort_key(rov_record):
        if not isinstance(rov_record, dict):
            return (datetime.min, datetime.min, "")

        start_dt = rov_record.get("start_dt") or datetime.min
        end_dt = rov_record.get("end_dt") or datetime.min
        issue_key = str(rov_record.get("key") or "")
        return (start_dt, end_dt, issue_key)

    return sorted(
        [record for record in (rov_records or []) if isinstance(record, dict)],
        key=_record_sort_key,
    )


def _clean_release_summary(summary):
    cleaned = re.sub(r"^\s*Р РµР»РёР·#\d+\s*", "", summary or "", flags=re.IGNORECASE)
    return cleaned.strip()


def _detect_system(prefix, summary, ke_name, system_info_text):
    raw_searchable = str(f"{summary} {ke_name} {system_info_text}" or "").lower()
    searchable = _normalize_text(raw_searchable)

    aist_markers = ("\u0430\u0438\u0441\u0442", "aist", "Р°РёСЃС‚".lower(), "РђРРЎРў".lower())
    if any(marker in raw_searchable or marker in searchable for marker in aist_markers):
        return "\u0410\u0418\u0421\u0422"
    if prefix in {"AIGAS", "HELPERAI", "DRMMMB"} or "ai-" in searchable or "ai " in searchable:
        return "AI-\u0410\u0433\u0435\u043d\u0442\u044b"
    if "clm" in searchable or prefix in {"SMECLM", "SMECSC"}:
        return "CLM"
    return "\u0424\u043e\u043a\u0443\u0441"


def _apply_template_system_classification(items):
    """Give AI_AGENTS templates priority over legacy prefix-based system rules."""
    try:
        catalog = build_runtime_template_catalog()
    except Exception as exc:
        logging.warning("Release monitor: failed to load template catalog for system classification: %s", exc)
        return items

    entries_by_ke = defaultdict(list)
    for entry in catalog:
        ke_id = str((entry or {}).get("ke") or "").strip()
        if ke_id:
            entries_by_ke[ke_id].append(entry)

    for item in items or []:
        if not isinstance(item, dict):
            continue

        item["template_category"] = ""
        item["is_ai_agent_template"] = False
        ke_id = str(item.get("ke_id") or "").strip()
        candidates = entries_by_ke.get(ke_id, [])
        if not candidates:
            continue

        summary = str(
            item.get("release_summary")
            or item.get("base_release_summary")
            or " ".join(item.get("release_name_lines") or [])
            or ""
        )
        selected = candidates[0] if len(candidates) == 1 else select_template_by_summary(candidates, summary)
        if selected:
            item["template_category"] = str(selected.get("category") or "").strip()

        ai_candidates = [
            candidate
            for candidate in candidates
            if is_ai_agents_template_category(candidate.get("category"))
        ]
        selected_is_ai = bool(
            selected and is_ai_agents_template_category(selected.get("category"))
        )
        all_candidates_are_ai = bool(ai_candidates and len(ai_candidates) == len(candidates))
        if selected_is_ai or all_candidates_are_ai:
            item["template_category"] = str(
                (selected or ai_candidates[0]).get("category") or "AI_AGENTS"
            ).strip()
            item["is_ai_agent_template"] = True
            item["system_name"] = AI_AGENTS_SYSTEM_NAME

    return items


def _build_release_name_lines(summary, release_ke_line, row_label="(\u0420\u0435\u043b\u0438\u0437)"):
    lines = []
    short_name = _clean_release_summary(summary)
    if short_name:
        lines.append(short_name)

    if release_ke_line:
        lines.append(release_ke_line)

    lines.append("(Релиз)")
    lines[-1] = row_label

    return lines


def _build_rov_record(issue, domain, resolved_fields):
    fields = issue.get("fields", {})
    rov_start = _parse_jira_date(_extract_field_value(fields.get(resolved_fields["rov_start"])))
    rov_end = _parse_jira_date(_extract_field_value(fields.get(resolved_fields["rov_end"])))

    return {
        "key": issue.get("key"),
        "summary": fields.get("summary", ""),
        "status": (fields.get("status") or {}).get("name", ""),
        "issue_type": (fields.get("issuetype") or {}).get("name", ""),
        "start_dt": rov_start,
        "end_dt": rov_end,
        "start": rov_start.strftime("%d.%m.%Y") if rov_start else "",
        "end": rov_end.strftime("%d.%m.%Y") if rov_end else "",
        "start_iso": rov_start.isoformat() if rov_start else "",
        "end_iso": rov_end.isoformat() if rov_end else "",
        "url": f"{domain}/browse/{issue.get('key')}",
    }


def _pick_release_year(rov_start, rov_end, planned_prom_start, planned_prom_end, created_dt):
    for dt in (rov_start, rov_end, planned_prom_start, planned_prom_end, created_dt):
        if dt:
            return dt.year
    return datetime.now().year


def _pick_release_sort_dt(rov_start, rov_end, planned_prom_start, planned_prom_end):
    return rov_start or rov_end or planned_prom_start or planned_prom_end


def _build_release_record(
    issue,
    domain,
    prefix,
    resolved_fields,
    rov_map,
    current_year,
    previous_year,
    *,
    attempt_outcomes=None,
):
    fields = issue.get("fields", {})
    status_name = (fields.get("status") or {}).get("name", "")
    summary = fields.get("summary", "")
    created_dt = _parse_jira_date(fields.get("created"))
    planned_prom_start = _parse_jira_date(_extract_field_value(fields.get(resolved_fields["planned_prom_start"])))
    planned_prom_end = _parse_jira_date(_extract_field_value(fields.get(resolved_fields["planned_prom_end"])))
    system_info_text = _extract_field_value(fields.get(resolved_fields["system_info"])) or ""

    ke_object = _extract_ke_object(fields, resolved_fields)
    ai_agent_release_context = _build_ai_agent_release_context(
        fields,
        resolved_fields,
        summary=summary,
        system_info_text=system_info_text,
        ke_object=ke_object,
    )
    dist_item = _extract_release_dist(
        fields,
        resolved_fields,
        release_context=ai_agent_release_context,
    )
    release_version = _extract_nested_version(dist_item)
    release_dist_url = _extract_nested_dist_url(dist_item)
    dist_ke_raw = extract_artifact_ke_id(dist_item)
    ke_distributive = _format_ke_id(dist_ke_raw)

    normalized_status = _normalize_text(status_name)
    is_cancelled = normalized_status == _normalize_text(CANCELLED_RELEASE_STATUS)
    is_final = normalized_status == _normalize_text(FINAL_RELEASE_STATUS)
    is_ready_for_prom = normalized_status == _normalize_text(READY_FOR_PROM_STATUS)
    is_non_final = not is_final and not is_cancelled
    is_pre_final = normalized_status in {_normalize_text(status) for status in PRE_FINAL_RELEASE_STATUSES}
    ke_name = ke_object.get("name") or ""
    ke_id = ke_object.get("id") or ""
    release_ke_line = f"{ke_name}({ke_id})" if ke_name and ke_id else (ke_name or "")
    if not release_ke_line:
        release_ke_line = (system_info_text or "").strip()
    linked_rov_keys = _extract_release_io_keys(issue)
    linked_rov_records = _sort_rov_records_for_release(
        [rov_map.get(key, {}) for key in linked_rov_keys if rov_map.get(key)]
    )
    row_variants = linked_rov_records or [{}]
    records = []
    now_dt = datetime.now()
    attempt_outcomes = (
        _load_release_attempt_outcomes()
        if attempt_outcomes is None
        else dict(attempt_outcomes or {})
    )
    has_successful_final_attempt_before = False
    stale_successful_attempt_key = ""
    if is_final and linked_rov_records:
        latest_rov = linked_rov_records[-1]
        latest_rov_key = str(latest_rov.get("key") or "").strip()
        if latest_rov_key:
            latest_row_key = f"{issue.get('key')}::{latest_rov_key}"
            if latest_row_key in attempt_outcomes:
                stale_successful_attempt_key = latest_row_key

    for index, rov_data in enumerate(row_variants):
        rov_key = rov_data.get("key", "")
        rov_start = rov_data.get("start_dt")
        rov_end = rov_data.get("end_dt")

        release_year = _pick_release_year(rov_start, rov_end, planned_prom_start, planned_prom_end, created_dt)
        if release_year not in {current_year, previous_year}:
            continue

        row_key = f"{issue.get('key')}::{rov_key or 'no-rov'}"
        is_deferred_attempt = row_key in attempt_outcomes and row_key != stale_successful_attempt_key
        is_reroll = bool(
            is_final
            and rov_key
            and len(linked_rov_records) > 1
            and index > 0
            and has_successful_final_attempt_before
            and not is_deferred_attempt
        )
        row_is_cancelled = is_cancelled
        row_is_final = is_final
        if is_reroll and (rov_start or rov_end) and not _is_release_window_expired(rov_end or rov_start, now_dt):
            row_is_final = False
        row_is_non_final = not row_is_final and not row_is_cancelled
        row_is_pre_final = is_pre_final and not row_is_final
        row_has_confirmed_attempt = _is_confirmed_deployment_attempt(
            {
                "is_pre_final": row_is_pre_final,
                "is_ready_for_prom": is_ready_for_prom and not row_is_final,
                "rov_status": rov_data.get("status", ""),
            }
        )

        is_overdue = bool(
            row_is_non_final
            and row_has_confirmed_attempt
            and _is_release_window_expired(rov_end, now_dt)
        )
        is_today = bool(
            row_is_non_final
            and _is_release_window_in_operational_day(rov_start, rov_end, now_dt)
        )

        if row_is_cancelled:
            row_state = "cancelled"
        elif row_is_final:
            row_state = "final"
        elif is_overdue:
            row_state = "overdue"
        elif is_today:
            row_state = "today"
        else:
            row_state = "planned"

        days_overdue = _release_days_overdue(rov_end, now_dt) if is_overdue else 0
        is_hotfix = str(release_version or "").upper().startswith("P-")
        if is_reroll:
            row_label = "(\u041f\u0435\u0440\u0435\u0440\u0430\u0441\u043a\u0430\u0442\u043a\u0430)"
        elif is_hotfix:
            row_label = "(\u0425\u043e\u0442\u0444\u0438\u043a\u0441)"
        else:
            row_label = "(\u0420\u0435\u043b\u0438\u0437)"

        is_unnumbered = (not rov_key) or (row_is_cancelled and not ke_distributive)

        records.append({
            "row_key": row_key,
            "year": release_year,
            "release_number": "",
            "release_key": issue.get("key"),
            "release_url": f"{domain}/browse/{issue.get('key')}",
            "release_status": status_name,
            "release_status_normalized": normalized_status,
            "release_summary": summary,
            "release_name_lines": _build_release_name_lines(summary, release_ke_line, row_label=row_label),
            "base_release_summary": summary,
            "base_release_name_lines": _build_release_name_lines(summary, release_ke_line, row_label=row_label),
            "is_reroll": is_reroll,
            "is_deferred_attempt": is_deferred_attempt,
            "row_label": row_label,
            "zni_key": "",
            "zni_url": "",
            "base_zni_key": "",
            "base_zni_url": "",
            "has_rollout_notes": False,
            "rollout_notes_level": "",
            "ke": ke_distributive,
            "base_ke": ke_distributive,
            "ke_name": ke_name,
            "ke_id": ke_id,
            "release_version": release_version,
            "release_dist_url": release_dist_url,
            "base_release_version": release_version,
            "base_release_dist_url": release_dist_url,
            "rov_key": rov_key,
            "rov_url": rov_data.get("url", ""),
            "rov_status": rov_data.get("status", ""),
            "has_rov": bool(rov_key),
            "deployment_start": rov_data.get("start", ""),
            "deployment_start_iso": rov_data.get("start_iso", ""),
            "deployment_end": rov_data.get("end", ""),
            "deployment_end_iso": rov_data.get("end_iso", ""),
            "source_deployment_start": rov_data.get("start", ""),
            "source_deployment_start_iso": rov_data.get("start_iso", ""),
            "source_deployment_end": rov_data.get("end", ""),
            "source_deployment_end_iso": rov_data.get("end_iso", ""),
            "psi_owner": "",
            "psi_responsibles": [],
            "psi_checker": "",
            "row_state": row_state,
            "is_final": row_is_final,
            "is_cancelled": row_is_cancelled,
            "is_non_final": row_is_non_final,
            "is_pre_final": row_is_pre_final,
            "is_ready_for_prom": is_ready_for_prom and not row_is_final,
            "is_overdue": is_overdue,
            "is_today": is_today,
            "days_overdue": days_overdue,
            "waits_for_rov": not rov_key and not is_cancelled,
            "rov_unlink_confirmed": bool(issue.get("_release_monitor_confirmed_missing_rov")),
            "is_unnumbered": is_unnumbered,
            "is_natural_unnumbered": is_unnumbered,
            "is_force_unnumbered": False,
            "source_prefix": prefix,
            "system_name": _detect_system(prefix, summary, ke_name, system_info_text),
            "sort_date": _pick_release_sort_dt(rov_start, rov_end, planned_prom_start, planned_prom_end).isoformat() if _pick_release_sort_dt(rov_start, rov_end, planned_prom_start, planned_prom_end) else "",
            "created_sort_date": created_dt.isoformat() if created_dt else "",
            "created": fields.get("created", ""),
        })

        if is_final and rov_key and not is_deferred_attempt:
            has_successful_final_attempt_before = True

    return records


def _sort_datetime_value(item, field_name="sort_date"):
    sort_value = item.get(field_name)
    if isinstance(sort_value, datetime):
        return sort_value
    if isinstance(sort_value, str) and sort_value:
        parsed = _parse_jira_date(sort_value)
        if parsed:
            return parsed
    return datetime.min


def _apply_group_manual_order(year, group_name, items):
    if not items:
        return items

    manual_order = _load_manual_order()
    year_payload = manual_order.get(str(year), {}) if isinstance(manual_order, dict) else {}
    group_payload = year_payload.get(group_name, []) if isinstance(year_payload, dict) else []

    item_by_key = {
        str(item.get("row_key") or "").strip(): item
        for item in items
        if str(item.get("row_key") or "").strip()
    }

    merged_items = list(items)

    if isinstance(group_payload, dict):
        bucket_orders = {
            str(bucket or "").strip(): [
                str(row_key or "").strip()
                for row_key in (row_keys or [])
                if str(row_key or "").strip()
            ]
            for bucket, row_keys in group_payload.items()
            if str(bucket or "").strip() and isinstance(row_keys, list)
        }
    else:
        legacy_order = [
            str(value or "").strip()
            for value in (group_payload or [])
            if str(value or "").strip()
        ]
        bucket_orders = defaultdict(list)
        for row_key in legacy_order:
            item = item_by_key.get(row_key)
            if item:
                bucket_orders[_get_release_order_bucket(item)].append(row_key)

    for bucket, ordered_row_keys in bucket_orders.items():
        present_manual_keys = [
            row_key
            for row_key in ordered_row_keys
            if row_key in item_by_key
            and _get_release_order_bucket(item_by_key[row_key]) == bucket
        ]
        if not present_manual_keys:
            continue

        manual_key_set = set(present_manual_keys)
        manual_slot_indexes = [
            index
            for index, item in enumerate(merged_items)
            if _get_release_order_bucket(item) == bucket
            and str(item.get("row_key") or "").strip() in manual_key_set
        ]
        if not manual_slot_indexes:
            continue

        ordered_manual_items = [item_by_key[row_key] for row_key in present_manual_keys]
        for slot_index, manual_item in zip(manual_slot_indexes, ordered_manual_items):
            merged_items[slot_index] = manual_item
    return merged_items


def _derive_natural_unnumbered(item):
    if not isinstance(item, dict):
        return False
    has_rov = bool(item.get("has_rov") or str(item.get("rov_key") or "").strip())
    if _is_cancelled_reroll_rov(item):
        return True
    is_cancelled = bool(item.get("is_cancelled"))
    if item.get("manual_release") and item.get("jira_duplicate_detected"):
        return True
    if item.get("manual_release"):
        return bool(is_cancelled)
    has_ke = bool(str(item.get("ke") or "").strip())
    has_manual_start = bool(str(item.get("manual_deployment_start") or "").strip())
    can_number_without_rov = bool(
        (not has_rov)
        and item.get("is_non_final")
        and has_manual_start
    )
    if can_number_without_rov:
        return bool(is_cancelled and not has_ke)
    return (not has_rov) or (is_cancelled and not has_ke)


def _apply_force_unnumbered_flags(year, items):
    manual_order = _load_manual_order()
    year_payload = manual_order.get(str(year), {}) if isinstance(manual_order, dict) else {}
    forced_keys = {
        str(row_key or "").strip()
        for row_key in (year_payload.get("force_unnumbered") or [])
        if str(row_key or "").strip()
    }
    forced_numbered_keys = {
        str(row_key or "").strip()
        for row_key in (year_payload.get("force_numbered") or [])
        if str(row_key or "").strip()
    }

    for item in items:
        row_key = str(item.get("row_key") or item.get("release_key") or "").strip()
        has_manual_start = bool(str(item.get("manual_deployment_start") or "").strip())
        manual_numbering_override = bool(
            has_manual_start
            and item.get("is_non_final")
            and not item.get("has_rov")
        )
        is_natural_unnumbered = _derive_natural_unnumbered(item)
        is_forced = row_key in forced_keys
        is_force_numbered = row_key in forced_numbered_keys
        is_cancelled_reroll_rov = _is_cancelled_reroll_rov(item)
        if is_force_numbered:
            is_forced = False
        item["is_natural_unnumbered"] = is_natural_unnumbered
        item["is_force_unnumbered"] = is_forced
        item["is_force_numbered"] = is_force_numbered
        item["is_cancelled_reroll_rov"] = is_cancelled_reroll_rov
        item["is_manual_numbering_override"] = manual_numbering_override
        item["is_unnumbered"] = bool(
            (is_natural_unnumbered or is_forced)
            and not manual_numbering_override
            and not is_force_numbered
        )


def _sort_and_number_records(records):
    records_by_year = defaultdict(list)
    for item in records:
        records_by_year[item["year"]].append(item)

    for year_items in records_by_year.values():
        _apply_force_unnumbered_flags(year_items[0]["year"], year_items)
        numbered_items = [item for item in year_items if not item.get("is_unnumbered")]
        waiting_items = [item for item in year_items if item.get("is_unnumbered")]

        numbered_items.sort(
            key=lambda item: (
                _sort_datetime_value(item, "sort_date"),
                _sort_datetime_value(item, "created_sort_date"),
                item.get("release_key", ""),
            ),
            reverse=True,
        )

        waiting_items.sort(
            key=lambda item: (
                _sort_datetime_value(item, "sort_date"),
                _sort_datetime_value(item, "created_sort_date"),
                item.get("release_key", ""),
            ),
            reverse=True,
        )

        waiting_items = _apply_group_manual_order(year_items[0]["year"], "waiting", waiting_items)
        numbered_items = _apply_group_manual_order(year_items[0]["year"], "numbered", numbered_items)

        total_numbered = len(numbered_items)
        for index, item in enumerate(numbered_items):
            item["release_number"] = total_numbered - index
            item["_group_rank"] = index

        for index, item in enumerate(waiting_items):
            item["release_number"] = ""
            item["_group_rank"] = index

    records.sort(
        key=lambda item: (
            -item.get("year", 0),
            0 if item.get("is_unnumbered") else 1,
            item.get("_group_rank", 0),
        )
    )
    return records


def save_release_monitor_manual_order(
    year,
    waiting_row_keys=None,
    numbered_row_keys=None,
    force_unnumbered_row_keys=None,
    force_numbered_row_keys=None,
):
    global _cached_data, _last_cache_update

    year = int(year or datetime.now().year)

    def _normalize_order_payload(value):
        if isinstance(value, dict):
            normalized = {}
            for bucket, row_keys in value.items():
                bucket_key = str(bucket or "").strip()
                if not bucket_key:
                    continue
                normalized_keys = []
                for row_key in (row_keys or []):
                    row_key = str(row_key or "").strip()
                    if row_key and row_key not in normalized_keys:
                        normalized_keys.append(row_key)
                if normalized_keys:
                    normalized[bucket_key] = normalized_keys
            return normalized

        normalized_keys = []
        for row_key in (value or []):
            row_key = str(row_key or "").strip()
            if row_key and row_key not in normalized_keys:
                normalized_keys.append(row_key)
        return normalized_keys

    normalized_waiting = _normalize_order_payload(waiting_row_keys)
    normalized_numbered = _normalize_order_payload(numbered_row_keys)
    normalized_force_unnumbered = []
    normalized_force_numbered = []

    for row_key in (force_unnumbered_row_keys or []):
        row_key = str(row_key or "").strip()
        if row_key and row_key not in normalized_force_unnumbered:
            normalized_force_unnumbered.append(row_key)

    for row_key in (force_numbered_row_keys or []):
        row_key = str(row_key or "").strip()
        if row_key and row_key not in normalized_force_numbered:
            normalized_force_numbered.append(row_key)

    normalized_force_unnumbered = [
        row_key for row_key in normalized_force_unnumbered if row_key not in normalized_force_numbered
    ]

    with _cache_lock:
        manual_order = _load_manual_order()
        manual_order[str(year)] = {
            "waiting": normalized_waiting,
            "numbered": normalized_numbered,
            "force_unnumbered": normalized_force_unnumbered,
            "force_numbered": normalized_force_numbered,
        }
        _save_manual_order(manual_order)

        payload = _rebuild_cached_payload_after_state_change_locked()
        if not payload.get("items"):
            disk_payload = _load_snapshot_from_disk()
            if disk_payload and disk_payload.get("items"):
                _cached_data = _hydrate_release_monitor_payload(disk_payload)
                _last_cache_update = _get_snapshot_mtime() or time.time()
                payload = _rebuild_cached_payload_after_state_change_locked()
        return {
            "year": year,
            "data": payload,
        }


def _build_summary(records, current_year, previous_year):
    summary = {
        "total": len(records),
        "non_final": 0,
        "overdue": 0,
        "today": 0,
        "pre_final": 0,
        "final": 0,
        "cancelled": 0,
        "by_status": defaultdict(int),
        "by_year": {
            str(current_year): {
                "total": 0,
                "non_final": 0,
                "overdue": 0,
                "today": 0,
                "pre_final": 0,
                "final": 0,
                "cancelled": 0,
            },
            str(previous_year): {
                "total": 0,
                "non_final": 0,
                "overdue": 0,
                "today": 0,
                "pre_final": 0,
                "final": 0,
                "cancelled": 0,
            },
        },
    }

    for item in records:
        year_bucket = summary["by_year"].setdefault(
            str(item["year"]),
            {
                "total": 0,
                "non_final": 0,
                "overdue": 0,
                "today": 0,
                "pre_final": 0,
                "final": 0,
                "cancelled": 0,
            },
        )
        summary["by_status"][item["release_status"] or "РќРµ СѓРєР°Р·Р°РЅ"] += 1
        year_bucket["total"] += 1

        if item["is_non_final"]:
            summary["non_final"] += 1
            year_bucket["non_final"] += 1
        if item["is_overdue"]:
            summary["overdue"] += 1
            year_bucket["overdue"] += 1
        if item["is_today"]:
            summary["today"] += 1
            year_bucket["today"] += 1
        if item["is_pre_final"]:
            summary["pre_final"] += 1
            year_bucket["pre_final"] += 1
        if item["is_final"]:
            summary["final"] += 1
            year_bucket["final"] += 1
        if item["is_cancelled"]:
            summary["cancelled"] += 1
            year_bucket["cancelled"] += 1

    summary["by_status"] = dict(
        sorted(summary["by_status"].items(), key=lambda pair: (-pair[1], pair[0]))
    )
    return summary


def _is_manual_release_monitor_item(item):
    if not isinstance(item, dict):
        return False
    row_key = str(item.get("row_key") or "").strip()
    return bool(
        item.get("manual_release")
        or item.get("source") == "manual"
        or row_key.startswith("manual::")
    )


def _candidate_prefix(item):
    prefix = str((item or {}).get("source_prefix") or "").strip().upper()
    if prefix:
        return prefix
    release_key = str((item or {}).get("release_key") or "").strip().upper()
    return release_key.split("-", 1)[0] if "-" in release_key else ""


def _build_snapshot_validation_profile(items, *, years=None):
    selected_years = set(years or [])
    profile = {
        "total": 0,
        "real_rov": 0,
        "by_year": {},
        "by_prefix": {},
    }
    for item in items or []:
        if not isinstance(item, dict) or _is_manual_release_monitor_item(item):
            continue
        try:
            item_year = int(item.get("year"))
        except (TypeError, ValueError):
            item_year = None
        if selected_years and item_year not in selected_years:
            continue

        profile["total"] += 1
        if _is_real_rov_key(item.get("rov_key")):
            profile["real_rov"] += 1
        if item_year is not None:
            year_key = str(item_year)
            profile["by_year"][year_key] = int(profile["by_year"].get(year_key) or 0) + 1
        prefix = _candidate_prefix(item)
        if prefix:
            profile["by_prefix"][prefix] = int(profile["by_prefix"].get(prefix) or 0) + 1
    return profile


def _release_monitor_item_years(items):
    years = set()
    for item in items or []:
        if not isinstance(item, dict) or _is_manual_release_monitor_item(item):
            continue
        try:
            years.add(int(item.get("year")))
        except (TypeError, ValueError):
            continue
    return years


def _build_raw_release_candidate(items, source_ledger, mode):
    current_year = datetime.now().year
    years = {current_year, current_year - 1}
    return {
        "items": [dict(item) for item in (items or []) if isinstance(item, dict)],
        "source_ledger": [dict(entry) for entry in (source_ledger or [])],
        "raw_counters": _build_snapshot_validation_profile(items, years=years),
        "meta": {
            "mode": mode,
            "generated_at": _utc_now_iso(),
            "years": sorted(years, reverse=True),
            "raw_candidate": True,
        },
    }


def _drop_exceeds_threshold(previous, candidate, absolute_minimum, ratio):
    previous = int(previous or 0)
    candidate = int(candidate or 0)
    loss = previous - candidate
    threshold = max(int(absolute_minimum), previous * float(ratio))
    return loss > threshold, loss, threshold


def _validate_release_candidate(candidate, baseline_payload):
    trace_started_at = time.monotonic()
    _rm_trace(
        "RM_VALIDATION",
        "start",
        baseline_items=_count_payload_items(baseline_payload),
        candidate_items=_count_payload_items(candidate),
        mode=((candidate or {}).get("meta") or {}).get("mode", ""),
    )
    current_year = datetime.now().year
    candidate_items = list((candidate or {}).get("items") or [])
    baseline_items = list((baseline_payload or {}).get("items") or [])
    expected_candidate_years = {
        int(year)
        for year in ((candidate or {}).get("meta") or {}).get("years", [])
        if str(year).isdigit()
    } or {current_year, current_year - 1}
    baseline_years = _release_monitor_item_years(baseline_items)
    comparison_years = baseline_years & expected_candidate_years
    if not comparison_years:
        comparison_years = expected_candidate_years
    candidate_profile = _build_snapshot_validation_profile(candidate_items, years=comparison_years)
    baseline_profile = _build_snapshot_validation_profile(baseline_items, years=comparison_years)
    reasons = []

    release_keys = []
    row_keys = []
    for item in candidate_items:
        release_key = str((item or {}).get("release_key") or "").strip()
        row_key = str((item or {}).get("row_key") or "").strip()
        if not release_key:
            reasons.append(
                {
                    "code": "empty_release_key",
                    "message": "Candidate contains a row without release_key",
                }
            )
        if release_key:
            release_keys.append(release_key)
        if row_key:
            row_keys.append(row_key)
        else:
            reasons.append(
                {
                    "code": "empty_row_key",
                    "message": f"Candidate row {release_key or '<unknown>'} has no row_key",
                }
            )

    duplicate_row_keys = sorted(
        row_key
        for row_key, count in {
            key: row_keys.count(key)
            for key in set(row_keys)
        }.items()
        if count > 1
    )
    if duplicate_row_keys:
        reasons.append(
            {
                "code": "duplicate_row_key",
                "message": f"Candidate contains duplicate row_key values: {', '.join(duplicate_row_keys[:10])}",
                "count": len(duplicate_row_keys),
            }
        )

    source_ledger = list((candidate or {}).get("source_ledger") or [])
    failed_sources = [entry for entry in source_ledger if entry.get("status") != "success"]
    incomplete_sources = [
        entry
        for entry in source_ledger
        if entry.get("status") == "success"
        and int(entry.get("fetched_total") or 0) != int(entry.get("expected_total") or 0)
    ]
    if failed_sources:
        reasons.append(
            {
                "code": "mandatory_source_failed",
                "message": f"{len(failed_sources)} mandatory Jira source(s) failed",
            }
        )
    if incomplete_sources:
        reasons.append(
            {
                "code": "incomplete_pagination",
                "message": f"{len(incomplete_sources)} Jira source(s) have incomplete pagination",
            }
        )

    if baseline_profile["total"] > 0:
        exceeds, loss, threshold = _drop_exceeds_threshold(
            baseline_profile["total"],
            candidate_profile["total"],
            10,
            0.05,
        )
        if exceeds:
            reasons.append(
                {
                    "code": "total_drop",
                    "message": (
                        f"Total Jira rows dropped from {baseline_profile['total']} "
                        f"to {candidate_profile['total']}"
                    ),
                    "loss": loss,
                    "threshold": threshold,
                }
            )

        if current_year in comparison_years:
            previous_current_year = int(baseline_profile["by_year"].get(str(current_year)) or 0)
            candidate_current_year = int(candidate_profile["by_year"].get(str(current_year)) or 0)
            exceeds, loss, threshold = _drop_exceeds_threshold(
                previous_current_year,
                candidate_current_year,
                5,
                0.03,
            )
            if exceeds:
                reasons.append(
                    {
                        "code": "current_year_drop",
                        "message": (
                            f"Current-year rows dropped from {previous_current_year} "
                            f"to {candidate_current_year}"
                        ),
                        "loss": loss,
                        "threshold": threshold,
                    }
                )

        exceeds, loss, threshold = _drop_exceeds_threshold(
            baseline_profile["real_rov"],
            candidate_profile["real_rov"],
            10,
            0.05,
        )
        if exceeds:
            reasons.append(
                {
                    "code": "real_rov_drop",
                    "message": (
                        f"release::REAL_ROV rows dropped from {baseline_profile['real_rov']} "
                        f"to {candidate_profile['real_rov']}"
                    ),
                    "loss": loss,
                    "threshold": threshold,
                }
            )

        for prefix in RELEASE_PREFIXES:
            previous_count = int(baseline_profile["by_prefix"].get(prefix) or 0)
            candidate_count = int(candidate_profile["by_prefix"].get(prefix) or 0)
            if previous_count > 0 and candidate_count == 0:
                reasons.append(
                    {
                        "code": "prefix_disappeared",
                        "message": f"Previously populated prefix {prefix} disappeared",
                        "prefix": prefix,
                        "previous": previous_count,
                    }
                )

    status = "accepted" if not reasons else "rejected"
    report = {
        "status": status,
        "validated_at": _utc_now_iso(),
        "reasons": reasons,
        "baseline": baseline_profile,
        "candidate": candidate_profile,
        "comparison_years": sorted(comparison_years, reverse=True),
        "thresholds": {
            "total_drop": {"minimum": 10, "ratio": 0.05},
            "current_year_drop": {"minimum": 5, "ratio": 0.03},
            "real_rov_drop": {"minimum": 10, "ratio": 0.05},
        },
    }
    _rm_trace(
        "RM_VALIDATION",
        "complete",
        started_at=trace_started_at,
        baseline_total=baseline_profile["total"],
        candidate_total=candidate_profile["total"],
        reasons=",".join(reason.get("code", "") for reason in reasons) or "none",
        status=status,
    )
    return report


def _compose_release_payload(all_records, mode):
    current_year = datetime.now().year
    previous_year = current_year - 1
    _apply_reviewer_assignments(all_records)
    _apply_date_overrides(all_records)
    _apply_duty_schedule_assignments(all_records, persist=False)
    _sort_and_number_records(all_records)
    payload = {
        "items": all_records,
        "summary": _build_summary(all_records, current_year, previous_year),
        "meta": {
            "final_status": FINAL_RELEASE_STATUS,
            "cancelled_status": CANCELLED_RELEASE_STATUS,
            "final_statuses": list(FINAL_RELEASE_STATUSES),
            "pre_final_statuses": list(PRE_FINAL_RELEASE_STATUSES),
            "prefixes": list(RELEASE_PREFIXES),
            "years": [current_year, previous_year],
            "current_year": current_year,
            "last_updated": _format_timestamp(),
            "last_sync_mode": mode,
            "last_auto_incremental_sync": None,
            "last_confluence_sync": None,
            "last_duty_schedule_upload": None,
            "quick_refresh_days": QUICK_REFRESH_DAYS,
            "auto_full_refresh_enabled": AUTO_FULL_REFRESH_ENABLED,
            "auto_full_refresh_hour": AUTO_FULL_REFRESH_HOUR,
            "auto_incremental_refresh_enabled": AUTO_INCREMENTAL_REFRESH_ENABLED,
            "auto_incremental_refresh_interval_seconds": AUTO_INCREMENTAL_REFRESH_INTERVAL_SECONDS,
            "auto_incremental_refresh_lookback_minutes": AUTO_INCREMENTAL_REFRESH_LOOKBACK_MINUTES,
            "data_revision": _read_data_revision(),
            "is_cached": True,
        },
    }
    payload["meta"] = _append_duty_schedule_meta(payload["meta"])
    return payload


def _fetch_release_monitor_data(*, reliable=False, base_items=None):
    current_year = datetime.now().year
    previous_year = current_year - 1
    all_records = []
    source_ledger = []
    mode = RELIABLE_FULL_REFRESH_MODE if reliable else "full"
    max_attempts = RELIABLE_SEARCH_MAX_ATTEMPTS if reliable else FULL_SEARCH_MAX_ATTEMPTS
    deadline_seconds = (
        RELIABLE_FULL_REFRESH_DEADLINE_SECONDS
        if reliable
        else FULL_REFRESH_DEADLINE_SECONDS
    )
    deadline_at = time.monotonic() + max(1, deadline_seconds)
    retry_kwargs = {
        "retry_server_errors": True,
        "retry_max_attempts": max_attempts,
        "retry_delay_seconds": RELIABLE_SEARCH_RETRY_DELAY_SECONDS,
        "deadline_at": deadline_at,
    }

    try:
        for (domain, token), prefixes in _get_domain_groups().items():
            _remaining_refresh_seconds(deadline_at)
            resolved_fields = _resolve_field_ids(domain, token)
            release_fields_to_load = {
                "key",
                "summary",
                "status",
                "resolution",
                "assignee",
                "reporter",
                "created",
                "updated",
                "issuetype",
                "priority",
                "issuelinks",
                resolved_fields["planned_prom_start"],
                resolved_fields["planned_prom_end"],
                resolved_fields["system_info"],
                resolved_fields["ke_object"],
                resolved_fields["release_distributive"],
                resolved_fields["delta_release_distributive"],
            }

            domain_release_issues = []
            for prefix in prefixes:
                issues = _execute_release_search(
                    domain,
                    token,
                    prefix,
                    previous_year,
                    release_fields_to_load,
                    strict_pagination=True,
                    source_ledger=source_ledger,
                    **retry_kwargs,
                )
                logging.info(
                    "Release monitor: loaded %s releases for prefix %s",
                    len(issues),
                    prefix,
                )
                domain_release_issues.extend((prefix, issue) for issue in issues)

            release_issues_by_key = {
                issue.get("key"): (prefix, issue)
                for prefix, issue in domain_release_issues
                if issue.get("key")
            }
            restored_rov_records = _restore_missing_rov_links_from_known_state(
                domain,
                token,
                release_issues_by_key,
                base_items or [],
                resolved_fields,
                release_fields_to_load,
                _release_monitor_rov_fields_to_load(resolved_fields),
                retry_kwargs=retry_kwargs,
                context_label=mode,
            )
            domain_release_issues = list(release_issues_by_key.values())

            rov_keys = sorted(
                {
                    rov_key
                    for _, issue in domain_release_issues
                    for rov_key in _extract_release_io_keys(issue)
                    if rov_key
                }
            )
            rov_map = {}
            if rov_keys:
                rov_fields_to_load = {
                    "key",
                    "summary",
                    "status",
                    "issuetype",
                    resolved_fields["rov_start"],
                    resolved_fields["rov_end"],
                }
                rov_issues = _execute_issue_keys_search(
                    domain,
                    token,
                    rov_keys,
                    rov_fields_to_load,
                    strict_pagination=True,
                    source_ledger=source_ledger,
                    **retry_kwargs,
                )
                loaded_rov_keys = {
                    str(issue.get("key") or "").strip()
                    for issue in rov_issues
                    if str(issue.get("key") or "").strip()
                }
                missing_rov_keys = sorted(set(rov_keys) - loaded_rov_keys)
                if missing_rov_keys:
                    raise RuntimeError(
                        f"Linked ROV query for {domain} did not return "
                        f"{len(missing_rov_keys)} requested issue(s): {', '.join(missing_rov_keys[:10])}"
                    )
                rov_map = {
                    issue.get("key"): _build_rov_record(issue, domain, resolved_fields)
                    for issue in rov_issues
                }
                rov_map.update(restored_rov_records)
            else:
                rov_map.update(restored_rov_records)

            for prefix, issue in domain_release_issues:
                records = _build_release_record(
                    issue,
                    domain,
                    prefix,
                    resolved_fields,
                    rov_map,
                    current_year,
                    previous_year,
                    attempt_outcomes={},
                )
                if records:
                    all_records.extend(records)
    except Exception as exc:
        candidate = _build_raw_release_candidate(all_records, source_ledger, mode)
        raise ReleaseMonitorSourceError(str(exc), candidate=candidate) from exc

    return _build_raw_release_candidate(all_records, source_ledger, mode)


def _merge_release_level_updates_into_known_rov_item(existing_item, update_item):
    merged = dict(existing_item or {})
    update_item = dict(update_item or {})
    release_level_fields = (
        "release_key",
        "release_url",
        "release_status",
        "release_status_normalized",
        "release_summary",
        "base_release_summary",
        "ke",
        "base_ke",
        "ke_name",
        "ke_id",
        "release_version",
        "release_dist_url",
        "base_release_version",
        "base_release_dist_url",
        "system_name",
        "base_system_name",
        "release_type",
        "base_release_type",
        "is_hotfix",
        "source_prefix",
        "created",
        "created_sort_date",
    )
    for field_name in release_level_fields:
        if field_name in update_item:
            merged[field_name] = update_item.get(field_name)

    update_lines = list(update_item.get("release_name_lines") or [])
    if update_lines:
        row_label = str(merged.get("row_label") or "").strip()
        if row_label:
            update_lines[-1] = row_label
        merged["release_name_lines"] = update_lines

    base_update_lines = list(update_item.get("base_release_name_lines") or [])
    if base_update_lines:
        row_label = str(merged.get("row_label") or "").strip()
        if row_label:
            base_update_lines[-1] = row_label
        merged["base_release_name_lines"] = base_update_lines

    merged["has_rov"] = True
    merged["rov_unlink_confirmed"] = False
    return merged


def _merge_release_records(existing_items, updated_items):
    current_year = datetime.now().year
    previous_year = current_year - 1
    existing_items = list(existing_items or [])
    updated_items = [dict(item) for item in (updated_items or []) if isinstance(item, dict)]

    existing_by_release = defaultdict(list)
    for item in existing_items:
        release_key = str(item.get("release_key") or "").strip()
        if release_key:
            existing_by_release[release_key].append(item)

    updated_by_release = defaultdict(list)
    for item in updated_items:
        release_key = str(item.get("release_key") or "").strip()
        if release_key:
            updated_by_release[release_key].append(item)

    protected_release_updates = {}
    for release_key, release_updated_items in updated_by_release.items():
        updated_real_rov_items = [
            item for item in release_updated_items if _is_real_rov_key(item.get("rov_key"))
        ]
        updated_no_rov_items = [
            item for item in release_updated_items if not _is_real_rov_key(item.get("rov_key"))
        ]
        existing_real_rov_items = [
            item
            for item in existing_by_release.get(release_key, [])
            if item.get("year") in {current_year, previous_year}
            and _is_real_rov_key(item.get("rov_key"))
        ]
        unlink_confirmed = any(bool(item.get("rov_unlink_confirmed")) for item in updated_no_rov_items)
        if updated_real_rov_items or not updated_no_rov_items or not existing_real_rov_items or unlink_confirmed:
            continue

        update_item = updated_no_rov_items[0]
        protected_release_updates[release_key] = {
            str(item.get("row_key") or item.get("release_key") or ""): _merge_release_level_updates_into_known_rov_item(
                item,
                update_item,
            )
            for item in existing_real_rov_items
            if str(item.get("row_key") or item.get("release_key") or "")
        }
        logging.warning(
            "Release monitor: preserved known ROV rows for %s because refresh returned only no-rov row",
            release_key,
        )

    updated_release_keys = {
        item.get("release_key")
        for item in updated_items
        if item.get("release_key")
    }

    records_by_key = {}
    for item in existing_items:
        item_key = item.get("row_key") or item.get("release_key")
        if not item_key or item.get("year") not in {current_year, previous_year}:
            continue
        release_key = item.get("release_key")
        if release_key in protected_release_updates:
            protected_item = protected_release_updates[release_key].get(str(item_key))
            if protected_item:
                records_by_key[item_key] = protected_item
            continue
        if release_key not in updated_release_keys:
            records_by_key[item_key] = dict(item)

    for item in updated_items:
        release_key = item.get("release_key")
        if release_key in protected_release_updates and not _is_real_rov_key(item.get("rov_key")):
            continue
        item_key = item.get("row_key") or item.get("release_key")
        if item and item_key:
            records_by_key[item_key] = item

    merged_items = list(records_by_key.values())
    return _compose_release_payload(merged_items, "quick")


def _build_auto_incremental_jira_diagnostics():
    return {
        "jira_checks_total": 0,
        "jira_checks_success": 0,
        "jira_checks_failed": 0,
        "jira_errors": [],
    }


def _record_auto_incremental_jira_check(diagnostics, *, label="", error=None):
    if not isinstance(diagnostics, dict):
        return
    diagnostics["jira_checks_total"] = int(diagnostics.get("jira_checks_total") or 0) + 1
    if error is None:
        diagnostics["jira_checks_success"] = int(diagnostics.get("jira_checks_success") or 0) + 1
        return

    diagnostics["jira_checks_failed"] = int(diagnostics.get("jira_checks_failed") or 0) + 1
    error_text = f"{label}: {error}" if label else str(error)
    errors = diagnostics.setdefault("jira_errors", [])
    if len(errors) < 8:
        errors.append(error_text)


def _apply_auto_incremental_jira_diagnostics(payload, diagnostics):
    payload = dict(payload or {})
    meta = dict(payload.get("meta") or {})
    diagnostics = dict(diagnostics or {})
    meta["auto_incremental_jira_checks_total"] = int(diagnostics.get("jira_checks_total") or 0)
    meta["auto_incremental_jira_checks_success"] = int(diagnostics.get("jira_checks_success") or 0)
    meta["auto_incremental_jira_checks_failed"] = int(diagnostics.get("jira_checks_failed") or 0)
    meta["auto_incremental_jira_errors"] = list(diagnostics.get("jira_errors") or [])
    payload["meta"] = meta
    return payload


def _get_auto_incremental_jira_failure_state(diagnostics):
    total = int((diagnostics or {}).get("jira_checks_total") or 0)
    failed = int((diagnostics or {}).get("jira_checks_failed") or 0)
    success = int((diagnostics or {}).get("jira_checks_success") or 0)
    if total <= 0 or failed <= 0:
        return ""
    if success <= 0 and failed >= total:
        return "failed"
    return "partial"


def _format_auto_incremental_jira_error(diagnostics):
    errors = list((diagnostics or {}).get("jira_errors") or [])
    total = int((diagnostics or {}).get("jira_checks_total") or 0)
    failed = int((diagnostics or {}).get("jira_checks_failed") or 0)
    success = int((diagnostics or {}).get("jira_checks_success") or 0)
    summary = f"Jira checks: success={success}, failed={failed}, total={total}"
    if errors:
        return f"{summary}. " + "; ".join(errors[:4])
    return summary


def _fetch_quick_release_monitor_data(base_items=None):
    current_year = datetime.now().year
    previous_year = current_year - 1
    updated_since = datetime.now() - timedelta(days=QUICK_REFRESH_DAYS)
    updated_records = []
    jira_diagnostics = _build_auto_incremental_jira_diagnostics()

    for (domain, token), prefixes in _get_domain_groups().items():
        resolved_fields = _resolve_field_ids(domain, token)
        release_fields_to_load = {
            "key",
            "summary",
            "status",
            "resolution",
            "assignee",
            "reporter",
            "created",
            "updated",
            "issuetype",
            "priority",
            "issuelinks",
            resolved_fields["planned_prom_start"],
            resolved_fields["planned_prom_end"],
            resolved_fields["system_info"],
            resolved_fields["ke_object"],
            resolved_fields["release_distributive"],
            resolved_fields["delta_release_distributive"],
        }

        domain_release_issues = []
        for prefix in prefixes:
            try:
                issues = _execute_quick_release_search(
                    domain,
                    token,
                    prefix,
                    updated_since,
                    release_fields_to_load,
                )
                logging.info(
                    "Release monitor: quick refresh loaded %s releases for prefix %s",
                    len(issues),
                    prefix,
                )
                _record_auto_incremental_jira_check(
                    jira_diagnostics,
                    label=f"{domain} {prefix} quick releases",
                )
                domain_release_issues.extend((prefix, issue) for issue in issues)
            except Exception as exc:
                logging.error(
                    "Release monitor: quick refresh failed for prefix %s: %s",
                    prefix,
                    exc,
                )
                _record_auto_incremental_jira_check(
                    jira_diagnostics,
                    label=f"{domain} {prefix} quick releases",
                    error=exc,
                )

        release_issues_by_key = {
            issue.get("key"): (prefix, issue)
            for prefix, issue in domain_release_issues
            if issue.get("key")
        }
        restored_rov_records = _restore_missing_rov_links_from_known_state(
            domain,
            token,
            release_issues_by_key,
            base_items or [],
            resolved_fields,
            release_fields_to_load,
            _release_monitor_rov_fields_to_load(resolved_fields),
            context_label="quick",
        )
        domain_release_issues = list(release_issues_by_key.values())

        rov_keys = sorted(
            {
                rov_key
                for _, issue in domain_release_issues
                for rov_key in _extract_release_io_keys(issue)
                if rov_key
            }
        )
        rov_map = {}
        if rov_keys:
            rov_fields_to_load = {
                "key",
                "summary",
                "status",
                "issuetype",
                resolved_fields["rov_start"],
                resolved_fields["rov_end"],
            }
            try:
                rov_issues = _execute_issue_keys_search(domain, token, rov_keys, rov_fields_to_load)
                rov_map = {
                    issue.get("key"): _build_rov_record(issue, domain, resolved_fields)
                    for issue in rov_issues
                }
                rov_map.update(restored_rov_records)
                _record_auto_incremental_jira_check(
                    jira_diagnostics,
                    label=f"{domain} quick linked ROV",
                )
            except Exception as exc:
                logging.error("Release monitor: quick refresh failed to load linked ROV issues from %s: %s", domain, exc)
                rov_map.update(restored_rov_records)
                _record_auto_incremental_jira_check(
                    jira_diagnostics,
                    label=f"{domain} quick linked ROV",
                    error=exc,
                )
        else:
            rov_map.update(restored_rov_records)

        for prefix, issue in domain_release_issues:
            records = _build_release_record(
                issue,
                domain,
                prefix,
                resolved_fields,
                rov_map,
                current_year,
                previous_year,
            )
            if records:
                updated_records.extend(records)

    payload = _merge_release_records(base_items or [], updated_records)
    meta = dict(payload.get("meta") or {})
    meta["quick_jira_checks_total"] = int(jira_diagnostics.get("jira_checks_total") or 0)
    meta["quick_jira_checks_success"] = int(jira_diagnostics.get("jira_checks_success") or 0)
    meta["quick_jira_checks_failed"] = int(jira_diagnostics.get("jira_checks_failed") or 0)
    payload["meta"] = meta
    return payload


def _fetch_incremental_release_monitor_data(base_items=None, lookback_minutes=AUTO_INCREMENTAL_REFRESH_LOOKBACK_MINUTES):
    current_year = datetime.now().year
    previous_year = current_year - 1
    updated_records = []
    jira_diagnostics = _build_auto_incremental_jira_diagnostics()
    auto_retry_kwargs = {
        "retry_server_errors": True,
        "retry_max_attempts": AUTO_INCREMENTAL_SEARCH_MAX_ATTEMPTS,
        "retry_delay_seconds": AUTO_INCREMENTAL_SEARCH_RETRY_DELAY_SECONDS,
    }

    for (domain, token), prefixes in _get_domain_groups().items():
        resolved_fields = _resolve_field_ids(domain, token)
        release_fields_to_load = {
            "key",
            "summary",
            "status",
            "resolution",
            "assignee",
            "reporter",
            "created",
            "updated",
            "issuetype",
            "priority",
            "issuelinks",
            resolved_fields["planned_prom_start"],
            resolved_fields["planned_prom_end"],
            resolved_fields["system_info"],
            resolved_fields["ke_object"],
            resolved_fields["release_distributive"],
            resolved_fields["delta_release_distributive"],
        }
        rov_fields_to_load = {
            "key",
            "summary",
            "status",
            "updated",
            "created",
            "issuetype",
            "issuelinks",
            resolved_fields["rov_start"],
            resolved_fields["rov_end"],
        }

        release_issues_by_key = {}
        updated_rov_issues_by_key = {}
        linked_release_keys = set()

        for prefix in prefixes:
            try:
                issues = _execute_incremental_release_search(
                    domain,
                    token,
                    prefix,
                    lookback_minutes,
                    release_fields_to_load,
                    **auto_retry_kwargs,
                )
                logging.info(
                    "Release monitor: auto incremental loaded %s updated releases for prefix %s",
                    len(issues),
                    prefix,
                )
                _record_auto_incremental_jira_check(jira_diagnostics, label=f"{domain} {prefix} releases")
                for issue in issues:
                    if _issue_type_name(issue) == RELEASE_ISSUE_TYPE and issue.get("key"):
                        release_issues_by_key[issue.get("key")] = (prefix, issue)
            except Exception as exc:
                logging.error(
                    "Release monitor: auto incremental failed to load releases for prefix %s: %s",
                    prefix,
                    exc,
                )
                _record_auto_incremental_jira_check(
                    jira_diagnostics,
                    label=f"{domain} {prefix} releases",
                    error=exc,
                )

            try:
                rov_issues = _execute_incremental_rov_search(
                    domain,
                    token,
                    prefix,
                    lookback_minutes,
                    rov_fields_to_load,
                    **auto_retry_kwargs,
                )
                logging.info(
                    "Release monitor: auto incremental loaded %s updated ROV issues for prefix %s",
                    len(rov_issues),
                    prefix,
                )
                _record_auto_incremental_jira_check(jira_diagnostics, label=f"{domain} {prefix} ROV")
                for issue in rov_issues:
                    if issue.get("key"):
                        updated_rov_issues_by_key[issue.get("key")] = issue
                    linked_release_keys.update(_extract_linked_release_keys_from_rov(issue))
            except Exception as exc:
                logging.error(
                    "Release monitor: auto incremental failed to load ROV issues for prefix %s: %s",
                    prefix,
                    exc,
                )
                _record_auto_incremental_jira_check(
                    jira_diagnostics,
                    label=f"{domain} {prefix} ROV",
                    error=exc,
                )

        missing_release_keys = sorted(
            key for key in linked_release_keys if key and key not in release_issues_by_key
        )
        if missing_release_keys:
            try:
                parent_issues = _execute_issue_keys_search(
                    domain,
                    token,
                    missing_release_keys,
                    release_fields_to_load,
                    **auto_retry_kwargs,
                )
                for issue in parent_issues:
                    if _issue_type_name(issue) != RELEASE_ISSUE_TYPE or not issue.get("key"):
                        continue
                    prefix = str(issue.get("key") or "").split("-", 1)[0]
                    release_issues_by_key[issue.get("key")] = (prefix, issue)
                logging.info(
                    "Release monitor: auto incremental loaded %s parent releases from updated ROV issues",
                    len(parent_issues),
                )
                _record_auto_incremental_jira_check(
                    jira_diagnostics,
                    label=f"{domain} parent releases from ROV",
                )
            except Exception as exc:
                logging.error(
                    "Release monitor: auto incremental failed to load parent releases from updated ROV issues: %s",
                    exc,
                )
                _record_auto_incremental_jira_check(
                    jira_diagnostics,
                    label=f"{domain} parent releases from ROV",
                    error=exc,
                )

        restored_rov_records = _restore_missing_rov_links_from_known_state(
            domain,
            token,
            release_issues_by_key,
            base_items or [],
            resolved_fields,
            release_fields_to_load,
            rov_fields_to_load,
            retry_kwargs=auto_retry_kwargs,
            context_label="auto_incremental",
        )

        release_issues = list(release_issues_by_key.values())
        rov_keys = sorted(
            {
                rov_key
                for _, issue in release_issues
                for rov_key in _extract_release_io_keys(issue)
                if rov_key
            }
            | set(updated_rov_issues_by_key)
        )
        rov_map = {}
        if rov_keys:
            try:
                rov_issues = _execute_issue_keys_search(
                    domain,
                    token,
                    rov_keys,
                    rov_fields_to_load,
                    **auto_retry_kwargs,
                )
                rov_map = {
                    issue.get("key"): _build_rov_record(issue, domain, resolved_fields)
                    for issue in rov_issues
                    if issue.get("key")
                }
                rov_map.update(restored_rov_records)
                _record_auto_incremental_jira_check(jira_diagnostics, label=f"{domain} linked ROV")
            except Exception as exc:
                logging.error("Release monitor: auto incremental failed to load linked ROV issues from %s: %s", domain, exc)
                _record_auto_incremental_jira_check(
                    jira_diagnostics,
                    label=f"{domain} linked ROV",
                    error=exc,
                )
                rov_map = {
                    key: _build_rov_record(issue, domain, resolved_fields)
                    for key, issue in updated_rov_issues_by_key.items()
                }
                rov_map.update(restored_rov_records)
        else:
            rov_map.update(restored_rov_records)

        for prefix, issue in release_issues:
            records = _build_release_record(
                issue,
                domain,
                prefix,
                resolved_fields,
                rov_map,
                current_year,
                previous_year,
            )
            if records:
                updated_records.extend(records)

    return _apply_auto_incremental_jira_diagnostics(
        _merge_release_records(base_items or [], updated_records),
        jira_diagnostics,
    )


def clear_release_monitor_cache():
    global _cached_data, _last_cache_update
    with _cache_lock:
        _cached_data = None
        _last_cache_update = None


def _get_cached_payload_copy():
    if _cached_data is None:
        return None
    manual_overrides = _load_manual_release_overrides()
    return {
        "items": list(_cached_data.get("items", [])),
        "manual_overrides": dict(manual_overrides),
        "summary": dict(_cached_data.get("summary", {})),
        "meta": dict(_cached_data.get("meta", {})),
    }


def _migrate_cached_display_snapshot_pair_if_needed(source_payload=None):
    global _cached_data, _snapshot_requires_display_migration

    if not _snapshot_requires_display_migration or not _is_valid_snapshot_payload(_cached_data):
        return
    trace_started_at = time.monotonic()
    _rm_trace(
        "RM_STARTUP",
        "display_snapshot_migration_start",
        items=_count_payload_items(_cached_data),
    )
    source_payload = source_payload or {}
    migrated = _prepare_accepted_snapshot(
        _cached_data,
        accepted_revision=(source_payload.get("meta") or {}).get("accepted_revision"),
        accepted_at=(source_payload.get("meta") or {}).get("accepted_at"),
    )
    _atomic_write_json(LAST_GOOD_SNAPSHOT_FILE, migrated)
    _atomic_write_json(SNAPSHOT_FILE, migrated)
    _atomic_write_text(REVISION_FILE, (migrated.get("meta") or {}).get("data_revision") or "")
    _cached_data = migrated
    _snapshot_requires_display_migration = False
    _rm_trace(
        "RM_STARTUP",
        "display_snapshot_migration_complete",
        started_at=trace_started_at,
        items=_count_payload_items(migrated),
        revision=(migrated.get("meta") or {}).get("accepted_revision", ""),
    )


def _ensure_cached_payload_loaded_locked():
    global _cached_data, _last_cache_update, _snapshot_requires_display_migration

    if _cached_data is not None:
        _migrate_cached_display_snapshot_pair_if_needed()
        return

    trace_started_at = time.monotonic()
    _rm_trace("RM_STARTUP", "snapshot_load_start")
    disk_payload = _load_snapshot_from_disk()
    if disk_payload is not None:
        _cached_data = _hydrate_release_monitor_payload(disk_payload)
        _last_cache_update = time.time()
        _migrate_cached_display_snapshot_pair_if_needed(disk_payload)
    _rm_trace(
        "RM_STARTUP",
        "snapshot_load_complete",
        started_at=trace_started_at,
        items=_count_payload_items(_cached_data),
        loaded=bool(disk_payload),
    )


def _rebuild_cached_payload_after_state_change_locked():
    global _cached_data

    _ensure_cached_payload_loaded_locked()
    payload = _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload())
    _cached_data = payload
    _mark_release_monitor_state_changed()
    return payload


def _release_monitor_payload_fingerprint(payload):
    normalized = _normalize_release_payload(payload or _build_empty_release_monitor_payload())
    comparable = {
        "items": normalized.get("items", []),
        "summary": normalized.get("summary", {}),
    }
    return json.dumps(comparable, ensure_ascii=False, sort_keys=True, default=str)


def _schedule_unassigned_confluence_auto_sync(
    payload,
    *,
    refresh_mode,
    force_notify_row_keys=None,
):
    try:
        from services.release_monitor_confluence_notification_service import (
            schedule_unassigned_auto_sync,
        )

        schedule_unassigned_auto_sync(
            payload,
            refresh_mode=refresh_mode,
            force_notify_row_keys=force_notify_row_keys,
        )
    except Exception:
        logging.exception(
            "Release monitor: failed to schedule Confluence unassigned auto-sync after %s refresh",
            refresh_mode,
        )


def _run_auto_incremental_release_monitor_refresh():
    global _cached_data, _last_cache_update, _last_auto_incremental_refresh_at

    trace_started_at = time.monotonic()
    try:
        started_at = _format_timestamp()
        _rm_trace("RM_SILENT_REFRESH", "start", started_at_text=started_at)
        with _cache_lock:
            manual_refresh_running = bool(
                (_refresh_thread and _refresh_thread.is_alive())
                or _refresh_status.get("state") == "refreshing"
            )
            if manual_refresh_running:
                logging.info("Release monitor: skipped auto incremental refresh because manual refresh is running")
                _rm_trace(
                    "RM_SILENT_REFRESH",
                    "skip_manual_refresh_running",
                    started_at=trace_started_at,
                )
                return

            _auto_incremental_status.update(
                {
                    "state": "running",
                    "last_started_at": started_at,
                    "last_error": None,
                    "jira_checks_total": 0,
                    "jira_checks_success": 0,
                    "jira_checks_failed": 0,
                }
            )

            base_payload = _get_cached_payload_copy()
            if base_payload is None:
                disk_payload = _load_snapshot_from_disk()
                if disk_payload is not None:
                    base_payload = _hydrate_release_monitor_payload(disk_payload)
                    _cached_data = base_payload
                    _last_cache_update = _get_snapshot_mtime() or time.time()

            if not base_payload or not base_payload.get("items"):
                logging.info("Release monitor: skipped auto incremental refresh because snapshot is empty")
                _auto_incremental_status.update(
                    {
                        "state": "waiting",
                        "last_finished_at": _format_timestamp(),
                        "last_changed": False,
                        "last_error": None,
                    }
                )
                _rm_trace(
                    "RM_SILENT_REFRESH",
                    "waiting_for_snapshot",
                    started_at=trace_started_at,
                )
                return

            base_items = list(base_payload.get("items", []))
            current_meta = dict(base_payload.get("meta", {}))
            before_fingerprint = _release_monitor_payload_fingerprint(base_payload)
            _rm_trace(
                "RM_SILENT_REFRESH",
                "baseline_ready",
                started_at=trace_started_at,
                items=len(base_items),
            )

        jira_started_at = time.monotonic()
        _rm_trace("RM_SILENT_REFRESH", "jira_fetch_start", items=len(base_items))
        data = _fetch_incremental_release_monitor_data(
            base_items,
            lookback_minutes=AUTO_INCREMENTAL_REFRESH_LOOKBACK_MINUTES,
        )
        _rm_trace(
            "RM_SILENT_REFRESH",
            "jira_fetch_complete",
            started_at=jira_started_at,
            fetched_items=_count_payload_items(data),
        )

        now_str = _format_timestamp()
        meta = dict(data.get("meta", {}))
        jira_diagnostics = {
            "jira_checks_total": meta.get("auto_incremental_jira_checks_total", 0),
            "jira_checks_success": meta.get("auto_incremental_jira_checks_success", 0),
            "jira_checks_failed": meta.get("auto_incremental_jira_checks_failed", 0),
            "jira_errors": meta.get("auto_incremental_jira_errors", []),
        }
        jira_failure_state = _get_auto_incremental_jira_failure_state(jira_diagnostics)
        jira_failure_text = _format_auto_incremental_jira_error(jira_diagnostics) if jira_failure_state else None
        if jira_failure_state == "failed":
            logging.warning("Release monitor: auto incremental refresh failed all Jira checks: %s", jira_failure_text)
            _update_auto_incremental_status(
                state="failed",
                last_finished_at=now_str,
                last_changed=False,
                last_error=jira_failure_text,
                jira_checks_total=int(jira_diagnostics.get("jira_checks_total") or 0),
                jira_checks_success=int(jira_diagnostics.get("jira_checks_success") or 0),
                jira_checks_failed=int(jira_diagnostics.get("jira_checks_failed") or 0),
            )
            _rm_trace(
                "RM_SILENT_REFRESH",
                "jira_checks_failed",
                started_at=trace_started_at,
                failed=jira_diagnostics.get("jira_checks_failed", 0),
                total=jira_diagnostics.get("jira_checks_total", 0),
            )
            return

        meta["last_updated"] = now_str
        meta["last_sync_mode"] = "auto_incremental"
        meta["quick_refresh_days"] = QUICK_REFRESH_DAYS
        meta["auto_full_refresh_enabled"] = AUTO_FULL_REFRESH_ENABLED
        meta["auto_full_refresh_hour"] = AUTO_FULL_REFRESH_HOUR
        meta["auto_incremental_refresh_enabled"] = AUTO_INCREMENTAL_REFRESH_ENABLED
        meta["auto_incremental_refresh_interval_seconds"] = AUTO_INCREMENTAL_REFRESH_INTERVAL_SECONDS
        meta["auto_incremental_refresh_lookback_minutes"] = AUTO_INCREMENTAL_REFRESH_LOOKBACK_MINUTES
        meta["last_full_sync"] = current_meta.get("last_full_sync")
        meta["last_quick_sync"] = current_meta.get("last_quick_sync")
        meta["last_auto_incremental_sync"] = now_str
        meta["last_confluence_sync"] = current_meta.get("last_confluence_sync")
        data["meta"] = meta

        manual_overrides = _load_manual_release_overrides()
        finalized = _finalize_release_monitor_payload(data, manual_overrides)
        after_fingerprint = _release_monitor_payload_fingerprint(finalized)
        if before_fingerprint == after_fingerprint:
            logging.info("Release monitor: auto incremental refresh completed without data changes")
            _update_auto_incremental_status(
                state="partial" if jira_failure_state == "partial" else "completed",
                last_finished_at=now_str,
                last_changed=False,
                last_error=jira_failure_text if jira_failure_state == "partial" else None,
                jira_checks_total=int(jira_diagnostics.get("jira_checks_total") or 0),
                jira_checks_success=int(jira_diagnostics.get("jira_checks_success") or 0),
                jira_checks_failed=int(jira_diagnostics.get("jira_checks_failed") or 0),
            )
            _rm_trace(
                "RM_SILENT_REFRESH",
                "complete_without_changes",
                started_at=trace_started_at,
                state="partial" if jira_failure_state == "partial" else "completed",
            )
            if jira_failure_state != "partial":
                _schedule_unassigned_confluence_auto_sync(
                    finalized,
                    refresh_mode="silent",
                )
            return

        with _cache_lock:
            manual_refresh_running = bool(
                (_refresh_thread and _refresh_thread.is_alive())
                or _refresh_status.get("state") == "refreshing"
            )
            if manual_refresh_running:
                logging.info("Release monitor: skipped saving auto incremental refresh because manual refresh started")
                _auto_incremental_status.update(
                    {
                        "state": "skipped",
                        "last_finished_at": _format_timestamp(),
                        "last_changed": False,
                        "last_error": None,
                    }
                )
                _rm_trace(
                    "RM_SILENT_REFRESH",
                    "skip_save_manual_refresh_started",
                    started_at=trace_started_at,
                )
                return
            _cached_data = finalized
            _last_cache_update = time.time()
            _save_snapshot_to_disk(_cached_data)
            _auto_incremental_status.update(
                {
                    "state": "partial" if jira_failure_state == "partial" else "completed",
                    "last_finished_at": now_str,
                    "last_changed": True,
                    "last_error": jira_failure_text if jira_failure_state == "partial" else None,
                    "jira_checks_total": int(jira_diagnostics.get("jira_checks_total") or 0),
                    "jira_checks_success": int(jira_diagnostics.get("jira_checks_success") or 0),
                    "jira_checks_failed": int(jira_diagnostics.get("jira_checks_failed") or 0),
                }
            )

        logging.info(
            "Release monitor: auto incremental refresh saved updated snapshot, items=%s",
            len(finalized.get("items", [])),
        )
        _rm_trace(
            "RM_SILENT_REFRESH",
            "complete_with_changes",
            started_at=trace_started_at,
            items=len(finalized.get("items", [])),
            state="partial" if jira_failure_state == "partial" else "completed",
        )
        if jira_failure_state != "partial":
            _schedule_unassigned_confluence_auto_sync(
                finalized,
                refresh_mode="silent",
            )
    except Exception as exc:
        logging.exception("Release monitor: auto incremental refresh failed")
        _update_auto_incremental_status(
            state="failed",
            last_finished_at=_format_timestamp(),
            last_changed=False,
            last_error=str(exc),
        )
        _rm_trace(
            "RM_SILENT_REFRESH",
            "failed",
            started_at=trace_started_at,
            error_type=type(exc).__name__,
        )
    finally:
        with _cache_lock:
            _last_auto_incremental_refresh_at = time.time()
            final_state = _auto_incremental_status.get("state")
        _rm_trace(
            "RM_SILENT_REFRESH",
            "worker_exit",
            started_at=trace_started_at,
            state=final_state,
        )


def _ensure_scheduler_started():
    global _scheduler_thread, _scheduler_started, _auto_incremental_thread

    if _scheduler_started:
        _rm_trace(
            "RM_SCHEDULER",
            "already_started",
            level=logging.DEBUG,
            thread_alive=bool(_scheduler_thread and _scheduler_thread.is_alive()),
        )
        return

    def _scheduler_loop():
        global _auto_incremental_thread
        _rm_trace("RM_SCHEDULER", "loop_started")
        while True:
            try:
                with _cache_lock:
                    snapshot = _cached_data or _load_snapshot_from_disk() or _build_empty_release_monitor_payload()
                    last_full_sync = (snapshot.get("meta") or {}).get("last_full_sync")
                    running = bool(
                        (_refresh_thread is not None and _refresh_thread.is_alive())
                        or _refresh_status.get("state") == "refreshing"
                    )
                    auto_running = _auto_incremental_thread is not None and _auto_incremental_thread.is_alive()
                    last_auto_ts = _last_auto_incremental_refresh_at

                now = datetime.now()
                last_full_date = None
                if last_full_sync:
                    try:
                        last_full_date = datetime.strptime(last_full_sync, "%d.%m.%Y %H:%M:%S").date()
                    except ValueError:
                        last_full_date = None

                should_run = (
                    AUTO_FULL_REFRESH_ENABLED
                    and now.hour >= AUTO_FULL_REFRESH_HOUR
                    and last_full_date != now.date()
                    and not running
                )

                if should_run:
                    _rm_trace("RM_SCHEDULER", "launch_auto_full")
                    start_release_monitor_refresh(mode="full", trigger="auto")
                    running = True

                now_ts = time.time()
                should_run_auto_incremental = (
                    AUTO_INCREMENTAL_REFRESH_ENABLED
                    and not running
                    and not auto_running
                    and (
                        last_auto_ts is None
                        or (now_ts - last_auto_ts) >= AUTO_INCREMENTAL_REFRESH_INTERVAL_SECONDS
                    )
                )

                if should_run_auto_incremental:
                    with _cache_lock:
                        manual_refresh_running = bool(
                            (_refresh_thread and _refresh_thread.is_alive())
                            or _refresh_status.get("state") == "refreshing"
                        )
                        if not manual_refresh_running:
                            _rm_trace(
                                "RM_SCHEDULER",
                                "launch_silent_refresh",
                                last_auto_ts=last_auto_ts,
                            )
                            _auto_incremental_thread = threading.Thread(
                                target=_run_auto_incremental_release_monitor_refresh,
                                daemon=True,
                                name="release-monitor-silent-refresh",
                            )
                            _auto_incremental_thread.start()
            except Exception:
                logging.exception("Release monitor: refresh scheduler failed")
                _rm_trace("RM_SCHEDULER", "loop_error")

            time.sleep(min(AUTO_REFRESH_CHECK_INTERVAL, AUTO_INCREMENTAL_REFRESH_CHECK_INTERVAL))

    _rm_trace(
        "RM_SCHEDULER",
        "start",
        auto_enabled=AUTO_INCREMENTAL_REFRESH_ENABLED,
        interval_seconds=AUTO_INCREMENTAL_REFRESH_INTERVAL_SECONDS,
    )
    _scheduler_thread = threading.Thread(
        target=_scheduler_loop,
        daemon=True,
        name="release-monitor-scheduler",
    )
    _scheduler_thread.start()
    _scheduler_started = True
    _rm_trace(
        "RM_SCHEDULER",
        "started",
        thread_alive=_scheduler_thread.is_alive(),
    )


def _prepare_display_ready_full_payload(candidate, mode, current_meta):
    data = _compose_release_payload(
        [dict(item) for item in (candidate or {}).get("items", []) if isinstance(item, dict)],
        mode,
    )
    now_str = _format_timestamp()
    meta = dict(data.get("meta", {}))
    meta["last_updated"] = now_str
    meta["last_sync_mode"] = mode
    meta["quick_refresh_days"] = QUICK_REFRESH_DAYS
    meta["auto_full_refresh_enabled"] = AUTO_FULL_REFRESH_ENABLED
    meta["auto_full_refresh_hour"] = AUTO_FULL_REFRESH_HOUR
    meta["auto_incremental_refresh_enabled"] = AUTO_INCREMENTAL_REFRESH_ENABLED
    meta["auto_incremental_refresh_interval_seconds"] = AUTO_INCREMENTAL_REFRESH_INTERVAL_SECONDS
    meta["auto_incremental_refresh_lookback_minutes"] = AUTO_INCREMENTAL_REFRESH_LOOKBACK_MINUTES
    meta["last_full_sync"] = now_str
    meta["last_quick_sync"] = current_meta.get("last_quick_sync")
    meta["last_auto_incremental_sync"] = current_meta.get("last_auto_incremental_sync")
    meta["last_confluence_sync"] = current_meta.get("last_confluence_sync")
    data["meta"] = meta
    return _finalize_release_monitor_payload(data, _load_manual_release_overrides())


def _confirmed_snapshot_timestamp(payload):
    meta = dict((payload or {}).get("meta") or {})
    return meta.get("accepted_at") or meta.get("last_full_sync") or meta.get("last_updated") or ""


def _execute_transactional_full_refresh(*, reliable, base_payload):
    mode = RELIABLE_FULL_REFRESH_MODE if reliable else "full"
    trace_tag = "RM_RELIABLE_REFRESH" if reliable else "RM_FULL_REFRESH"
    trace_started_at = time.monotonic()
    _rm_trace(
        trace_tag,
        "candidate_pipeline_start",
        baseline_items=_count_payload_items(base_payload),
    )
    base_payload = base_payload or _build_empty_release_monitor_payload()
    candidate = _fetch_release_monitor_data(
        reliable=reliable,
        base_items=list(base_payload.get("items") or []),
    )
    _rm_trace(
        trace_tag,
        "candidate_loaded",
        started_at=trace_started_at,
        candidate_items=_count_payload_items(candidate),
    )
    validation_report = _validate_release_candidate(candidate, base_payload)
    _save_candidate_diagnostic(
        candidate,
        validation_report,
        state=validation_report.get("status") or "candidate",
    )
    if validation_report.get("status") != "accepted":
        _rm_trace(
            trace_tag,
            "candidate_rejected",
            started_at=trace_started_at,
            reasons=",".join(
                reason.get("code", "")
                for reason in validation_report.get("reasons") or []
            ) or "unknown",
        )
        primary_reason = next(iter(validation_report.get("reasons") or []), {})
        raise ReleaseMonitorCandidateRejected(
            primary_reason.get("message") or "Release monitor candidate was rejected",
            candidate=candidate,
            validation_report=validation_report,
        )

    try:
        display_payload = _prepare_display_ready_full_payload(
            candidate,
            mode,
            dict(base_payload.get("meta") or {}),
        )
        committed = _commit_accepted_snapshot(display_payload, mode=mode)
    except Exception as exc:
        _save_candidate_diagnostic(
            candidate,
            validation_report,
            state="commit_failed",
            error=str(exc),
        )
        raise
    _rm_trace(
        trace_tag,
        "candidate_committed",
        started_at=trace_started_at,
        items=_count_payload_items(committed),
        revision=(committed.get("meta") or {}).get("accepted_revision", ""),
    )
    return committed, validation_report


def get_release_monitor_data(force_refresh=False):
    global _cached_data, _last_cache_update

    with _cache_lock:
        _ensure_scheduler_started()
        _ensure_cached_payload_loaded_locked()

        now = time.time()
        if (
            not force_refresh
            and _cached_data is not None
            and _last_cache_update is not None
            and (now - _last_cache_update) < DASHBOARD_CACHE_TTL
        ):
            return _cached_data

    with _cache_lock:
        base_payload = copy.deepcopy(_cached_data or _build_empty_release_monitor_payload())

    try:
        data, _ = _execute_transactional_full_refresh(
            reliable=False,
            base_payload=base_payload,
        )
    except (ReleaseMonitorSourceError, ReleaseMonitorCandidateRejected) as exc:
        candidate = getattr(exc, "candidate", None)
        validation_report = getattr(exc, "validation_report", None) or {}
        if candidate is not None and isinstance(exc, ReleaseMonitorSourceError):
            _save_candidate_diagnostic(
                candidate,
                validation_report,
                state="failed",
                error=str(exc),
            )
        if base_payload.get("items"):
            logging.warning("Release monitor: force refresh was not applied; using confirmed snapshot: %s", exc)
            return base_payload
        raise

    with _cache_lock:
        _cached_data = data
        _last_cache_update = time.time()
        return _cached_data


def _run_release_monitor_refresh(mode="full", trigger="manual"):
    global _cached_data, _last_cache_update

    trace_tag = (
        "RM_RELIABLE_REFRESH"
        if mode == RELIABLE_FULL_REFRESH_MODE
        else "RM_FULL_REFRESH"
        if mode == "full"
        else "RM_QUICK_REFRESH"
    )
    trace_started_at = time.monotonic()
    auto_sync_allowed = True
    try:
        logging.info("Release monitor: background %s refresh started", mode)
        _rm_trace(trace_tag, "worker_start", mode=mode, trigger=trigger)
        validation_report = None
        now_str = _format_timestamp()
        with _cache_lock:
            _ensure_cached_payload_loaded_locked()
            base_payload = copy.deepcopy(_cached_data or _build_empty_release_monitor_payload())
            base_items = list(base_payload.get("items", []))
            current_meta = dict(base_payload.get("meta", {}))
            if not base_items:
                disk_payload = _load_snapshot_from_disk()
                if disk_payload:
                    base_payload = _hydrate_release_monitor_payload(disk_payload)
                    base_items = list(base_payload.get("items", []))
                    current_meta = dict(base_payload.get("meta", {}))

        reliable = mode == RELIABLE_FULL_REFRESH_MODE
        if mode == "quick" and base_items:
            manual_overrides = _load_manual_release_overrides()
            data = _fetch_quick_release_monitor_data(base_items)
            auto_sync_allowed = int(
                ((data.get("meta") or {}).get("quick_jira_checks_failed")) or 0
            ) == 0
            now_str = _format_timestamp()
            meta = dict(data.get("meta", {}))
            meta["last_updated"] = now_str
            meta["last_sync_mode"] = mode
            meta["quick_refresh_days"] = QUICK_REFRESH_DAYS
            meta["auto_full_refresh_enabled"] = AUTO_FULL_REFRESH_ENABLED
            meta["auto_full_refresh_hour"] = AUTO_FULL_REFRESH_HOUR
            meta["auto_incremental_refresh_enabled"] = AUTO_INCREMENTAL_REFRESH_ENABLED
            meta["auto_incremental_refresh_interval_seconds"] = AUTO_INCREMENTAL_REFRESH_INTERVAL_SECONDS
            meta["auto_incremental_refresh_lookback_minutes"] = AUTO_INCREMENTAL_REFRESH_LOOKBACK_MINUTES
            meta["last_full_sync"] = current_meta.get("last_full_sync")
            meta["last_quick_sync"] = now_str
            meta["last_auto_incremental_sync"] = current_meta.get("last_auto_incremental_sync")
            meta["last_confluence_sync"] = current_meta.get("last_confluence_sync")
            data["meta"] = meta
            data = _finalize_release_monitor_payload(data, manual_overrides)
            with _cache_lock:
                _cached_data = data
                _last_cache_update = time.time()
                _save_snapshot_to_disk(_cached_data)
        else:
            if mode == "quick":
                mode = "full"
            data, validation_report = _execute_transactional_full_refresh(
                reliable=reliable,
                base_payload=base_payload,
            )
            with _cache_lock:
                _cached_data = data
                _last_cache_update = time.time()

        now_str = _format_timestamp()
        with _cache_lock:
            _refresh_status.update(
                {
                    "state": "completed",
                    "message": "Р”Р°РЅРЅС‹Рµ РїРѕ СЂРµР»РёР·Р°Рј РѕР±РЅРѕРІР»РµРЅС‹",
                    "started_at": _refresh_status.get("started_at"),
                    "finished_at": now_str,
                    "error": None,
                    "mode": mode,
                    "trigger": trigger,
                    "validation_report": validation_report if mode in {"full", RELIABLE_FULL_REFRESH_MODE} else None,
                }
            )
        logging.info("Release monitor: background %s refresh completed, items=%s", mode, len(data.get("items", [])))
        _rm_trace(
            trace_tag,
            "worker_complete",
            started_at=trace_started_at,
            items=len(data.get("items", [])),
            mode=mode,
        )
        if auto_sync_allowed:
            _schedule_unassigned_confluence_auto_sync(
                data,
                refresh_mode=mode,
            )
    except ReleaseMonitorCandidateRejected as exc:
        logging.warning("Release monitor: background %s refresh rejected: %s", mode, exc)
        report = dict(exc.validation_report or {})
        baseline = dict(report.get("baseline") or {})
        candidate_profile = dict(report.get("candidate") or {})
        with _cache_lock:
            _refresh_status.update(
                {
                    "state": "rejected",
                    "message": "Обновление не применено: Jira вернула неполные данные",
                    "finished_at": _format_timestamp(),
                    "error": str(exc),
                    "mode": mode,
                    "trigger": trigger,
                    "validation_report": report,
                    "previous_total": int(baseline.get("total") or 0),
                    "candidate_total": int(candidate_profile.get("total") or 0),
                    "confirmed_snapshot_at": _confirmed_snapshot_timestamp(_cached_data),
                }
            )
        _rm_trace(
            trace_tag,
            "worker_rejected",
            started_at=trace_started_at,
            error_type=type(exc).__name__,
        )
    except ReleaseMonitorSourceError as exc:
        logging.exception("Release monitor: background %s refresh failed on mandatory source", mode)
        _save_candidate_diagnostic(
            exc.candidate,
            state="failed",
            error=str(exc),
        )
        with _cache_lock:
            _refresh_status.update(
                {
                    "state": "failed",
                    "message": "Обновление не применено: обязательный источник Jira недоступен",
                    "finished_at": _format_timestamp(),
                    "error": str(exc),
                    "mode": mode,
                    "trigger": trigger,
                    "previous_total": _count_payload_items(_cached_data),
                    "candidate_total": _count_payload_items(exc.candidate),
                    "confirmed_snapshot_at": _confirmed_snapshot_timestamp(_cached_data),
                }
            )
        _rm_trace(
            trace_tag,
            "worker_source_failed",
            started_at=trace_started_at,
            error_type=type(exc).__name__,
        )
    except Exception as exc:
        logging.exception("Release monitor: background %s refresh failed", mode)
        with _cache_lock:
            _refresh_status.update(
                {
                    "state": "failed",
                    "message": "РћС€РёР±РєР° РѕР±РЅРѕРІР»РµРЅРёСЏ СЂРµР»РёР·РѕРІ",
                    "finished_at": _format_timestamp(),
                    "error": str(exc),
                    "mode": mode,
                    "trigger": trigger,
                }
            )
        _rm_trace(
            trace_tag,
            "worker_failed",
            started_at=trace_started_at,
            error_type=type(exc).__name__,
        )


def start_release_monitor_refresh(mode="full", trigger="manual"):
    global _refresh_thread

    with _cache_lock:
        _ensure_scheduler_started()
        if _refresh_thread and _refresh_thread.is_alive():
            return {
                "started": False,
                "status": dict(_refresh_status),
            }

        _refresh_status.update(
            {
                "state": "refreshing",
                "message": "РРґРµС‚ РїРѕР»РЅРѕРµ РѕР±РЅРѕРІР»РµРЅРёРµ СЂРµР»РёР·РѕРІ РёР· Jira" if mode == "full" else f"РРґРµС‚ Р±С‹СЃС‚СЂРѕРµ РѕР±РЅРѕРІР»РµРЅРёРµ СЂРµР»РёР·РѕРІ Р·Р° РїРѕСЃР»РµРґРЅРёРµ {QUICK_REFRESH_DAYS} РґРЅРµР№",
                "started_at": _format_timestamp(),
                "finished_at": None,
                "error": None,
                "mode": mode,
                "trigger": trigger,
                "validation_report": None,
                "previous_total": None,
                "candidate_total": None,
                "confirmed_snapshot_at": None,
            }
        )
        _refresh_thread = threading.Thread(
            target=_run_release_monitor_refresh,
            kwargs={"mode": mode, "trigger": trigger},
            daemon=True,
            name=f"release-monitor-{mode}-refresh",
        )
        _refresh_thread.start()

        return {
            "started": True,
            "status": dict(_refresh_status),
        }


def _normalize_release_payload(payload):
    if not isinstance(payload, dict):
        return _build_empty_release_monitor_payload()

    normalized_items = [dict(item) for item in (payload.get("items") or []) if isinstance(item, dict)]
    _apply_manual_releases(normalized_items)
    _apply_manual_release_overrides(normalized_items, payload.get("manual_overrides") or {})
    _apply_reviewer_assignments(normalized_items)
    _apply_release_status_consistency(normalized_items)
    _apply_date_overrides(normalized_items)
    _apply_release_attempt_outcomes(normalized_items)
    _apply_duty_schedule_assignments(normalized_items, persist=False)
    _apply_zni_assignments(normalized_items)
    _apply_manual_release_overrides(normalized_items, payload.get("manual_overrides") or {})
    _apply_template_system_classification(normalized_items)
    _apply_work_marks(normalized_items)
    _apply_manual_duplicate_reconciliation(normalized_items)
    _apply_week_control_flags(normalized_items)
    _apply_release_week_buckets(normalized_items)
    _sort_and_number_records(normalized_items)
    current_year = datetime.now().year
    previous_year = current_year - 1

    return {
        "items": normalized_items,
        "manual_overrides": _normalize_manual_release_overrides(payload.get("manual_overrides") or {}),
        "summary": _build_summary(normalized_items, current_year, previous_year),
        "meta": _append_auto_incremental_meta(
            _append_revision_meta(_append_duty_schedule_meta(dict(payload.get("meta", {}))))
        ),
    }


def _finalize_release_monitor_payload(payload, manual_overrides=None):
    payload = dict(payload or {})
    normalized_manual_overrides = _normalize_manual_release_overrides(
        manual_overrides if manual_overrides is not None else payload.get("manual_overrides") or {}
    )
    payload["manual_overrides"] = dict(normalized_manual_overrides)
    finalized_payload = _normalize_release_payload(payload)
    finalized_payload["manual_overrides"] = dict(normalized_manual_overrides)
    return finalized_payload


def _hydrate_release_monitor_payload(payload):
    if payload is None:
        return None
    return _finalize_release_monitor_payload(payload, _load_manual_release_overrides())


def get_release_monitor_refresh_status():
    with _cache_lock:
        _ensure_scheduler_started()
        _reload_snapshot_from_disk_if_newer()

        payload = {
            "status": dict(_refresh_status),
        }
        data = copy.deepcopy(
            _get_cached_payload_copy() or _build_empty_release_monitor_payload()
        )
        data["meta"] = _append_auto_incremental_meta(data.get("meta", {}))
        payload["data"] = data
        persisted_state = str(
            ((_cached_data or {}).get("meta") or {})
            .get("auto_incremental_status", {})
            .get("state", "")
        )
        live_status = data["meta"].get("auto_incremental_status") or {}
        _rm_trace(
            "RM_SILENT_REFRESH",
            "status_poll",
            level=logging.DEBUG,
            live_running=live_status.get("running", False),
            live_state=live_status.get("state", ""),
            persisted_state=persisted_state,
            scheduler_alive=bool(_scheduler_thread and _scheduler_thread.is_alive()),
            worker_alive=bool(_auto_incremental_thread and _auto_incremental_thread.is_alive()),
        )
        return payload


def get_release_monitor_snapshot():
    with _cache_lock:
        _ensure_scheduler_started()
        _reload_snapshot_from_disk_if_newer()
        if _cached_data is None:
            return _build_empty_release_monitor_payload()

        payload = _normalize_release_payload(_get_cached_payload_copy())
        payload["meta"] = {
            **payload.get("meta", {}),
            "is_cached": True,
        }
        return payload


def get_release_monitor_reviewer_options():
    return list(OPLOT_VALUES)


def get_release_monitor_week_control(snapshot=None):
    snapshot = snapshot if isinstance(snapshot, dict) else (get_release_monitor_snapshot() or {})
    items = snapshot.get("items", []) if isinstance(snapshot, dict) else []
    week_start, week_end = _get_current_week_bounds()

    week_items = [
        item for item in items
        if _is_release_assignment_relevant_for_week(item, week_start, week_end)
    ]
    missing_week_items = [
        item for item in week_items
        if not _has_release_responsible(item)
    ]
    assignment_dates = {
        _get_release_start_date(item)
        for item in missing_week_items
        if _get_release_start_date(item)
    }
    if not assignment_dates:
        assignment_dates = {
            _get_release_start_date(item)
            for item in week_items
            if _get_release_start_date(item)
        }
    candidate_groups = _collect_week_candidate_availability(
        week_start,
        week_end,
        target_dates=assignment_dates,
    )
    missing_responsible = []
    for item in missing_week_items:
        release_date = _get_release_start_date(item)
        missing_responsible.append({
            "row_key": _get_assignment_key_for_item(item),
            "release_key": item.get("release_key", ""),
            "rov_key": item.get("rov_key", ""),
            "release_summary": item.get("release_summary", ""),
            "system_name": item.get("system_name", ""),
            "release_status": item.get("release_status", ""),
            "deployment_start": item.get("deployment_start", ""),
            "deployment_end": item.get("deployment_end", ""),
            "release_url": item.get("release_url", ""),
            "rov_url": item.get("rov_url", ""),
            "availability_date": release_date.isoformat() if release_date else "",
            "candidate_availability": _candidate_availability_for_release_date(
                candidate_groups,
                release_date,
            ),
        })

    assigned_load = defaultdict(int)
    for item in week_items:
        responsibles = item.get("psi_responsibles") or []
        if not isinstance(responsibles, list):
            responsibles = [responsibles] if responsibles else []
        for responsible in responsibles:
            responsible_name = str(responsible or "").strip()
            if responsible_name:
                assigned_load[responsible_name] += 1

    available_count = max(1, len(candidate_groups.get("available") or []))
    reserve_allowed = bool(missing_responsible and (len(missing_responsible) / available_count) > 5)

    return {
        "period": {
            "start": week_start.isoformat(),
            "end": week_end.isoformat(),
            "label": f"{week_start.strftime('%d.%m.%Y')} - {week_end.strftime('%d.%m.%Y')}",
        },
        "statistics": {
            "week_releases": len(week_items),
            "missing_responsible": len(missing_responsible),
            "available_candidates": len(candidate_groups.get("available") or []),
            "reserve_candidates": len(candidate_groups.get("reserve") or []),
            "excluded_candidates": len(candidate_groups.get("excluded") or []),
            "reserve_allowed": reserve_allowed,
        },
        "missing_responsible": missing_responsible,
        "candidates": candidate_groups,
        "assigned_load": dict(sorted(assigned_load.items(), key=lambda pair: (-pair[1], pair[0]))),
    }


def _release_assignment_center_period_stats(items, reference_dt=None):
    reference_dt = reference_dt or datetime.now()
    reference_date = reference_dt.date()
    week_start, week_end = _get_current_week_bounds(reference_dt)
    quarter_start_month = ((reference_date.month - 1) // 3) * 3 + 1
    quarter_start = reference_date.replace(month=quarter_start_month, day=1)
    if quarter_start_month == 10:
        quarter_end = reference_date.replace(
            year=reference_date.year + 1,
            month=1,
            day=1,
        ) - timedelta(days=1)
    else:
        quarter_end = reference_date.replace(
            month=quarter_start_month + 3,
            day=1,
        ) - timedelta(days=1)

    stats = {
        name: {
            "active": 0,
            "week": 0,
            "quarter": 0,
            "year": 0,
        }
        for name in OPLOT_VALUES
    }

    for item in items or []:
        if item.get("is_cancelled"):
            continue
        release_date = _get_release_start_date(item)
        if not release_date:
            continue
        raw_responsibles = item.get("psi_responsibles") or []
        if not isinstance(raw_responsibles, list):
            raw_responsibles = [raw_responsibles] if raw_responsibles else []
        responsibles = {
            str(value or "").strip()
            for value in raw_responsibles
            if str(value or "").strip() in stats
        }
        if not responsibles:
            continue

        is_week = week_start <= release_date <= week_end
        is_quarter = quarter_start <= release_date <= quarter_end
        is_year = release_date.year == reference_date.year
        is_active = is_week and not item.get("is_final")
        for responsible in responsibles:
            values = stats[responsible]
            if is_active:
                values["active"] += 1
            if is_week:
                values["week"] += 1
            if is_quarter:
                values["quarter"] += 1
            if is_year:
                values["year"] += 1
    return stats


def get_release_monitor_assignment_center_data():
    snapshot = get_release_monitor_snapshot() or {}
    items = snapshot.get("items", []) if isinstance(snapshot, dict) else []
    control = get_release_monitor_week_control(snapshot=snapshot)
    period_stats = _release_assignment_center_period_stats(items)
    meta = dict(snapshot.get("meta") or {})

    availability_by_name = {}
    enriched_candidates = {"available": [], "reserve": [], "excluded": []}
    for group_name in ("available", "reserve", "excluded"):
        for candidate in (control.get("candidates") or {}).get(group_name, []):
            candidate_name = str(candidate.get("name") or "").strip()
            if not candidate_name:
                continue
            enriched = {
                **candidate,
                "metrics": dict(period_stats.get(candidate_name) or {}),
            }
            enriched_candidates[group_name].append(enriched)
            availability_by_name[candidate_name] = {
                "availability": group_name,
                "reasons": list(candidate.get("reasons") or []),
                "statuses": list(candidate.get("statuses") or []),
            }

    for group_name in enriched_candidates:
        enriched_candidates[group_name].sort(
            key=lambda candidate: (
                int((candidate.get("metrics") or {}).get("active") or 0),
                int((candidate.get("metrics") or {}).get("week") or 0),
                int((candidate.get("metrics") or {}).get("year") or 0),
                str(candidate.get("name") or ""),
            )
        )

    item_by_row_key = {
        _get_assignment_key_for_item(item): item
        for item in items
        if _get_assignment_key_for_item(item)
    }
    missing_items = []
    for missing in control.get("missing_responsible") or []:
        row_key = str(missing.get("row_key") or "").strip()
        source_item = item_by_row_key.get(row_key) or {}
        missing_items.append({
            **missing,
            "ke_id": source_item.get("ke_id", ""),
            "ke_name": source_item.get("ke_name", ""),
            "release_version": source_item.get("release_version", ""),
            "rov_status": source_item.get("rov_status", ""),
            "row_state": source_item.get("row_state", ""),
            "is_final": bool(source_item.get("is_final")),
            "duty_owner": source_item.get("psi_owner", ""),
            "duty_owner_source": source_item.get("psi_owner_source", ""),
        })

    week_start, _ = _get_current_week_bounds()
    week_key = f"{week_start.isocalendar().year}-W{week_start.isocalendar().week:02d}"
    view_payload = {
        "week_key": week_key,
        "missing_responsible": missing_items,
        "candidates": enriched_candidates,
        "statistics": control.get("statistics") or {},
        "employee_metrics": period_stats,
    }
    view_revision = hashlib.sha256(
        json.dumps(view_payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()

    return {
        **control,
        "missing_responsible": missing_items,
        "candidates": enriched_candidates,
        "employee_metrics": period_stats,
        "availability_by_name": availability_by_name,
        "meta": {
            "data_revision": str(meta.get("data_revision") or _read_data_revision() or ""),
            "accepted_revision": str(meta.get("accepted_revision") or ""),
            "snapshot_at": _confirmed_snapshot_timestamp(snapshot),
            "week_key": week_key,
            "view_revision": view_revision,
            "generated_at": _format_timestamp(),
        },
    }


def _extract_json_object(text_value):
    text_value = str(text_value or "").strip()
    if not text_value:
        return {}
    try:
        parsed = json.loads(text_value)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        pass

    match = re.search(r"\{.*\}", text_value, re.DOTALL)
    if not match:
        return {}
    try:
        parsed = json.loads(match.group(0))
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def _build_responsible_history(items, candidate_names):
    candidate_set = set(candidate_names)
    history = {
        candidate: {
            "total": 0,
            "by_system": defaultdict(int),
            "by_prefix": defaultdict(int),
        }
        for candidate in candidate_names
    }

    for item in items:
        responsibles = item.get("psi_responsibles") or []
        if not isinstance(responsibles, list):
            responsibles = [responsibles] if responsibles else []

        prefix = str(item.get("release_key") or "").split("-", 1)[0]
        system_name = str(item.get("system_name") or "").strip() or prefix or "unknown"
        for responsible in responsibles:
            responsible_name = str(responsible or "").strip()
            if responsible_name not in candidate_set:
                continue
            history[responsible_name]["total"] += 1
            history[responsible_name]["by_system"][system_name] += 1
            history[responsible_name]["by_prefix"][prefix or "unknown"] += 1

    return {
        candidate: {
            "total": values["total"],
            "by_system": dict(sorted(values["by_system"].items(), key=lambda pair: (-pair[1], pair[0]))[:8]),
            "by_prefix": dict(sorted(values["by_prefix"].items(), key=lambda pair: (-pair[1], pair[0]))[:8]),
        }
        for candidate, values in history.items()
    }


def _normalize_giga_recommendation_name(value, allowed_names):
    value = str(value or "").strip()
    if not value:
        return ""
    if value in allowed_names:
        return value
    matched = _match_oplot_name(value)
    return matched if matched in allowed_names else ""


WEEK_RECOMMENDATION_MAX_LOAD_GAP = 2
WEEK_RECOMMENDATION_CLM_PRIMARY = (
    "Кашкин С.Н.",
    "Гапоненко Д.А.",
)


def _week_recommendation_system_group(item):
    item = item or {}
    system_name = str(item.get("system_name") or "").strip()
    normalized_system = _normalize_text(system_name).lower()
    release_prefix = str(item.get("release_key") or "").split("-", 1)[0].upper()

    if "ai-" in normalized_system or "агент" in normalized_system:
        return "AI"
    if "emrm" in normalized_system or release_prefix == "EMRM":
        return "EMRM"
    if "clm" in normalized_system or release_prefix == "SMECLM":
        return "CLM"
    if "аист" in normalized_system or release_prefix == "SMECSC":
        return "AIST"
    return system_name or release_prefix or "OTHER"


def _week_candidate_affinity(candidate_name, item, history):
    candidate_history = (history or {}).get(candidate_name) or {}
    system_name = str((item or {}).get("system_name") or "").strip()
    release_prefix = str((item or {}).get("release_key") or "").split("-", 1)[0]
    return int((candidate_history.get("by_system") or {}).get(system_name) or 0) + int(
        (candidate_history.get("by_prefix") or {}).get(release_prefix) or 0
    )


def _week_balanced_candidate_pool(item, allowed_candidates, projected_load):
    candidates = list(dict.fromkeys(
        str(name or "").strip()
        for name in (allowed_candidates or [])
        if str(name or "").strip()
    ))
    if not candidates:
        return []

    if _week_recommendation_system_group(item) == "EMRM":
        non_clm_primary = [
            name for name in candidates
            if name not in WEEK_RECOMMENDATION_CLM_PRIMARY
        ]
        if non_clm_primary:
            candidates = non_clm_primary

    minimum_load = min(int(projected_load.get(name, 0) or 0) for name in candidates)
    return [
        name for name in candidates
        if int(projected_load.get(name, 0) or 0)
        <= minimum_load + WEEK_RECOMMENDATION_MAX_LOAD_GAP
    ]


def _select_week_balanced_candidate(item, candidates, projected_load, history):
    system_group = _week_recommendation_system_group(item)

    def rank(candidate_name):
        load = int(projected_load.get(candidate_name, 0) or 0)
        affinity = _week_candidate_affinity(candidate_name, item, history)
        if system_group == "CLM":
            return (
                0 if candidate_name in WEEK_RECOMMENDATION_CLM_PRIMARY else 1,
                load,
                -affinity,
                candidate_name,
            )
        return (load, -affinity, candidate_name)

    return min(candidates, key=rank) if candidates else ""


def _balance_week_responsible_recommendations(
    recommendations,
    *,
    missing_by_row_key,
    allowed_candidates,
    allowed_candidates_by_row=None,
    current_week_load,
    history,
):
    projected_load = {
        candidate: int((current_week_load or {}).get(candidate) or 0)
        for candidate in allowed_candidates or []
    }
    missing_order = {
        row_key: index
        for index, row_key in enumerate(missing_by_row_key)
    }
    balanced = []
    correction_count = 0

    for recommendation in sorted(
        recommendations or [],
        key=lambda item: missing_order.get(item.get("row_key"), len(missing_order)),
    ):
        row_key = recommendation.get("row_key")
        source_item = missing_by_row_key.get(row_key) or {}
        if allowed_candidates_by_row is not None and row_key in allowed_candidates_by_row:
            row_allowed_candidates = list(allowed_candidates_by_row.get(row_key) or [])
        else:
            row_allowed_candidates = list(allowed_candidates or [])
        candidate_pool = _week_balanced_candidate_pool(
            source_item,
            row_allowed_candidates,
            projected_load,
        )
        if not candidate_pool:
            continue

        original_name = str(recommendation.get("recommended") or "").strip()
        selected_name = original_name
        if selected_name not in candidate_pool:
            selected_name = _select_week_balanced_candidate(
                source_item,
                candidate_pool,
                projected_load,
                history,
            )

        next_item = dict(recommendation)
        next_item["recommended"] = selected_name
        if selected_name != original_name:
            correction_count += 1
            original_load = int(projected_load.get(original_name, 0) or 0)
            selected_load = int(projected_load.get(selected_name, 0) or 0)
            next_item["reason"] = (
                f"Скорректировано по недельной нагрузке: "
                f"{selected_name} ({selected_load}) вместо "
                f"{original_name} ({original_load})."
            )
            next_item["confidence"] = "high"
            next_item["balance_adjusted"] = True
        else:
            next_item["balance_adjusted"] = False

        backup_pool = [name for name in candidate_pool if name != selected_name]
        current_backup = str(next_item.get("backup") or "").strip()
        if current_backup not in backup_pool:
            current_backup = _select_week_balanced_candidate(
                source_item,
                backup_pool,
                projected_load,
                history,
            )
        next_item["backup"] = current_backup

        projected_load[selected_name] = int(projected_load.get(selected_name, 0) or 0) + 1
        next_item["projected_week_load"] = projected_load[selected_name]
        balanced.append(next_item)

    return balanced, correction_count


def get_release_monitor_week_responsible_recommendations():
    snapshot = get_release_monitor_snapshot() or {}
    items = snapshot.get("items", []) if isinstance(snapshot, dict) else []
    control = get_release_monitor_week_control(snapshot=snapshot)
    missing = control.get("missing_responsible") or []
    week_start, week_end = _get_current_week_bounds()
    week_items = [
        {
            "row_key": _get_assignment_key_for_item(item),
            "release_key": item.get("release_key", ""),
            "rov_key": item.get("rov_key", ""),
            "release_summary": item.get("release_summary", ""),
            "system_name": item.get("system_name", ""),
            "release_status": item.get("release_status", ""),
            "deployment_start": item.get("deployment_start", ""),
            "deployment_end": item.get("deployment_end", ""),
            "responsibles": item.get("psi_responsibles", []),
            "row_state": item.get("row_state", ""),
            "is_overdue": bool(item.get("is_overdue")),
        }
        for item in items
        if _is_release_assignment_relevant_for_week(item, week_start, week_end)
    ]

    reserve_allowed = bool(control.get("statistics", {}).get("reserve_allowed"))
    allowed_candidates = []
    allowed_candidates_by_row = {}
    missing_with_rules = []
    for missing_item in missing:
        row_key = str(missing_item.get("row_key") or "").strip()
        candidate_availability = missing_item.get("candidate_availability") or {}
        row_available = []
        row_reserve = []
        row_excluded = []
        for candidate_name in OPLOT_VALUES:
            availability = str(
                (candidate_availability.get(candidate_name) or {}).get("availability")
                or "available"
            ).strip()
            if availability == "excluded":
                row_excluded.append(candidate_name)
            elif availability == "reserve":
                row_reserve.append(candidate_name)
            else:
                row_available.append(candidate_name)

        row_allowed = list(row_available)
        if reserve_allowed:
            row_allowed.extend(name for name in row_reserve if name not in row_allowed)
        allowed_candidates_by_row[row_key] = row_allowed
        for candidate_name in row_allowed:
            if candidate_name not in allowed_candidates:
                allowed_candidates.append(candidate_name)
        missing_with_rules.append({
            **missing_item,
            "allowed_candidates": row_allowed,
            "reserve_candidates": row_reserve,
            "excluded_candidates": row_excluded,
        })

    if missing and not allowed_candidates:
        return {
            "control": control,
            "recommendations": [],
            "source": "rules",
            "message": "Нет доступных кандидатов по графику дежурств.",
        }

    history = _build_responsible_history(items, allowed_candidates)
    recommendation_context = {
        "period": control.get("period", {}),
        "rules": {
            "allowed_candidates": allowed_candidates,
            "reserve_allowed": reserve_allowed,
            "do_not_assign_checkers": True,
            "max_week_load_gap": WEEK_RECOMMENDATION_MAX_LOAD_GAP,
            "clm_primary_candidates": list(WEEK_RECOMMENDATION_CLM_PRIMARY),
            "clm_primary_only_within_load_gap": True,
            "emrm_use_clm_primary_only_if_no_other_candidate": True,
        },
        "current_week_load": control.get("assigned_load", {}),
        "history": history,
        "current_week_releases": week_items,
        "releases_without_responsible": missing_with_rules,
    }

    try:
        from services.gigachat_service import GIGA_HELPER
    except Exception as exc:
        logging.warning("Release week control: failed to import GigaChat helper: %s", exc)
        return {
            "control": control,
            "recommendations": [],
            "source": "unavailable",
            "message": f"GigaChat недоступен: {exc}",
        }

    if not getattr(GIGA_HELPER, "client", None):
        return {
            "control": control,
            "recommendations": [],
            "source": "unavailable",
            "message": "GigaChat недоступен или не инициализирован.",
        }

    prompt = f"""
Ты помогаешь планировать ответственных за ПСИ по релизам.

Сначала сформируй краткую операционную сводку по current_week_releases.
Если releases_without_responsible пустой, recommendations верни пустым массивом и в summary напиши, что по ответственным критичных пробелов нет.
Если releases_without_responsible не пустой, предложи ответственного только для релизов из этого списка.
Для каждого релиза используй только ФИО из его поля allowed_candidates.
Нельзя предлагать сотрудника из excluded_candidates конкретного релиза.
reserve_candidates конкретного релиза уже включены в allowed_candidates только когда rules.reserve_allowed=true.
Проверяющих не назначай и не анализируй.

Учитывай:
- историю назначений по похожим системам, prefix и category/system_name;
- текущую недельную нагрузку current_week_load;
- равномерность распределения считать обязательным ограничением, а не пожеланием;
- сначала найти минимальную нагрузку среди допустимых кандидатов;
- нельзя рекомендовать кандидата, чья текущая или проектная нагрузка больше минимума более чем на rules.max_week_load_gap;
- учитывать рекомендации в текущем ответе последовательно: после выбора сотрудника мысленно увеличить его нагрузку на 1;
- для CLM Кашкин С.Н. и Гапоненко Д.А. являются приоритетными специалистами только внутри допустимого коридора нагрузки;
- для EMRM Кашкин С.Н. и Гапоненко Д.А. допустимы только если нет других разрешенных кандидатов;
- причины исключения из графика.

Верни строго JSON без markdown:
{{
  "recommendations": [
    {{
      "row_key": "...",
      "release_key": "...",
      "recommended": "ФИО из allowed_candidates",
      "backup": "ФИО из allowed_candidates или пусто",
      "confidence": "high|medium|low",
      "reason": "коротко почему выбран кандидат"
    }}
  ],
  "summary": "краткая сводка по неделе и назначениям"
}}

Данные:
{json.dumps(recommendation_context, ensure_ascii=False, indent=2)}
"""

    try:
        response = GIGA_HELPER.client.chat(prompt)
        content = response.choices[0].message.content
        parsed = _extract_json_object(content)
    except Exception as exc:
        logging.warning("Release week control: GigaChat recommendation failed: %s", exc)
        return {
            "control": control,
            "recommendations": [],
            "source": "error",
            "message": f"Ошибка GigaChat: {exc}",
        }

    raw_recommendations = parsed.get("recommendations") if isinstance(parsed, dict) else []
    if not isinstance(raw_recommendations, list):
        raw_recommendations = []

    missing_by_row_key = {
        item.get("row_key"): item
        for item in missing_with_rules
        if item.get("row_key")
    }
    normalized_recommendations = []
    for raw_item in raw_recommendations:
        if not isinstance(raw_item, dict):
            continue
        row_key = str(raw_item.get("row_key") or "").strip()
        if row_key not in missing_by_row_key:
            continue
        row_allowed_set = set(allowed_candidates_by_row.get(row_key) or [])
        recommended = _normalize_giga_recommendation_name(
            raw_item.get("recommended"),
            row_allowed_set,
        )
        if not recommended:
            continue
        backup = _normalize_giga_recommendation_name(
            raw_item.get("backup"),
            row_allowed_set,
        )
        confidence = str(raw_item.get("confidence") or "medium").strip().lower()
        if confidence not in {"high", "medium", "low"}:
            confidence = "medium"
        source_item = missing_by_row_key[row_key]
        normalized_recommendations.append({
            "row_key": row_key,
            "release_key": source_item.get("release_key", ""),
            "rov_key": source_item.get("rov_key", ""),
            "recommended": recommended,
            "backup": backup,
            "confidence": confidence,
            "reason": str(raw_item.get("reason") or "").strip(),
        })

    normalized_recommendations, correction_count = _balance_week_responsible_recommendations(
        normalized_recommendations,
        missing_by_row_key=missing_by_row_key,
        allowed_candidates=allowed_candidates,
        allowed_candidates_by_row=allowed_candidates_by_row,
        current_week_load=control.get("assigned_load") or {},
        history=history,
    )
    summary = str(parsed.get("summary") or "").strip() if isinstance(parsed, dict) else ""
    if correction_count:
        balance_note = (
            f"По недельной нагрузке скорректировано рекомендаций: {correction_count}."
        )
        summary = f"{summary} {balance_note}".strip()

    return {
        "control": control,
        "recommendations": normalized_recommendations,
        "source": "gigachat",
        "summary": summary,
        "message": "" if normalized_recommendations else "GigaChat не вернул применимых рекомендаций.",
    }


def upload_release_monitor_duty_schedules(uploaded_files):
    global _cached_data, _last_cache_update

    files = [file_storage for file_storage in (uploaded_files or []) if getattr(file_storage, "filename", "")]
    if not files:
        raise ValueError("\u041d\u0435 \u0432\u044b\u0431\u0440\u0430\u043d\u044b \u0444\u0430\u0439\u043b\u044b \u0433\u0440\u0430\u0444\u0438\u043a\u0430")

    existing_payload = _load_duty_schedule_payload()
    merged_payload = dict(existing_payload)
    uploaded_names = []
    parsed_months = []
    warnings = []

    for file_storage in files:
        filename = str(getattr(file_storage, "filename", "") or "").strip()
        if not filename:
            continue

        file_bytes = file_storage.read()
        if not file_bytes:
            continue

        parsed_payload = _parse_duty_schedule_workbook(file_bytes, filename)
        parsed_payload["last_upload"] = _format_timestamp()
        merged_payload = _merge_duty_schedule_payload(merged_payload, parsed_payload)
        uploaded_names.append(filename)
        parsed_months.extend(parsed_payload.get("months", []))
        warnings.extend(parsed_payload.get("warnings", []))

    if not uploaded_names:
        raise ValueError("\u041d\u0435 \u0443\u0434\u0430\u043b\u043e\u0441\u044c \u043f\u0440\u043e\u0447\u0438\u0442\u0430\u0442\u044c \u043d\u0438 \u043e\u0434\u0438\u043d \u0444\u0430\u0439\u043b \u0433\u0440\u0430\u0444\u0438\u043a\u0430")

    merged_payload["last_upload"] = _format_timestamp()
    _save_duty_schedule_payload(merged_payload)

    with _cache_lock:
        _ensure_scheduler_started()
        if _cached_data is None:
            disk_payload = _load_snapshot_from_disk()
            if disk_payload is not None:
                _cached_data = _hydrate_release_monitor_payload(disk_payload)
                _last_cache_update = _get_snapshot_mtime() or time.time()

        applied_count = 0
        duty_debug_rows = []
        if _cached_data is not None:
            items = _cached_data.get("items") or []
            _apply_reviewer_assignments(items)
            _apply_date_overrides(items)
            duty_result = _apply_duty_schedule_assignments(items, persist=False, force=True, debug_limit=20)
            applied_count = duty_result.get("applied_count", 0) if isinstance(duty_result, dict) else int(duty_result or 0)
            duty_debug_rows = duty_result.get("debug_rows", []) if isinstance(duty_result, dict) else []
            _sort_and_number_records(items)
            meta = _cached_data.setdefault("meta", {})
            meta["last_duty_schedule_upload"] = merged_payload.get("last_upload")
            _mark_release_monitor_state_changed()

        payload = _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload())

    return {
        "uploaded_files": uploaded_names,
        "parsed_months": list(dict.fromkeys(parsed_months)),
        "warnings": warnings,
        "applied_count": applied_count,
        "duty_debug_rows": duty_debug_rows,
        "data": payload,
    }


def create_release_monitor_zni(release_key, reporter=""):
    global _cached_data, _last_cache_update

    release_key = str(release_key or "").strip()
    reporter = str(reporter or "").strip()
    if not release_key:
        raise ValueError("Не указан ключ строки релиза")

    with _cache_lock:
        _ensure_scheduler_started()
        if _cached_data is None:
            disk_payload = _load_snapshot_from_disk()
            if disk_payload is not None:
                _cached_data = _hydrate_release_monitor_payload(disk_payload)
                _last_cache_update = _get_snapshot_mtime() or time.time()

        if _cached_data is None:
            raise ValueError("Таблица релизов еще не загружена")

        items = _cached_data.get("items") or []
        _apply_reviewer_assignments(items)
        _apply_date_overrides(items)
        _apply_duty_schedule_assignments(items, persist=False)
        _apply_zni_assignments(items)
        target_item = None
        for item in items:
            if _get_assignment_key_for_item(item) == release_key:
                target_item = item
                break

        if not target_item:
            raise ValueError("Не удалось найти строку релиза")
        if target_item.get("zni_key"):
            return {
                "issue": {
                    "key": target_item.get("zni_key"),
                    "url": target_item.get("zni_url", ""),
                    "already_exists": True,
                },
                "data": _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload()),
            }

    issue_item = _prepare_item_for_zni_creation(target_item)
    issue = create_oplot_release_issue(issue_item, reporter_name=reporter)

    with _cache_lock:
        assignments = _load_zni_assignments()
        assignments[release_key] = {
            "key": issue.get("key", ""),
            "url": issue.get("url", ""),
            "summary": issue.get("summary", ""),
            "created_at": _format_timestamp(),
            "assignee": issue.get("assignee", ""),
            "reporter": issue.get("reporter", ""),
        }
        _save_zni_assignments(assignments)

        overrides = _load_manual_release_overrides()
        current_override = dict(overrides.get(release_key) or {})
        for field_name in (
            "clear_zni",
            "zni_key",
            "zni_url",
            "base_zni_key",
            "base_zni_url",
        ):
            current_override.pop(field_name, None)
        current_override = _normalize_manual_release_override(current_override)
        if current_override:
            overrides[release_key] = current_override
        else:
            overrides.pop(release_key, None)
        _save_manual_release_overrides(overrides)

        if _cached_data is not None:
            _cached_data["manual_overrides"] = dict(overrides)
            items = _cached_data.get("items") or []
            for item in items:
                if _get_assignment_key_for_item(item) == release_key:
                    item["zni_key"] = issue.get("key", "")
                    item["zni_url"] = issue.get("url", "")
                    item["base_zni_key"] = issue.get("key", "")
                    item["base_zni_url"] = issue.get("url", "")
                    item["manual_clear_zni"] = False
                    break
            _mark_release_monitor_state_changed()

        payload = _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload())

    return {
        "issue": issue,
        "data": payload,
    }


def set_release_monitor_work_mark(row_key, mark=""):
    global _cached_data

    row_key = str(row_key or "").strip()
    if not row_key:
        raise ValueError("Не указан ключ строки релиза")

    mark = str(mark or "").strip()

    with _cache_lock:
        if _cached_data is None:
            disk_payload = _load_snapshot_from_disk()
            if disk_payload is not None:
                _cached_data = _hydrate_release_monitor_payload(disk_payload)

        current_payload = _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload())
        target_item = next(
            (
                item for item in current_payload.get("items") or []
                if _get_assignment_key_for_item(item) == row_key
            ),
            None,
        )
        if target_item is None:
            raise ValueError("Строка релиза не найдена в актуальном наборе данных")

        marks = _load_work_marks()
        changed = False
        if mark:
            existing_mark = str((marks.get(row_key) or {}).get("mark") or "").strip()
            next_payload = {
                "mark": mark,
                "updated_at": _format_timestamp(),
            }
            changed = existing_mark != mark
            if changed:
                marks[row_key] = next_payload
        else:
            changed = row_key in marks
            marks.pop(row_key, None)
        if changed:
            _save_work_marks(marks)

        if _cached_data is not None:
            for item in _cached_data.get("items") or []:
                if _get_assignment_key_for_item(item) == row_key:
                    item["work_mark"] = mark
                    item["work_mark_updated_at"] = marks.get(row_key, {}).get("updated_at", "") if mark else ""
                    break
            if changed:
                _mark_release_monitor_state_changed()
        else:
            if changed:
                _touch_release_monitor_revision()

        payload = _normalize_release_payload(_get_cached_payload_copy() or current_payload)

    return {
        "row_key": row_key,
        "work_mark": mark,
        "data": payload,
    }


def set_release_monitor_rollout_notes(release_key, enabled=False, level=""):
    global _cached_data

    release_key = str(release_key or "").strip()
    if not release_key:
        raise ValueError("Не указан ключ строки релиза")

    enabled = bool(enabled)
    level = str(level or "").strip().lower()
    if enabled and level not in {"success", "warning", "danger", "none"}:
        level = "warning"
    if not enabled:
        level = ""
    should_clear_work_mark = bool(enabled and level in {"success", "warning", "danger"})
    work_mark_cleared = False
    work_mark_cleanup_failed = False

    with _cache_lock:
        flags = _load_rollout_note_flags()
        if enabled:
            flags[release_key] = {
                "has_rollout_notes": level != "none",
                "rollout_notes_level": level,
                "updated_at": _format_timestamp(),
            }
        else:
            flags.pop(release_key, None)
        _save_rollout_note_flags(flags)

        if should_clear_work_mark:
            try:
                work_mark_cleared = _clear_release_work_mark(release_key)
            except Exception as exc:
                work_mark_cleanup_failed = True
                logging.warning(
                    "Release monitor: failed to clear work mark after rollout color for %s: %s",
                    release_key,
                    exc,
                )

        if _cached_data is None:
            disk_payload = _load_snapshot_from_disk()
            if disk_payload is not None:
                _cached_data = _hydrate_release_monitor_payload(disk_payload)

        if _cached_data is not None:
            for item in _cached_data.get("items") or []:
                if _get_assignment_key_for_item(item) == release_key:
                    item["has_rollout_notes"] = bool(enabled and level != "none")
                    item["rollout_notes_level"] = level
                    if should_clear_work_mark and not work_mark_cleanup_failed:
                        item["work_mark"] = ""
                        item["work_mark_updated_at"] = ""
                    break
            _mark_release_monitor_state_changed()
        else:
            _touch_release_monitor_revision()

        payload = _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload())

    return {
        "release_key": release_key,
        "has_rollout_notes": bool(enabled and level != "none"),
        "rollout_notes_level": level,
        "work_mark_cleared": work_mark_cleared,
        "work_mark_cleanup_failed": work_mark_cleanup_failed,
        "data": payload,
    }


def set_release_monitor_date_override(release_key, start_value="", end_value="", reset=False):
    global _cached_data, _last_cache_update

    release_key = str(release_key or "").strip()
    if not release_key:
        raise ValueError("Не указан ключ строки релиза")

    overrides = _load_date_overrides()
    if reset:
        overrides.pop(release_key, None)
    else:
        normalized_start = _format_release_monitor_date(_parse_release_monitor_date(start_value)) if start_value else ""
        normalized_end = _format_release_monitor_date(_parse_release_monitor_date(end_value)) if end_value else ""
        if not normalized_start and not normalized_end:
            overrides.pop(release_key, None)
        else:
            overrides[release_key] = {
                "start": normalized_start,
                "end": normalized_end,
                "updated_at": _format_timestamp(),
            }

    _save_date_overrides(overrides)
    _remove_row_from_manual_order(release_key)

    with _cache_lock:
        if _cached_data is None:
            disk_payload = _load_snapshot_from_disk()
            if disk_payload:
                _cached_data = _hydrate_release_monitor_payload(disk_payload)
                _last_cache_update = _get_snapshot_mtime() or time.time()

        if _cached_data is not None:
            items = _cached_data.get("items") or []
            _apply_reviewer_assignments(items)
            _apply_date_overrides(items)
            _apply_duty_schedule_assignments(items, persist=False)
            _sort_and_number_records(items)
            _mark_release_monitor_state_changed()

        payload = _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload())
        return payload


def set_release_monitor_manual_override(
    release_key,
    *,
    release_summary="",
    release_version="",
    release_dist_url="",
    ke="",
    zni_key="",
    zni_url="",
    clear_zni=False,
    reset=False,
):
    global _cached_data, _last_cache_update

    release_key = str(release_key or "").strip()
    if not release_key:
        raise ValueError("Не указан ключ строки релиза")

    with _cache_lock:
        if _cached_data is None:
            disk_payload = _load_snapshot_from_disk()
            if disk_payload is not None:
                _cached_data = _hydrate_release_monitor_payload(disk_payload)
                _last_cache_update = _get_snapshot_mtime() or time.time()

        if _cached_data is None:
            raise ValueError("Таблица релизов еще не загружена")

        current_payload = _get_cached_payload_copy() or _build_empty_release_monitor_payload()
        current_items = current_payload.get("items") or []
        target_item = next(
            (item for item in current_items if _get_assignment_key_for_item(item) == release_key),
            None,
        )
        if not target_item:
            raise ValueError("Не удалось найти строку релиза")

        base_summary = str(target_item.get("base_release_summary") or target_item.get("release_summary") or "").strip()
        base_version = str(target_item.get("base_release_version") or target_item.get("release_version") or "").strip()
        base_url = str(target_item.get("base_release_dist_url") or target_item.get("release_dist_url") or "").strip()
        base_ke = str(target_item.get("base_ke") or target_item.get("ke") or "").strip()
        base_zni_key = str(target_item.get("base_zni_key") or target_item.get("zni_key") or "").strip()
        base_zni_url = str(target_item.get("base_zni_url") or target_item.get("zni_url") or "").strip()

        if reset:
            overrides = _load_manual_release_overrides()
            removed_override = dict(overrides.get(release_key) or {})
            overrides.pop(release_key, None)
            _save_manual_release_overrides(overrides)

            if _cached_data is not None:
                _cached_data["manual_overrides"] = dict(overrides)
                items = _cached_data.get("items") or []
                reset_base_summary = str(
                    target_item.get("base_release_summary")
                    or removed_override.get("base_release_summary")
                    or ""
                ).strip()
                reset_base_version = str(
                    target_item.get("base_release_version")
                    or removed_override.get("base_release_version")
                    or ""
                ).strip()
                reset_base_url = str(
                    target_item.get("base_release_dist_url")
                    or removed_override.get("base_release_dist_url")
                    or ""
                ).strip()
                reset_base_ke = str(
                    target_item.get("base_ke")
                    or removed_override.get("base_ke")
                    or ""
                ).strip()
                reset_base_zni_key = str(
                    target_item.get("base_zni_key")
                    or removed_override.get("base_zni_key")
                    or ""
                ).strip()
                reset_base_zni_url = str(
                    target_item.get("base_zni_url")
                    or removed_override.get("base_zni_url")
                    or ""
                ).strip()
                _apply_reviewer_assignments(items)
                _apply_date_overrides(items)
                _apply_duty_schedule_assignments(items, persist=False)
                _apply_zni_assignments(items)
                for item in items:
                    if _get_assignment_key_for_item(item) != release_key:
                        continue
                    item["base_release_summary"] = reset_base_summary
                    item["base_release_version"] = reset_base_version
                    item["base_release_dist_url"] = reset_base_url
                    item["base_ke"] = reset_base_ke
                    item["base_zni_key"] = reset_base_zni_key
                    item["base_zni_url"] = reset_base_zni_url
                    item["release_summary"] = reset_base_summary
                    item["release_version"] = reset_base_version
                    item["release_dist_url"] = reset_base_url
                    item["ke"] = reset_base_ke
                    item["zni_key"] = reset_base_zni_key
                    item["zni_url"] = reset_base_zni_url
                    if isinstance(item.get("base_release_name_lines"), list):
                        item["release_name_lines"] = list(item.get("base_release_name_lines") or [])
                    break
                _apply_manual_release_overrides(items)
                _sort_and_number_records(items)
                _mark_release_monitor_state_changed()

            payload = _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload())
            return payload

        normalized_summary = str(release_summary or "").strip()
        normalized_version = str(release_version or "").strip()
        normalized_url = _normalize_artifact_url(str(release_dist_url or "").strip())
        normalized_ke = str(ke or "").strip()
        normalized_zni_key = str(zni_key or "").strip()
        normalized_zni_url = str(zni_url or "").strip()
        normalized_clear_zni = bool(clear_zni)
        zni_cache_removed = False
        if normalized_clear_zni:
            zni_cache_removed = _remove_zni_assignment_for_keys(
                release_key,
                target_item.get("release_key"),
            )
            if zni_cache_removed:
                _clear_cached_zni_fields_for_keys(release_key, target_item.get("release_key"))
                base_zni_key = ""
                base_zni_url = ""

        overrides = _load_manual_release_overrides()
        current_override = dict(overrides.get(release_key) or {})
        if zni_cache_removed:
            current_override.pop("base_zni_key", None)
            current_override.pop("base_zni_url", None)
        if base_summary and not str(current_override.get("base_release_summary") or "").strip():
            current_override["base_release_summary"] = base_summary
        if base_version and not str(current_override.get("base_release_version") or "").strip():
            current_override["base_release_version"] = base_version
        if base_url and not str(current_override.get("base_release_dist_url") or "").strip():
            current_override["base_release_dist_url"] = base_url
        if base_ke and not str(current_override.get("base_ke") or "").strip():
            current_override["base_ke"] = base_ke
        if base_zni_key and not str(current_override.get("base_zni_key") or "").strip():
            current_override["base_zni_key"] = base_zni_key
        if base_zni_url and not str(current_override.get("base_zni_url") or "").strip():
            current_override["base_zni_url"] = base_zni_url

        if normalized_summary and normalized_summary != base_summary:
            current_override["release_summary"] = normalized_summary
        else:
            current_override.pop("release_summary", None)

        if normalized_version and normalized_version != base_version:
            current_override["release_version"] = normalized_version
        else:
            current_override.pop("release_version", None)

        if normalized_url and normalized_url != base_url:
            current_override["release_dist_url"] = normalized_url
        else:
            current_override.pop("release_dist_url", None)

        if normalized_ke and normalized_ke != base_ke:
            current_override["ke"] = _format_ke_id(normalized_ke) if re.fullmatch(r"\d+", normalized_ke) else normalized_ke
        else:
            current_override.pop("ke", None)

        if normalized_clear_zni and zni_cache_removed:
            current_override.pop("clear_zni", None)
            current_override.pop("zni_key", None)
            current_override.pop("zni_url", None)
        elif normalized_clear_zni and base_zni_key:
            current_override["clear_zni"] = True
            current_override.pop("zni_key", None)
            current_override.pop("zni_url", None)
        else:
            current_override.pop("clear_zni", None)
            if normalized_zni_key and normalized_zni_key != base_zni_key:
                current_override["zni_key"] = normalized_zni_key
                derived_url = _resolve_manual_zni_url(normalized_zni_key, normalized_zni_url)
                if derived_url and derived_url != base_zni_url:
                    current_override["zni_url"] = derived_url
                else:
                    current_override.pop("zni_url", None)
            else:
                current_override.pop("zni_key", None)
                current_override.pop("zni_url", None)

        current_override = _normalize_manual_release_override(current_override)
        if current_override:
            current_override["updated_at"] = _format_timestamp()
            overrides[release_key] = current_override
        else:
            overrides.pop(release_key, None)

        _save_manual_release_overrides(overrides)

        if _cached_data is None:
            disk_payload = _load_snapshot_from_disk()
            if disk_payload is not None:
                _cached_data = _hydrate_release_monitor_payload(disk_payload)

        if _cached_data is not None:
            _cached_data["manual_overrides"] = dict(overrides)
            items = _cached_data.get("items") or []
            _apply_reviewer_assignments(items)
            _apply_date_overrides(items)
            _apply_duty_schedule_assignments(items, persist=False)
            _apply_zni_assignments(items)
            _apply_manual_release_overrides(items)
            _sort_and_number_records(items)
            _mark_release_monitor_state_changed()

        payload = _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload())
        return payload


def _normalize_manual_release_input_fields(data, *, partial=False):
    if not isinstance(data, dict):
        return {}

    fields = {}
    for field_name in MANUAL_RELEASE_SCALAR_FIELDS:
        if field_name not in data:
            continue
        if field_name == "year":
            raw_year = str(data.get(field_name) or "").strip()
            if raw_year:
                fields[field_name] = raw_year
            elif not partial:
                fields[field_name] = ""
            continue
        normalized_value = _normalize_manual_scalar_field(field_name, data.get(field_name))
        if normalized_value or not partial:
            fields[field_name] = normalized_value

    for field_name in MANUAL_OVERRIDE_DICT_FIELDS:
        if field_name not in data:
            continue
        normalized_value = _normalize_manual_dict_field(data.get(field_name))
        if normalized_value or not partial:
            fields[field_name] = normalized_value

    return fields


def _normalize_manual_override_input_fields(data):
    if not isinstance(data, dict):
        return {}

    fields = {}
    for field_name in MANUAL_OVERRIDE_SCALAR_FIELDS:
        if field_name not in data:
            continue
        fields[field_name] = _normalize_manual_scalar_field(field_name, data.get(field_name))

    for field_name in MANUAL_OVERRIDE_DICT_FIELDS:
        if field_name not in data:
            continue
        fields[field_name] = _normalize_manual_dict_field(data.get(field_name))

    if "clear_zni" in data:
        fields["clear_zni"] = bool(data.get("clear_zni"))

    return fields


def _manual_release_form_fields_from_item(item):
    if not isinstance(item, dict):
        return {}

    return {
        "release_key": str(item.get("release_key") or "").strip(),
        "release_type": normalize_release_type(item.get("release_type"), default="release"),
        "release_summary": str(item.get("release_summary") or "").strip(),
        "deployment_start": str(item.get("deployment_start") or "").strip(),
        "deployment_end": str(item.get("deployment_end") or "").strip(),
        "rov_key": str(item.get("rov_key") or "").strip(),
        "release_url": str(item.get("release_url") or "").strip(),
        "rov_url": str(item.get("rov_url") or "").strip(),
        "release_status": str(item.get("release_status") or "").strip(),
        "rov_status": str(item.get("rov_status") or "").strip(),
        "ke_id": str(item.get("ke_id") or "").strip(),
        "ke": str(item.get("ke") or "").strip(),
        "release_version": str(item.get("release_version") or "").strip(),
        "release_dist_url": str(item.get("release_dist_url") or "").strip(),
        "system_name": str(item.get("system_name") or "").strip(),
        "zni_key": str(item.get("zni_key") or "").strip(),
        "zni_url": str(item.get("zni_url") or "").strip(),
    }


def _fetch_jira_release_records_for_manual_lookup(release_key):
    normalized_key = str(release_key or "").strip().upper()
    if not normalized_key:
        raise ValueError("release_key is required")
    if not re.fullmatch(r"[A-Z][A-Z0-9]+-\d+", normalized_key):
        raise ValueError("release_key format is invalid")

    prefix = normalized_key.split("-", 1)[0]
    domain, token = get_jira_domain_and_token(normalized_key)
    resolved_fields = _resolve_field_ids(domain, token)
    release_fields_to_load = _release_monitor_release_fields_to_load(resolved_fields)
    release_issues = _execute_issue_keys_search(domain, token, [normalized_key], release_fields_to_load)
    release_issue = next(
        (
            issue
            for issue in release_issues
            if str(issue.get("key") or "").upper() == normalized_key
            and _issue_type_name(issue) == RELEASE_ISSUE_TYPE
        ),
        None,
    )
    if not release_issue:
        return []

    rov_map = {}
    rov_keys = sorted(key for key in _extract_release_io_keys(release_issue) if key)
    if rov_keys:
        rov_fields_to_load = _release_monitor_rov_fields_to_load(resolved_fields)
        rov_issues = _execute_issue_keys_search(domain, token, rov_keys, rov_fields_to_load)
        rov_map = {
            issue.get("key"): _build_rov_record(issue, domain, resolved_fields)
            for issue in rov_issues
            if issue.get("key")
        }

    current_year = datetime.now().year
    previous_year = current_year - 1
    return _build_release_record(
        release_issue,
        domain,
        prefix,
        resolved_fields,
        rov_map,
        current_year,
        previous_year,
    )


def lookup_release_monitor_manual_release_jira(release_key):
    normalized_key = str(release_key or "").strip().upper()
    if not normalized_key:
        raise ValueError("release_key is required")

    records = _fetch_jira_release_records_for_manual_lookup(normalized_key)
    form_fields = _manual_release_form_fields_from_item(records[0]) if records else {"release_key": normalized_key}

    with _cache_lock:
        _ensure_cached_payload_loaded_locked()
        current_payload = _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload())
        warnings = _find_manual_release_duplicate_warnings(form_fields, current_payload.get("items") or [])

    return {
        "found": bool(records),
        "release_key": normalized_key,
        "fields": form_fields,
        "records": [_manual_release_form_fields_from_item(item) for item in records],
        "warnings": warnings,
    }


def _autofill_manual_release_fields_from_jira(fields):
    fields = dict(fields or {})
    release_key = str(fields.get("release_key") or "").strip()
    if not release_key:
        return fields

    try:
        records = _fetch_jira_release_records_for_manual_lookup(release_key)
    except Exception as exc:
        logging.warning("Release monitor: manual release Jira autofill failed for %s: %s", release_key, exc)
        return fields

    if len(records) != 1:
        return fields

    jira_fields = _manual_release_form_fields_from_item(records[0])
    for field_name, value in jira_fields.items():
        value = str(value or "").strip()
        if not value:
            continue
        current_value = str(fields.get(field_name) or "").strip()
        if field_name == "release_key":
            fields[field_name] = value
        elif field_name == "release_type":
            if not current_value or (current_value == "release" and value != "release"):
                fields[field_name] = value
        elif not current_value:
            fields[field_name] = value
    return fields


def create_release_monitor_manual_release(data, updated_by=""):
    global _cached_data

    if isinstance(data, dict) and data.get("release_type") and not normalize_release_type(data.get("release_type")):
        raise ValueError("release_type is invalid")
    fields = _normalize_manual_release_input_fields(data or {}, partial=False)
    if not str(fields.get("release_key") or "").strip():
        raise ValueError("release_key is required")
    fields = _autofill_manual_release_fields_from_jira(fields)
    fields["release_type"] = normalize_release_type(fields.get("release_type"), default="release")
    if fields.get("deployment_start") and not fields.get("deployment_end"):
        fields["deployment_end"] = fields.get("deployment_start")

    now_text = _format_timestamp()
    row_key = f"manual::{uuid4().hex}"
    manual_release = {
        **fields,
        "row_key": row_key,
        "created_at": now_text,
        "updated_at": now_text,
        "updated_by": str(updated_by or (data or {}).get("updated_by") or "").strip(),
    }

    validation_errors = _validate_manual_release_payload(manual_release)
    if validation_errors:
        raise ValueError("; ".join(validation_errors))

    with _cache_lock:
        _ensure_cached_payload_loaded_locked()
        current_payload = _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload())
        duplicate_warnings = _find_manual_release_duplicate_warnings(
            manual_release,
            current_payload.get("items") or [],
            include_manual=True,
        )
        blocking_duplicate = next((warning for warning in duplicate_warnings if warning.get("blocking")), None)
        if blocking_duplicate:
            duplicate_row = blocking_duplicate.get("row_key") or ""
            raise ValueError(f"Релиз уже есть в таблице: {duplicate_row}".strip())

        manual_payload = _load_manual_releases_payload()
        manual_payload.setdefault("items", {})[row_key] = manual_release
        manual_payload["meta"] = {
            **(manual_payload.get("meta") or {}),
            "updated_at": now_text,
            "updated_by": manual_release.get("updated_by", ""),
        }
        _save_manual_releases_payload(manual_payload)

        payload = _rebuild_cached_payload_after_state_change_locked()
        return {
            "row_key": row_key,
            "manual_release": _normalize_manual_release(row_key, manual_release),
            "warnings": duplicate_warnings,
            "data": payload,
        }


def update_release_monitor_manual_release(row_key, data, updated_by=""):
    row_key = str(row_key or "").strip()
    if not row_key.startswith("manual::"):
        raise ValueError("Manual release row_key must start with manual::")
    if isinstance(data, dict) and data.get("row_key") and str(data.get("row_key")).strip() != row_key:
        raise ValueError("row_key is immutable")

    if isinstance(data, dict) and data.get("release_type") and not normalize_release_type(data.get("release_type")):
        raise ValueError("release_type is invalid")
    fields = _normalize_manual_release_input_fields(data or {}, partial=True)
    if not fields and not updated_by:
        raise ValueError("No manual release fields were provided")
    if "release_type" in fields:
        fields["release_type"] = normalize_release_type(fields.get("release_type"))
        if not fields["release_type"]:
            raise ValueError("release_type is invalid")

    now_text = _format_timestamp()
    with _cache_lock:
        manual_payload = _load_manual_releases_payload()
        current = dict((manual_payload.get("items") or {}).get(row_key) or {})
        if not current:
            raise ValueError("Manual release was not found")

        date_changed = any(field_name in fields for field_name in ("deployment_start", "deployment_end"))
        current.update(fields)
        current["row_key"] = row_key
        current["updated_at"] = now_text
        current["updated_by"] = str(updated_by or (data or {}).get("updated_by") or current.get("updated_by") or "").strip()

        if current.get("deployment_start") and not current.get("deployment_end"):
            current["deployment_end"] = current.get("deployment_start")

        validation_errors = _validate_manual_release_payload(current)
        if validation_errors:
            raise ValueError("; ".join(validation_errors))

        current_payload = _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload())
        duplicate_warnings = _find_manual_release_duplicate_warnings(
            current,
            current_payload.get("items") or [],
            include_manual=True,
            ignore_row_key=row_key,
        )
        blocking_duplicate = next((warning for warning in duplicate_warnings if warning.get("blocking")), None)
        if blocking_duplicate:
            duplicate_row = blocking_duplicate.get("row_key") or ""
            raise ValueError(f"Релиз уже есть в таблице: {duplicate_row}".strip())

        _remove_row_from_manual_order(row_key) if date_changed else None
        manual_payload.setdefault("items", {})[row_key] = current
        manual_payload["meta"] = {
            **(manual_payload.get("meta") or {}),
            "updated_at": now_text,
            "updated_by": current.get("updated_by", ""),
        }
        _save_manual_releases_payload(manual_payload)

        payload = _rebuild_cached_payload_after_state_change_locked()
        return {
            "row_key": row_key,
            "manual_release": _normalize_manual_release(row_key, current),
            "warnings": duplicate_warnings,
            "data": payload,
        }


def update_release_monitor_manual_override_fields(row_key, fields, updated_by=""):
    global _cached_data

    row_key = str(row_key or "").strip()
    if not row_key:
        raise ValueError("Release row_key is required")

    if isinstance(fields, dict) and fields.get("release_type") and not normalize_release_type(fields.get("release_type")):
        raise ValueError("release_type is invalid")
    normalized_fields = _normalize_manual_override_input_fields(fields or {})
    if not normalized_fields:
        raise ValueError("No manual override fields were provided")
    zni_key_was_provided = "zni_key" in normalized_fields
    zni_url_was_provided = "zni_url" in normalized_fields

    with _cache_lock:
        _ensure_cached_payload_loaded_locked()
        current_payload = _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload())
        target_item = next(
            (item for item in current_payload.get("items") or [] if _get_assignment_key_for_item(item) == row_key),
            None,
        )
        if not target_item:
            raise ValueError("Release row was not found")

        overrides = _load_manual_release_overrides()
        current_override = dict(overrides.get(row_key) or {})

        for field_name, value in normalized_fields.items():
            if field_name == "clear_zni":
                if value:
                    _remove_zni_assignment_for_keys(row_key, target_item.get("release_key"))
                    _clear_cached_zni_fields_for_keys(row_key, target_item.get("release_key"))
                    current_override.pop("clear_zni", None)
                    current_override.pop("zni_key", None)
                    current_override.pop("zni_url", None)
                    current_override.pop("base_zni_key", None)
                    current_override.pop("base_zni_url", None)
                else:
                    current_override.pop("clear_zni", None)
                continue

            if field_name in MANUAL_OVERRIDE_DICT_FIELDS:
                if value:
                    current_override[field_name] = value
                else:
                    current_override.pop(field_name, None)
                continue

            if value:
                current_override[field_name] = value
            else:
                current_override.pop(field_name, None)

        if zni_key_was_provided and not zni_url_was_provided:
            current_override.pop("zni_url", None)

        current_override = _normalize_manual_release_override(current_override)
        if current_override:
            current_override["updated_at"] = _format_timestamp()
            current_override["updated_by"] = str(updated_by or (fields or {}).get("updated_by") or current_override.get("updated_by") or "").strip()
            overrides[row_key] = current_override
        else:
            overrides.pop(row_key, None)

        _save_manual_release_overrides(overrides)
        if _cached_data is not None:
            _cached_data["manual_overrides"] = dict(overrides)
        payload = _rebuild_cached_payload_after_state_change_locked()
        return payload


def reset_release_monitor_manual_override(row_key):
    global _cached_data

    row_key = str(row_key or "").strip()
    if not row_key:
        raise ValueError("Release row_key is required")

    with _cache_lock:
        overrides = _load_manual_release_overrides()
        overrides.pop(row_key, None)
        _save_manual_release_overrides(overrides)
        if _cached_data is not None:
            _cached_data["manual_overrides"] = dict(overrides)
        payload = _rebuild_cached_payload_after_state_change_locked()
        return payload


def sync_release_monitor_jira_fields(row_key="", release_key="", release_version="", ke=""):
    """Update snapshot base fields with data just read from Jira for one release row."""
    global _cached_data, _last_cache_update

    normalized_row_key = str(row_key or "").strip()
    normalized_release_key = str(release_key or "").strip()
    normalized_version = str(release_version or "").strip()
    normalized_ke = str(ke or "").strip()
    if normalized_ke:
        normalized_ke = _format_ke_id(normalized_ke) if re.fullmatch(r"(?:CI)?\d+", normalized_ke, re.IGNORECASE) else normalized_ke

    if not normalized_row_key and not normalized_release_key:
        return {"updated": False, "fields": {}}
    if not normalized_version and not normalized_ke:
        return {
            "updated": False,
            "row_key": normalized_row_key,
            "release_key": normalized_release_key,
            "fields": {},
        }

    with _cache_lock:
        if _cached_data is None:
            disk_payload = _load_snapshot_from_disk()
            if disk_payload is not None:
                _cached_data = _hydrate_release_monitor_payload(disk_payload)
                _last_cache_update = _get_snapshot_mtime() or time.time()

        if _cached_data is None:
            return {
                "updated": False,
                "row_key": normalized_row_key,
                "release_key": normalized_release_key,
                "fields": {},
            }

        items = _cached_data.get("items") or []
        target_item = next(
            (
                item for item in items
                if normalized_row_key and _get_assignment_key_for_item(item) == normalized_row_key
            ),
            None,
        )
        if target_item is None and normalized_release_key:
            target_item = next(
                (
                    item for item in items
                    if str(item.get("release_key") or "").strip() == normalized_release_key
                ),
                None,
            )

        if target_item is None:
            return {
                "updated": False,
                "row_key": normalized_row_key,
                "release_key": normalized_release_key,
                "fields": {},
            }

        changed_fields = {}

        def update_base_field(field, base_field, manual_field, value):
            if not value:
                return
            current_base = str(target_item.get(base_field) or target_item.get(field) or "").strip()
            if current_base == value:
                return
            target_item[base_field] = value
            if not str(target_item.get(manual_field) or "").strip():
                target_item[field] = value
            changed_fields[field] = value
            changed_fields[base_field] = value

        update_base_field("release_version", "base_release_version", "manual_release_version", normalized_version)
        update_base_field("ke", "base_ke", "manual_ke", normalized_ke)

        if changed_fields:
            _save_snapshot_to_disk(_cached_data)
            _last_cache_update = _get_snapshot_mtime() or time.time()

        normalized_payload = _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload())
        normalized_items = normalized_payload.get("items") or []
        effective_row_key = normalized_row_key or _get_assignment_key_for_item(target_item)
        normalized_item = next(
            (
                item for item in normalized_items
                if _get_assignment_key_for_item(item) == effective_row_key
            ),
            None,
        )

        return {
            "updated": bool(changed_fields),
            "row_key": effective_row_key,
            "release_key": normalized_release_key or str(target_item.get("release_key") or "").strip(),
            "fields": changed_fields,
            "item": normalized_item or {},
            "data_revision": (normalized_payload.get("meta") or {}).get("data_revision", ""),
        }


def set_release_monitor_manual_distribution_override(release_key, release_version="", ke=""):
    global _cached_data, _last_cache_update

    release_key = str(release_key or "").strip()
    normalized_version = str(release_version or "").strip()
    normalized_ke = str(ke or "").strip()
    if normalized_ke:
        normalized_ke = _format_ke_id(normalized_ke) if re.fullmatch(r"(?:CI)?\d+", normalized_ke, re.IGNORECASE) else normalized_ke

    if not release_key:
        raise ValueError("Не указан ключ строки релиза")
    if not normalized_version:
        raise ValueError("Укажите версию сборки")
    if not normalized_ke:
        raise ValueError("Укажите КЭ дистрибутива")

    with _cache_lock:
        if _cached_data is None:
            disk_payload = _load_snapshot_from_disk()
            if disk_payload is not None:
                _cached_data = _hydrate_release_monitor_payload(disk_payload)
                _last_cache_update = _get_snapshot_mtime() or time.time()

        if _cached_data is None:
            raise ValueError("Таблица релизов еще не загружена")

        current_payload = _get_cached_payload_copy() or _build_empty_release_monitor_payload()
        current_items = current_payload.get("items") or []
        target_item = next(
            (item for item in current_items if _get_assignment_key_for_item(item) == release_key),
            None,
        )
        if not target_item:
            raise ValueError("Не удалось найти строку релиза")

        base_version = str(target_item.get("base_release_version") or target_item.get("release_version") or "").strip()
        base_ke = str(target_item.get("base_ke") or target_item.get("ke") or "").strip()

        overrides = _load_manual_release_overrides()
        current_override = dict(overrides.get(release_key) or {})
        if base_version and not str(current_override.get("base_release_version") or "").strip():
            current_override["base_release_version"] = base_version
        if base_ke and not str(current_override.get("base_ke") or "").strip():
            current_override["base_ke"] = base_ke

        current_override["release_version"] = normalized_version
        current_override["ke"] = normalized_ke

        current_override = _normalize_manual_release_override(current_override)
        if current_override:
            current_override["updated_at"] = _format_timestamp()
            overrides[release_key] = current_override
        else:
            overrides.pop(release_key, None)

        _save_manual_release_overrides(overrides)

        _cached_data["manual_overrides"] = dict(overrides)
        items = _cached_data.get("items") or []
        _apply_reviewer_assignments(items)
        _apply_date_overrides(items)
        _apply_duty_schedule_assignments(items, persist=False)
        _apply_zni_assignments(items)
        _apply_manual_release_overrides(items)
        _sort_and_number_records(items)
        _mark_release_monitor_state_changed()

        payload = _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload())
        return payload


def sync_release_monitor_assignments_from_confluence(year):
    global _cached_data, _last_cache_update

    year = int(year or datetime.now().year)
    with _cache_lock:
        _ensure_scheduler_started()
        if _cached_data is None:
            disk_payload = _load_snapshot_from_disk()
            if disk_payload is not None:
                _cached_data = _hydrate_release_monitor_payload(disk_payload)
                _last_cache_update = time.time()

        if _cached_data is None:
            raise ValueError("Таблица релизов еще не загружена. Сначала выполните обновление релизов.")

        payload = _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload())
        items = [item for item in payload.get("items", []) if int(item.get("year", 0) or 0) == year]
        if not items:
            raise ValueError(f"Для {year} года нет строк для выгрузки в Confluence")

        page_data = _fetch_confluence_release_page(year)
        replacement_table = _build_confluence_release_table(items, year)
        updated_storage, replaced = _replace_release_monitor_table_in_storage(page_data["storage_html"], replacement_table)
        if not replaced:
            updated_storage = replacement_table

        token = str(TOKENS.get("confluence_delta_token", "") or "").strip()
        url = f"{CONFLUENCE_DELTA_BASE}/rest/api/content/{page_data['page_id']}"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "Accept": "application/json",
        }
        next_version = max(int(page_data.get("version") or 0), 1) + 1
        page_title = page_data["title"] or f"Блок релизов {year}"
        payload_body = {
            "id": page_data["page_id"],
            "type": "page",
            "title": page_title,
            "version": {
                "number": next_version,
                "minorEdit": True,
                "message": f"Автообновление таблицы релизов за {year} год",
            },
            "body": {
                "storage": {
                    "value": updated_storage,
                    "representation": "storage",
                }
            },
        }

        response = requests.put(url, headers=headers, json=payload_body, verify=False, timeout=60)
        if not response.ok:
            detail = _extract_confluence_error_detail(response)
            logging.error(
                "Confluence release push failed: page_id=%s year=%s status=%s detail=%s payload_keys=%s",
                page_data["page_id"],
                year,
                response.status_code,
                detail,
                list(payload_body.keys()),
            )
            raise ValueError(f"Confluence PUT failed ({response.status_code}): {detail}")
        result_data = response.json()
        published_version = int(((result_data.get("version") or {}).get("number")) or next_version)
        page_url = f"{CONFLUENCE_DELTA_BASE}/pages/viewpage.action?pageId={page_data['page_id']}"

        _cached_data.setdefault("meta", {})["last_confluence_sync"] = _format_timestamp()
        _save_snapshot_to_disk(_cached_data)

        return {
            "year": year,
            "page_id": page_data["page_id"],
            "page_url": page_url,
            "page_title": page_title,
            "page_version": published_version,
            "rows_pushed": len(items),
            "replaced": replaced,
            "data": _normalize_release_payload(_get_cached_payload_copy() or _build_empty_release_monitor_payload()),
        }

    with _cache_lock:
        _ensure_scheduler_started()
        if _cached_data is None:
            disk_payload = _load_snapshot_from_disk()
            if disk_payload:
                _cached_data = _hydrate_release_monitor_payload(disk_payload)
                _last_cache_update = time.time()

        if _cached_data is None:
            raise ValueError("РўР°Р±Р»РёС†Р° СЂРµР»РёР·РѕРІ РµС‰Рµ РЅРµ Р·Р°РіСЂСѓР¶РµРЅР°. РЎРЅР°С‡Р°Р»Р° РІС‹РїРѕР»РЅРёС‚Рµ РѕР±РЅРѕРІР»РµРЅРёРµ СЂРµР»РёР·РѕРІ.")

        assignments = _load_reviewer_assignments()
        matched_rows = 0

        for item in _cached_data.get("items", []):
            if int(item.get("year", 0) or 0) != year:
                continue

            row_key = _get_assignment_key_for_item(item)
            release_key = item.get("release_key")
            fallback_row_key = f"{release_key}::no-rov" if release_key else ""
            source = (
                confluence_assignments.get(row_key)
                or confluence_assignments.get(fallback_row_key)
            )
            if not source:
                continue

            current_assignment = dict(assignments.get(row_key) or assignments.get(release_key) or {})
            current_responsibles = current_assignment.get("responsibles")
            if not isinstance(current_responsibles, list):
                current_responsibles = []
            current_responsibles = [str(value or "").strip() for value in current_responsibles if str(value or "").strip()]
            current_checker = str(current_assignment.get("checker", "") or "").strip()

            item_responsibles = item.get("psi_responsibles")
            if not isinstance(item_responsibles, list):
                item_responsibles = []
            item_responsibles = [str(value or "").strip() for value in item_responsibles if str(value or "").strip()]
            item_checker = str(item.get("psi_checker", "") or "").strip()

            effective_responsibles = current_responsibles or item_responsibles
            effective_checker = current_checker or item_checker
            applied = False

            if not effective_responsibles and source.get("responsibles"):
                new_responsibles = list(source.get("responsibles", []))
                current_assignment["responsibles"] = new_responsibles
                item["psi_responsibles"] = new_responsibles
                applied = True

            if not effective_checker and source.get("checker"):
                new_checker = str(source.get("checker", "") or "").strip()
                current_assignment["checker"] = new_checker
                item["psi_checker"] = new_checker
                applied = True

            if (
                current_assignment.get("reviewer")
                or current_assignment.get("checker")
                or current_assignment.get("responsibles")
            ):
                assignments[row_key] = current_assignment

            if applied:
                matched_rows += 1

        _save_reviewer_assignments(assignments)
        _cached_data.setdefault("meta", {})["last_confluence_sync"] = _format_timestamp()
        _mark_release_monitor_state_changed()

        return {
            "matched_rows": matched_rows,
            "source_rows": len(confluence_assignments),
            "year": year,
            "data": _get_cached_payload_copy() or _build_empty_release_monitor_payload(),
        }


def set_release_monitor_reviewer(release_key, reviewer):
    global _cached_data

    release_key = (release_key or "").strip()
    reviewer = (reviewer or "").strip()
    if not release_key:
        raise ValueError("РќРµ СѓРєР°Р·Р°РЅ РєР»СЋС‡ СЂРµР»РёР·Р°")

    if reviewer and reviewer not in OPLOT_VALUES:
        raise ValueError("Р’С‹Р±СЂР°РЅРЅС‹Р№ РїСЂРѕРІРµСЂСЏСЋС‰РёР№ РѕС‚СЃСѓС‚СЃС‚РІСѓРµС‚ РІ СЃРїРёСЃРєРµ РћРџР›РћРў")

    assignments = _load_reviewer_assignments()
    current_assignment = dict(assignments.get(release_key, {}))
    current_assignment["reviewer"] = reviewer
    current_assignment["reviewer_source"] = "manual" if reviewer else ""
    current_assignment["reviewer_date"] = "manual" if reviewer else ""

    if current_assignment.get("reviewer") or current_assignment.get("checker") or current_assignment.get("responsibles"):
        assignments[release_key] = current_assignment
    else:
        assignments.pop(release_key, None)
    _save_reviewer_assignments(assignments)

    with _cache_lock:
        if _cached_data is None:
            disk_payload = _load_snapshot_from_disk()
            if disk_payload is not None:
                _cached_data = _hydrate_release_monitor_payload(disk_payload)
        if _cached_data is not None:
            for item in _cached_data.get("items", []):
                item_key = _get_assignment_key_for_item(item)
                if item_key == release_key:
                    item["psi_owner"] = reviewer
                    item["psi_owner_source"] = current_assignment.get("reviewer_source", "")
                    item["psi_owner_date"] = current_assignment.get("reviewer_date", "")
                    item["psi_checker"] = current_assignment.get("checker", "")
                    item["psi_responsibles"] = list(current_assignment.get("responsibles", []))
                    break
            _mark_release_monitor_state_changed()

    return reviewer


def set_release_monitor_assignment(release_key, reviewer, checker, responsibles=None, reviewer_source=None, zni_reviewer=None):
    global _cached_data

    release_key = (release_key or "").strip()
    reviewer = (reviewer or "").strip()
    checker = (checker or "").strip()
    if not release_key:
        raise ValueError("РќРµ СѓРєР°Р·Р°РЅ РєР»СЋС‡ СЂРµР»РёР·Р°")

    if reviewer and reviewer_source != "manual_text" and reviewer not in OPLOT_VALUES:
        raise ValueError("Р’С‹Р±СЂР°РЅРЅС‹Р№ РґРµР¶СѓСЂРЅС‹Р№ РѕС‚СЃСѓС‚СЃС‚РІСѓРµС‚ РІ СЃРїРёСЃРєРµ РћРџР›РћРў")

    normalized_responsibles = []
    for responsible in (responsibles or []):
        responsible_name = str(responsible or "").strip()
        if not responsible_name:
            continue
        if responsible_name not in OPLOT_VALUES:
            raise ValueError("Р’С‹Р±СЂР°РЅРЅС‹Р№ РѕС‚РІРµС‚СЃС‚РІРµРЅРЅС‹Р№ РѕС‚СЃСѓС‚СЃС‚РІСѓРµС‚ РІ СЃРїРёСЃРєРµ РћРџР›РћРў")
        if responsible_name not in normalized_responsibles:
            normalized_responsibles.append(responsible_name)

    assignments = _load_reviewer_assignments()
    current_assignment = dict(assignments.get(release_key, {}))
    if reviewer_source is None:
        reviewer_source = current_assignment.get("reviewer_source")
    reviewer_source = str(reviewer_source or "").strip()
    if zni_reviewer is None:
        zni_reviewer = current_assignment.get("zni_reviewer", "")
    zni_reviewer = str(zni_reviewer or "").strip()
    if zni_reviewer and zni_reviewer not in OPLOT_VALUES:
        zni_reviewer = _match_oplot_name(zni_reviewer)
    if zni_reviewer and zni_reviewer not in OPLOT_VALUES:
        raise ValueError("Выбранный дежурный для ЗНИ отсутствует в списке ОПЛОТ")

    if reviewer_source == "manual_text" and reviewer:
        resolved_reviewer_source = "manual_text"
        reviewer_date = "manual"
    elif reviewer_source == "duty_schedule" and reviewer:
        resolved_reviewer_source = "duty_schedule"
        reviewer_date = str(current_assignment.get("reviewer_date") or "").strip()
    else:
        resolved_reviewer_source = "manual" if reviewer else ""
        reviewer_date = "manual" if reviewer else ""

    if resolved_reviewer_source != "manual_text":
        zni_reviewer = ""
    elif not zni_reviewer:
        previous_reviewer = str(current_assignment.get("reviewer") or "").strip()
        previous_source = str(current_assignment.get("reviewer_source") or "").strip()
        if previous_source != "manual_text" and previous_reviewer in OPLOT_VALUES:
            zni_reviewer = previous_reviewer

    if reviewer or checker or normalized_responsibles:
        assignments[release_key] = {
            "reviewer": reviewer,
            "reviewer_source": resolved_reviewer_source,
            "reviewer_date": reviewer_date,
            "zni_reviewer": zni_reviewer,
            "checker": checker,
            "responsibles": normalized_responsibles,
        }
    else:
        assignments.pop(release_key, None)
    _save_reviewer_assignments(assignments)

    unassigned_snapshot = None
    with _cache_lock:
        if _cached_data is None:
            disk_payload = _load_snapshot_from_disk()
            if disk_payload is not None:
                _cached_data = _hydrate_release_monitor_payload(disk_payload)
        if _cached_data is not None:
            for item in _cached_data.get("items", []):
                item_key = _get_assignment_key_for_item(item)
                if item_key == release_key:
                    had_responsible = _has_release_responsible(item)
                    item["psi_owner"] = reviewer
                    item["psi_owner_source"] = resolved_reviewer_source
                    item["psi_owner_date"] = reviewer_date
                    item["psi_zni_reviewer"] = zni_reviewer
                    item["psi_checker"] = checker
                    item["psi_responsibles"] = list(normalized_responsibles)
                    is_week_scope = _is_release_assignment_relevant_for_week(item)
                    item["is_current_week_assignment_scope"] = is_week_scope
                    item["is_missing_week_responsible"] = bool(
                        is_week_scope and not normalized_responsibles
                    )
                    if (
                        had_responsible
                        and item["is_missing_week_responsible"]
                    ):
                        unassigned_snapshot = copy.deepcopy(_cached_data)
                    break
            _mark_release_monitor_state_changed()
        else:
            _touch_release_monitor_revision()

    result = {
        "reviewer": reviewer,
        "reviewer_source": resolved_reviewer_source,
        "reviewer_date": reviewer_date,
        "zni_reviewer": zni_reviewer,
        "checker": checker,
        "responsibles": normalized_responsibles,
        "data_revision": _read_data_revision(),
    }
    if unassigned_snapshot is not None:
        unassigned_snapshot.setdefault("meta", {})["data_revision"] = result["data_revision"]
        _schedule_unassigned_confluence_auto_sync(
            unassigned_snapshot,
            refresh_mode="assignment_change",
            force_notify_row_keys=[release_key],
        )
    return result


def assign_release_monitor_responsible_if_expected(
    release_key,
    responsible,
    expected_responsibles=None,
):
    global _cached_data

    release_key = str(release_key or "").strip()
    responsible = str(responsible or "").strip()
    if not release_key:
        raise ValueError("Не указан ключ строки релиза")
    if responsible not in OPLOT_VALUES:
        raise ValueError("Выбранный ответственный отсутствует в списке ОПЛОТ")

    expected = []
    for value in expected_responsibles or []:
        normalized_value = str(value or "").strip()
        if normalized_value and normalized_value not in expected:
            expected.append(normalized_value)

    with _cache_lock:
        _ensure_cached_payload_loaded_locked()
        target_item = next(
            (
                item
                for item in (_cached_data or {}).get("items", [])
                if _get_assignment_key_for_item(item) == release_key
            ),
            None,
        )
        if not target_item:
            raise ValueError("Строка релиза больше не существует в актуальной таблице")

        assignments = _load_reviewer_assignments()
        legacy_key = str(target_item.get("release_key") or "").strip()
        current_assignment = dict(
            assignments.get(release_key)
            or assignments.get(legacy_key)
            or {}
        )
        raw_current_responsibles = (
            current_assignment.get("responsibles")
            if "responsibles" in current_assignment
            else target_item.get("psi_responsibles")
        )
        current_responsibles = []
        for value in raw_current_responsibles or []:
            normalized_value = str(value or "").strip()
            if normalized_value and normalized_value not in current_responsibles:
                current_responsibles.append(normalized_value)

        if responsible in current_responsibles:
            return {
                **current_assignment,
                "responsibles": current_responsibles,
                "data_revision": str(
                    ((_cached_data or {}).get("meta") or {}).get("data_revision")
                    or _read_data_revision()
                    or ""
                ),
                "idempotent": True,
            }

        if current_responsibles != expected:
            raise ReleaseMonitorAssignmentConflict(
                "Ответственный уже назначен другим пользователем",
                assignment={
                    **current_assignment,
                    "responsibles": current_responsibles,
                    "data_revision": str(
                        ((_cached_data or {}).get("meta") or {}).get("data_revision")
                        or _read_data_revision()
                        or ""
                    ),
                },
            )

        next_responsibles = [responsible, *current_responsibles]
        next_assignment = {
            "reviewer": str(current_assignment.get("reviewer") or target_item.get("psi_owner") or "").strip(),
            "reviewer_source": str(
                current_assignment.get("reviewer_source")
                or target_item.get("psi_owner_source")
                or ""
            ).strip(),
            "reviewer_date": str(
                current_assignment.get("reviewer_date")
                or target_item.get("psi_owner_date")
                or ""
            ).strip(),
            "zni_reviewer": str(
                current_assignment.get("zni_reviewer")
                or target_item.get("psi_zni_reviewer")
                or ""
            ).strip(),
            "checker": str(
                current_assignment.get("checker")
                or target_item.get("psi_checker")
                or ""
            ).strip(),
            "responsibles": next_responsibles,
        }
        assignments[release_key] = next_assignment
        _save_reviewer_assignments(assignments)

        target_item["psi_responsibles"] = list(next_responsibles)
        target_item["is_missing_week_responsible"] = False
        revision = _mark_release_monitor_state_changed()
        return {
            **next_assignment,
            "data_revision": revision,
            "idempotent": False,
        }
